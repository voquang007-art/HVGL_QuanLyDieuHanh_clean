# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import uuid
import logging
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from anyio import from_thread

from fastapi import APIRouter, Depends, Form, HTTPException, Request
from sqlalchemy import text
from sqlalchemy.orm import Session
from starlette.responses import JSONResponse, RedirectResponse, StreamingResponse
from starlette.templating import Jinja2Templates

from ..chat.realtime import manager
from ..config import settings
from ..database import get_db
from ..models import (
    RoleCode,
    Roles,
    UnitStatus,
    Units,
    UserRoles,
    UserUnitMemberships,
    Users,
)
from ..security.deps import login_required
from ..work_access import get_work_access_from_session

logger = logging.getLogger("app.leave_schedule")

router = APIRouter(prefix="/leave-schedule", tags=["leave_schedule"])
templates = Jinja2Templates(directory=os.path.join(os.path.dirname(__file__), "..", "templates"))

templates.env.globals["work_access"] = get_work_access_from_session
templates.env.globals["pending_user_approval_count"] = lambda: 0
templates.env.globals["user_unit_names"] = lambda user_id: []
templates.env.globals["leave_schedule_nav"] = lambda: {
    "blocks": [
        {"key": "HANH_CHINH", "label": "Khối Hành chính"},
        {"key": "CHUYEN_MON", "label": "Khối Chuyên môn"},
    ],
    "sub_tabs": [
        {"key": "register", "label": "Đăng ký nghỉ"},
        {"key": "roster", "label": "Phân lịch trực"},
        {"key": "view_calendar", "label": "Xem lịch nghỉ-trực"},
        {"key": "groups", "label": "Tạo nhóm đăng ký nghỉ"},
        {"key": "leave_notice", "label": "Thông báo lịch nghỉ"},
        {"key": "roster_notice", "label": "Thông báo lịch trực"},
    ],
}


def _role_code(name: str) -> str:
    member = getattr(RoleCode, name, None)
    return getattr(member, "value", name)


ROLE_TRUONG_KHOA = _role_code("ROLE_TRUONG_KHOA")
ROLE_PHO_TRUONG_KHOA = _role_code("ROLE_PHO_TRUONG_KHOA")
ROLE_KTV_TRUONG = _role_code("ROLE_KY_THUAT_VIEN_TRUONG")
ROLE_ADMIN = _role_code("ROLE_ADMIN")

MANAGER_ROLE_CODES = {
    ROLE_ADMIN,
    ROLE_TRUONG_KHOA,
    ROLE_PHO_TRUONG_KHOA,
    ROLE_KTV_TRUONG,
}

APPROVER_ROLE_CODES = {
    ROLE_TRUONG_KHOA,
    ROLE_PHO_TRUONG_KHOA,
}

LEAVE_TYPE_META = {
    "F": {"label": "Nghỉ phép năm", "max_year": 12, "max_month": None, "symbol": "F"},
    "P": {"label": "Nghỉ phép tháng", "max_year": None, "max_month": 4, "symbol": "P"},
    "NL": {"label": "Nghỉ lễ, tết", "max_year": 12, "max_month": None, "symbol": "NL"},
    "DB": {"label": "Phép đặc biệt", "max_year": None, "max_month": None, "symbol": "ĐB"},
    "H": {"label": "Đi học", "max_year": None, "max_month": None, "symbol": "H"},
    "CT": {"label": "Đi công tác", "max_year": None, "max_month": None, "symbol": "CT"},
}

LEAVE_TYPES_BYPASS_GROUP_RULES = {"DB", "H", "CT"}

CREATE_LEAVE_REQUESTS_SQL = """
CREATE TABLE IF NOT EXISTS leave_requests (
    id TEXT PRIMARY KEY,
    user_id TEXT NOT NULL,
    user_name TEXT,
    unit_id TEXT,
    unit_name TEXT,
    group_key TEXT NOT NULL,
    group_label TEXT NOT NULL,
    leave_type TEXT NOT NULL,
    symbol TEXT NOT NULL,
    start_date TEXT NOT NULL,
    end_date TEXT NOT NULL,
    day_count REAL NOT NULL DEFAULT 0,
    reason TEXT,
    status TEXT NOT NULL DEFAULT 'PENDING',
    approved_by_id TEXT,
    approved_by_name TEXT,
    approved_role TEXT,
    approved_at TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
)
"""

CREATE_LEAVE_ADJUSTMENTS_SQL = """
CREATE TABLE IF NOT EXISTS leave_year_adjustments (
    id TEXT PRIMARY KEY,
    user_id TEXT NOT NULL,
    leave_year INTEGER NOT NULL,
    extra_days INTEGER NOT NULL DEFAULT 0,
    reason TEXT,
    created_by_id TEXT,
    created_at TEXT NOT NULL
)
"""

CREATE_INDEXES_SQL = [
    "CREATE INDEX IF NOT EXISTS idx_leave_requests_user_status ON leave_requests (user_id, status)",
    "CREATE INDEX IF NOT EXISTS idx_leave_requests_group_dates ON leave_requests (group_key, start_date, end_date)",
    "CREATE INDEX IF NOT EXISTS idx_leave_adjustments_user_year ON leave_year_adjustments (user_id, leave_year)",
]

CREATE_LEAVE_NOTICES_SQL = """
CREATE TABLE IF NOT EXISTS leave_schedule_notices (
    id TEXT PRIMARY KEY,
    user_id TEXT NOT NULL,
    block_code TEXT,
    notice_type TEXT NOT NULL,
    title TEXT NOT NULL,
    message TEXT,
    related_request_id TEXT,
    is_read INTEGER NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL
)
"""

CREATE_NOTICE_INDEXES_SQL = [
    "CREATE INDEX IF NOT EXISTS idx_leave_notice_user_read ON leave_schedule_notices (user_id, is_read, created_at)",
]

CREATE_LEAVE_SWAP_REQUESTS_SQL = """
CREATE TABLE IF NOT EXISTS leave_swap_requests (
    id TEXT PRIMARY KEY,
    block_code TEXT,
    scope_unit_id TEXT,
    scope_subunit_id TEXT,
    user_a_id TEXT NOT NULL,
    user_a_name TEXT,
    request_a_id TEXT NOT NULL,
    user_b_id TEXT NOT NULL,
    user_b_name TEXT,
    request_b_id TEXT NOT NULL,
    swap_note TEXT,
    status TEXT NOT NULL DEFAULT 'PENDING',
    approved_by_id TEXT,
    approved_by_name TEXT,
    approved_at TEXT,
    created_by_id TEXT,
    created_by_name TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
)
"""

CREATE_SWAP_INDEXES_SQL = [
    "CREATE INDEX IF NOT EXISTS idx_leave_swap_scope_status ON leave_swap_requests (block_code, scope_unit_id, scope_subunit_id, status)",
]

CREATE_APPROVED_SCHEDULE_INDEXES_SQL = [
    "CREATE INDEX IF NOT EXISTS idx_leave_requests_scope_status_dates ON leave_requests (status, unit_id, start_date, end_date)",
]

CREATE_ROSTER_PLANS_SQL = """
CREATE TABLE IF NOT EXISTS leave_roster_plans (
    id TEXT PRIMARY KEY,
    block_code TEXT NOT NULL,
    scope_unit_id TEXT,
    scope_unit_name TEXT,
    scope_subunit_id TEXT,
    scope_subunit_name TEXT,
    target_month TEXT NOT NULL,
    planner_user_id TEXT NOT NULL,
    planner_user_name TEXT,
    planner_role_codes TEXT,
    status TEXT NOT NULL DEFAULT 'DRAFT',
    approved_by_id TEXT,
    approved_by_name TEXT,
    approved_at TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
)
"""

CREATE_ROSTER_PLAN_ITEMS_SQL = """
CREATE TABLE IF NOT EXISTS leave_roster_plan_items (
    id TEXT PRIMARY KEY,
    plan_id TEXT NOT NULL,
    user_id TEXT NOT NULL,
    user_name TEXT,
    work_date TEXT NOT NULL,
    shift_code TEXT NOT NULL DEFAULT 'TRUC',
    note TEXT,
    status TEXT NOT NULL DEFAULT 'PLANNED',
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
)
"""

CREATE_ROSTER_ADJUSTMENTS_SQL = """
CREATE TABLE IF NOT EXISTS leave_roster_adjustments (
    id TEXT PRIMARY KEY,
    plan_id TEXT NOT NULL,
    roster_item_id TEXT NOT NULL,
    user_id TEXT NOT NULL,
    user_name TEXT,
    old_work_date TEXT NOT NULL,
    new_work_date TEXT NOT NULL,
    reason TEXT,
    status TEXT NOT NULL DEFAULT 'PENDING',
    requested_by_id TEXT NOT NULL,
    requested_by_name TEXT,
    approved_by_id TEXT,
    approved_by_name TEXT,
    approved_at TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
)
"""

CREATE_ROSTER_INDEXES_SQL = [
    "CREATE INDEX IF NOT EXISTS idx_roster_plans_scope_month ON leave_roster_plans (block_code, scope_unit_id, scope_subunit_id, target_month, status)",
    "CREATE INDEX IF NOT EXISTS idx_roster_items_plan_date ON leave_roster_plan_items (plan_id, work_date, user_id)",
    "CREATE INDEX IF NOT EXISTS idx_roster_adjustments_plan_status ON leave_roster_adjustments (plan_id, status, created_at)",
]

CM_GROUP_DEFS = [
    {"type": "LIMIT_1", "label": "Nhóm hạn mức 1 người nghỉ/ngày", "max_off_per_day": 1},
    {"type": "LIMIT_2", "label": "Nhóm hạn mức 2 người nghỉ/ngày", "max_off_per_day": 2},
    {"type": "LIMIT_3", "label": "Nhóm hạn mức 3 người nghỉ/ngày", "max_off_per_day": 3},
    {"type": "ROTATION", "label": "Nhóm nghỉ luân phiên", "max_off_per_day": 0},
    {"type": "MANAGER_KHOA", "label": "Nhóm Quản lý khoa", "max_off_per_day": 0},
    {"type": "MANAGER_SUBUNIT", "label": "Nhóm Quản lý đơn vị thuộc khoa", "max_off_per_day": 0},
    {"type": "UNLIMITED", "label": "Nhóm nghỉ không hạn mức", "max_off_per_day": 0},
]

HC_GROUP_DEFS = [
    {"type": "LIMIT_1", "label": "Nhóm hạn mức 1 người nghỉ/ngày", "max_off_per_day": 1},
    {"type": "LIMIT_2", "label": "Nhóm hạn mức 2 người nghỉ/ngày", "max_off_per_day": 2},
    {"type": "ROTATION", "label": "Nhóm nghỉ luân phiên", "max_off_per_day": 0},
    {"type": "MANAGER_PHONG", "label": "Nhóm Quản lý Phòng", "max_off_per_day": 0},
    {"type": "MANAGER_TO", "label": "Nhóm Quản lý Tổ", "max_off_per_day": 0},
    {"type": "UNLIMITED", "label": "Nhóm nghỉ không hạn mức", "max_off_per_day": 0},
]

CM_MANAGER_KHOA_ROLES = {
    "ROLE_TRUONG_KHOA",
    "ROLE_PHO_TRUONG_KHOA",
    "ROLE_DIEU_DUONG_TRUONG",
    "ROLE_KY_THUAT_VIEN_TRUONG",
}

CM_MANAGER_SUBUNIT_ROLES = {
    "ROLE_TRUONG_DON_VI",
    "ROLE_PHO_DON_VI",
    "ROLE_DIEU_DUONG_TRUONG_DON_VI",
    "ROLE_KY_THUAT_VIEN_TRUONG_DON_VI",
}

HC_MANAGER_PHONG_ROLES = {
    "ROLE_TRUONG_PHONG",
    "ROLE_PHO_PHONG",
}

HC_MANAGER_TO_ROLES = {
    "ROLE_TO_TRUONG",
    "ROLE_PHO_TO",
}

CREATE_LEAVE_GROUPS_SQL = """
CREATE TABLE IF NOT EXISTS leave_group_configs (
    id TEXT PRIMARY KEY,
    block_code TEXT NOT NULL,
    scope_unit_id TEXT,
    scope_unit_name TEXT,
    scope_subunit_id TEXT,
    scope_subunit_name TEXT,
    group_type TEXT NOT NULL,
    group_name TEXT NOT NULL,
    max_off_per_day INTEGER NOT NULL DEFAULT 0,
    is_active INTEGER NOT NULL DEFAULT 1,
    created_by_id TEXT,
    created_by_name TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
)
"""

CREATE_LEAVE_GROUP_MEMBERS_SQL = """
CREATE TABLE IF NOT EXISTS leave_group_members (
    id TEXT PRIMARY KEY,
    group_id TEXT NOT NULL,
    user_id TEXT NOT NULL,
    user_name TEXT,
    created_at TEXT NOT NULL
)
"""

CREATE_GROUP_INDEXES_SQL = [
    "CREATE INDEX IF NOT EXISTS idx_leave_group_scope ON leave_group_configs (block_code, scope_unit_id, scope_subunit_id, is_active)",
    "CREATE INDEX IF NOT EXISTS idx_leave_group_member_group ON leave_group_members (group_id)",
    "CREATE INDEX IF NOT EXISTS idx_leave_group_member_user ON leave_group_members (user_id)",
]

PRIORITY_PARENT_GROUP_TYPE = "PRIORITY_PARENT"
PRIORITY_CHILD_GROUP_TYPE = "PRIORITY_CHILD"
PRIORITY_STEP_DAYS_FIXED = 1

def _now_vn() -> datetime:
    return datetime.utcnow() + timedelta(hours=7)


def _today_vn() -> date:
    return _now_vn().date()


def _ensure_column(db: Session, table_name: str, column_name: str, column_sql: str) -> None:
    cols = db.execute(text(f"PRAGMA table_info({table_name})")).fetchall()
    col_names = {str(r[1]) for r in cols if r and len(r) > 1}
    if column_name not in col_names:
        db.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_sql}"))


def _ensure_leave_group_extra_columns(db: Session) -> None:
    _ensure_column(db, "leave_group_configs", "parent_group_id", "TEXT")
    _ensure_column(db, "leave_group_configs", "priority_base_day", "INTEGER NOT NULL DEFAULT 0")
    _ensure_column(db, "leave_group_configs", "priority_order", "INTEGER NOT NULL DEFAULT 0")
    _ensure_column(db, "leave_group_configs", "custom_deadline_day", "INTEGER")
    _ensure_column(db, "leave_group_configs", "custom_max_off_per_day", "INTEGER")
    _ensure_column(db, "leave_group_members", "member_order", "INTEGER NOT NULL DEFAULT 0")

def _ensure_tables(db: Session) -> None:
    db.execute(text(CREATE_LEAVE_REQUESTS_SQL))
    db.execute(text(CREATE_LEAVE_ADJUSTMENTS_SQL))
    db.execute(text(CREATE_LEAVE_NOTICES_SQL))
    db.execute(text(CREATE_LEAVE_GROUPS_SQL))
    db.execute(text(CREATE_LEAVE_GROUP_MEMBERS_SQL))
    db.execute(text(CREATE_LEAVE_SWAP_REQUESTS_SQL))
    db.execute(text(CREATE_ROSTER_PLANS_SQL))
    db.execute(text(CREATE_ROSTER_PLAN_ITEMS_SQL))
    db.execute(text(CREATE_ROSTER_ADJUSTMENTS_SQL))
    _ensure_leave_group_extra_columns(db)

    _ensure_column(db, "leave_requests", "reject_reason", "TEXT")
    _ensure_column(db, "leave_requests", "block_code", "TEXT")

    for sql in CREATE_INDEXES_SQL:
        db.execute(text(sql))
    for sql in CREATE_NOTICE_INDEXES_SQL:
        db.execute(text(sql))
    for sql in CREATE_GROUP_INDEXES_SQL:
        db.execute(text(sql))
    for sql in CREATE_SWAP_INDEXES_SQL:
        db.execute(text(sql))
    for sql in CREATE_APPROVED_SCHEDULE_INDEXES_SQL:
        db.execute(text(sql))
    for sql in CREATE_ROSTER_INDEXES_SQL:
        db.execute(text(sql))

    db.commit()


def _parse_date(value: str, field_name: str) -> date:
    try:
        return datetime.strptime((value or "").strip(), "%Y-%m-%d").date()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"{field_name} không hợp lệ.") from exc


def _date_to_str(value: date | None) -> str | None:
    return value.isoformat() if value else None


def _dt_to_str(value: datetime | None) -> str | None:
    return value.strftime("%Y-%m-%d %H:%M:%S") if value else None

async def _notify_leave_schedule_users(user_ids: list[str], payload: dict[str, Any]) -> None:
    clean_ids: list[str] = []
    for raw_user_id in user_ids or []:
        uid = str(raw_user_id or "").strip()
        if uid and uid not in clean_ids:
            clean_ids.append(uid)

    if not clean_ids:
        return

    await manager.notify_users_json(clean_ids, payload)


def _manager_user_ids_for_leave_scope(db: Session, unit_id: str | None, block_code: str | None = None) -> list[str]:
    if not unit_id:
        return []

    manager_roles = {
        "ROLE_ADMIN",
        "ROLE_LANH_DAO",
        "ROLE_TRUONG_PHONG",
        "ROLE_PHO_PHONG",
        "ROLE_TO_TRUONG",
        "ROLE_PHO_TO",
        "ROLE_TRUONG_KHOA",
        "ROLE_PHO_TRUONG_KHOA",
        "ROLE_DIEU_DUONG_TRUONG",
        "ROLE_KY_THUAT_VIEN_TRUONG",
        "ROLE_TRUONG_DON_VI",
        "ROLE_PHO_DON_VI",
        "ROLE_DIEU_DUONG_TRUONG_DON_VI",
        "ROLE_KY_THUAT_VIEN_TRUONG_DON_VI",
        "ROLE_TRUONG_NHOM",
        "ROLE_PHO_NHOM",
    }

    rows = (
        db.query(Users.id)
        .join(UserUnitMemberships, UserUnitMemberships.user_id == Users.id)
        .join(UserRoles, UserRoles.user_id == Users.id)
        .join(Roles, Roles.id == UserRoles.role_id)
        .filter(UserUnitMemberships.unit_id == unit_id)
        .filter(Roles.code.in_(list(manager_roles)))
        .distinct()
        .all()
    )
    return [str(user_id) for (user_id,) in rows if user_id]


def _leave_schedule_payload(event_type: str, **kwargs) -> dict[str, Any]:
    payload = {
        "module": "leave_schedule",
        "type": event_type,
        "sent_at": _dt_to_str(_now_vn()),
    }
    payload.update(kwargs)
    return payload


def _month_cutoff(any_day: date) -> int:
    if any_day.month == 2:
        return 23
    if any_day.month in {4, 6, 9, 11}:
        return 25
    return 26


def _iter_workdays(start_date: date, end_date: date):
    current = start_date
    while current <= end_date:
        if current.weekday() != 6:
            yield current
        current += timedelta(days=1)


def _count_workdays(start_date: date, end_date: date) -> int:
    return sum(1 for _ in _iter_workdays(start_date, end_date))

def _next_workday(value: date) -> date:
    return value + timedelta(days=1)


def _parse_selected_dates_csv(selected_dates: str) -> list[date]:
    raw_items = [x.strip() for x in (selected_dates or "").split(",") if x.strip()]
    out: list[date] = []
    seen: set[str] = set()
    for item in raw_items:
        d = _parse_date(item, "Ngày nghỉ")
        key = d.isoformat()
        if key not in seen:
            seen.add(key)
            out.append(d)
    out.sort()
    return out


def _leave_symbol(leave_type: str) -> str:
    meta = LEAVE_TYPE_META.get((leave_type or "").strip().upper(), {})
    return str(meta.get("symbol") or (leave_type or "").strip().upper())


def _validate_quota_for_selected_dates(db: Session, user: Users, leave_type: str, selected_dates: list[date]) -> None:
    if not selected_dates:
        raise HTTPException(status_code=400, detail="Chưa có ngày nghỉ hợp lệ được chọn.")

    statuses = ("PENDING", "APPROVED")

    if leave_type == "F":
        leave_year = selected_dates[0].year
        total_extra = _sum_extra_days(db, user.id, leave_year)
        used_days = _sum_days_by_year(db, user.id, leave_type, leave_year, statuses)
        requested_days = len(selected_dates)
        max_days = int(LEAVE_TYPE_META["F"]["max_year"] or 0) + total_extra
        if used_days + requested_days > max_days:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Nghỉ phép năm vượt giới hạn. Đã dùng/đang chờ: {used_days:g} ngày; "
                    f"được phép: {max_days:g} ngày."
                ),
            )

    if leave_type == "NL":
        leave_year = selected_dates[0].year
        used_days = _sum_days_by_year(db, user.id, leave_type, leave_year, statuses)
        requested_days = len(selected_dates)
        max_days = int(LEAVE_TYPE_META["NL"]["max_year"] or 0)
        if used_days + requested_days > max_days:
            raise HTTPException(
                status_code=400,
                detail=f"Nghỉ lễ, tết vượt mức cấu hình {max_days:g} ngày/năm.",
            )

    if leave_type == "P":
        month_map: dict[str, int] = {}
        for d in selected_dates:
            ym = d.strftime("%Y-%m")
            month_map[ym] = month_map.get(ym, 0) + 1

        for ym, add_days in month_map.items():
            used_days = _sum_days_by_month(db, user.id, leave_type, ym, statuses)
            max_days = int(LEAVE_TYPE_META["P"]["max_month"] or 0)
            if used_days + add_days > max_days:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"Nghỉ phép tháng vượt giới hạn 4 ngày công/tháng. "
                        f"Tháng {ym} đã dùng/đang chờ: {used_days:g} ngày."
                    ),
                )


def _find_group_conflict_for_selected_dates(
    db: Session,
    user: Users,
    group_key: str,
    selected_dates: list[date],
    max_off_per_day: int = 1,
    leave_type: str | None = None,
) -> dict[str, Any] | None:
    if not selected_dates:
        return None

    start_date = min(selected_dates)
    end_date = max(selected_dates)

    rows = db.execute(
        text(
            """
            SELECT *
            FROM leave_requests
            WHERE group_key = :group_key
              AND status IN ('PENDING', 'APPROVED')
              AND NOT (end_date < :start_date OR start_date > :end_date)
            ORDER BY start_date ASC, created_at ASC
            """
        ),
        {
            "group_key": group_key,
            "start_date": _date_to_str(start_date),
            "end_date": _date_to_str(end_date),
        },
    ).fetchall()

    wanted_dates = {d.isoformat() for d in selected_dates}
    existing_count_by_date: dict[str, int] = {}

    for row in rows:
        item = _row_to_dict(row)
        if str(item.get("user_id") or "") == str(user.id):
            for ds in wanted_dates:
                if item.get("start_date") == ds and item.get("end_date") == ds:
                    return {
                        "message": f"Ngày {ds} đã có đăng ký nghỉ trước đó của chính người dùng.",
                        "row": item,
                    }

        row_start = _parse_date(item["start_date"], "Ngày bắt đầu")
        row_end = _parse_date(item["end_date"], "Ngày kết thúc")
        for d in _iter_workdays(row_start, row_end):
            ds = d.isoformat()
            if ds in wanted_dates:
                existing_count_by_date[ds] = existing_count_by_date.get(ds, 0) + 1

    if leave_type in {"F", "P"} and str(group_key).startswith("GROUP::"):
        group_id = str(group_key).split("GROUP::", 1)[1].strip()
        group_row = db.execute(
            text("SELECT id, group_type FROM leave_group_configs WHERE id = :id"),
            {"id": group_id},
        ).fetchone()

        if group_row:
            group_item = _row_to_dict(group_row)
            group_type = str(group_item.get("group_type") or "")
            if group_type in {"MANAGER_KHOA", "MANAGER_SUBUNIT"}:
                members_with_roles = _load_group_members_with_roles(db, group_id)
                rules = _manager_group_runtime_rules(
                    group_type=group_type,
                    members=members_with_roles,
                )
                bucket_roles = set(rules.get("bucket_roles") or set())
                bucket_limit = int(rules.get("bucket_limit") or 0)
                current_user_roles = _get_role_codes(db, str(user.id))

                if bucket_roles and bucket_limit >= 0 and (current_user_roles & bucket_roles):
                    bucket_user_ids = {
                        str(m.get("user_id") or "")
                        for m in members_with_roles
                        if set(m.get("role_codes") or []) & bucket_roles
                    }
                    existing_bucket_count_by_date: dict[str, int] = {}

                    for row in rows:
                        item = _row_to_dict(row)
                        if str(item.get("user_id") or "") not in bucket_user_ids:
                            continue

                        row_start = _parse_date(item["start_date"], "Ngày bắt đầu")
                        row_end = _parse_date(item["end_date"], "Ngày kết thúc")
                        for d in _iter_workdays(row_start, row_end):
                            ds = d.isoformat()
                            if ds in wanted_dates:
                                existing_bucket_count_by_date[ds] = existing_bucket_count_by_date.get(ds, 0) + 1

                    for ds in sorted(wanted_dates):
                        if existing_bucket_count_by_date.get(ds, 0) >= bucket_limit and bucket_limit >= 0:
                            if group_type == "MANAGER_KHOA":
                                msg = f"Ngày {ds} đã đạt hạn mức riêng của nhóm quản lý khoa đối với vị trí Trưởng/Phó khoa."
                            else:
                                msg = f"Ngày {ds} đã đạt hạn mức riêng của nhóm quản lý đơn vị đối với vị trí Trưởng/Phó đơn vị."
                            return {
                                "message": msg,
                                "row": None,
                            }

    if max_off_per_day and max_off_per_day > 0:
        for ds in sorted(wanted_dates):
            if existing_count_by_date.get(ds, 0) >= int(max_off_per_day):
                return {
                    "message": f"Ngày {ds} đã đạt hạn mức {int(max_off_per_day)} người nghỉ trong nhóm.",
                    "row": None,
                }

    return None


def _build_calendar_days(month_str: str) -> list[dict[str, Any]]:
    year, month = [int(x) for x in month_str.split("-")]
    first_day = date(year, month, 1)
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)
    last_day = next_month - timedelta(days=1)

    out: list[dict[str, Any]] = []
    current = first_day
    while current <= last_day:
        out.append(
            {
                "date": current.isoformat(),
                "day": current.day,
                "weekday": current.weekday(),
                "is_sunday": current.weekday() == 6,
            }
        )
        current += timedelta(days=1)
    return out


def _load_group_rule_blocked_dates(db: Session, user: Users, month_str: str) -> list[str]:
    year, month = [int(x) for x in month_str.split("-")]
    first_day = date(year, month, 1)
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)
    last_day = next_month - timedelta(days=1)

    blocked: set[str] = set()
    group_key, group_label, unit, max_off_per_day, deadline_day = _resolve_effective_leave_group(db, user, first_day)

    current = first_day
    while current <= last_day:
        if current.weekday() == 6:
            current += timedelta(days=1)
            continue

        if max_off_per_day and max_off_per_day > 0:
            conflict = _find_group_conflict_for_selected_dates(
                db,
                user,
                group_key,
                [current],
                max_off_per_day=max_off_per_day,
            )
            if conflict:
                blocked.add(current.isoformat())

        current += timedelta(days=1)

    return sorted(blocked)

def _can_show_register_button(
    db: Session,
    user: Users,
    *,
    target_month_str: str,
) -> tuple[bool, str]:
    year, month = [int(x) for x in target_month_str.split("-")]
    target_month_base = date(year, month, 1)
    today = _today_vn()
    current_month_base = date(today.year, today.month, 1)

    if current_month_base.month == 12:
        next_month_base = date(current_month_base.year + 1, 1, 1)
    else:
        next_month_base = date(current_month_base.year, current_month_base.month + 1, 1)

    if target_month_base != next_month_base:
        return False, "Chỉ mở đăng ký cho tháng kế tiếp."

    cutoff_day = _month_cutoff(current_month_base)

    group_key, group_label, unit, max_off_per_day, deadline_day = _resolve_effective_leave_group(
        db,
        user,
        target_month_base,
    )

    # Nhóm thường: trong tháng hiện tại, được đăng ký cho tháng kế tiếp đến hết ngày khóa
    if deadline_day is None:
        return today.day <= cutoff_day, (
            f"Đã vào giai đoạn khóa đăng ký tháng sau (sau ngày {cutoff_day} của tháng hiện tại)."
        )

    # Nhóm luân phiên: chỉ mở đúng ngày ưu tiên trong tháng hiện tại,
    # đồng thời ngày đó không được vượt mốc khóa của tháng hiện tại
    if int(deadline_day) > int(cutoff_day):
        return False, (
            f"Ngày đăng ký của nhóm con ({deadline_day}) rơi sau mốc khóa tháng hiện tại ({cutoff_day}), "
            f"nên tháng này không mở đăng ký cho tháng sau."
        )

    return today.day == int(deadline_day), (
        f"Nhóm của bạn chỉ được đăng ký tháng sau đúng ngày {deadline_day} của tháng hiện tại."
    )

def _row_to_dict(row: Any) -> dict[str, Any]:
    if hasattr(row, "_mapping"):
        return dict(row._mapping)
    return dict(row)


def _enum_str(value: Any) -> str:
    return str(getattr(value, "value", value) or "").upper().strip()


def _unit_category_value(unit: Units | None) -> str:
    if not unit:
        return ""
    return _enum_str(getattr(unit, "unit_category", None))


def _unit_block_value(unit: Units | None) -> str:
    if not unit:
        return ""
    return _enum_str(getattr(unit, "block_code", None))


def _active_membership_units(db: Session, user_id: str) -> list[Units]:
    rows = (
        db.query(Units)
        .join(UserUnitMemberships, UserUnitMemberships.unit_id == Units.id)
        .filter(UserUnitMemberships.user_id == user_id)
        .filter(Units.trang_thai == UnitStatus.ACTIVE)
        .order_by(Units.cap_do.asc(), Units.order_index.asc(), Units.ten_don_vi.asc())
        .all()
    )
    dedup: dict[str, Units] = {}
    for u in rows:
        if u and getattr(u, "id", None):
            dedup[u.id] = u
    return list(dedup.values())


def _primary_membership(db: Session, user_id: str) -> UserUnitMemberships | None:
    q = db.query(UserUnitMemberships).filter(UserUnitMemberships.user_id == user_id)
    if hasattr(UserUnitMemberships, "is_primary"):
        q = q.order_by(UserUnitMemberships.is_primary.desc())
    if hasattr(UserUnitMemberships, "id"):
        q = q.order_by(UserUnitMemberships.id.desc())
    return q.first()


def _primary_unit(db: Session, user_id: str) -> Units | None:
    membership = _primary_membership(db, user_id)
    if not membership:
        return None
    return db.get(Units, getattr(membership, "unit_id", None))


def _resolve_unit_scope(db: Session, user_id: str) -> tuple[Units | None, Units | None]:
    primary_unit = _primary_unit(db, user_id)
    if not primary_unit:
        return None, None

    if _unit_category_value(primary_unit) == "SUBUNIT":
        parent_unit = db.get(Units, getattr(primary_unit, "parent_id", None))
        return parent_unit, primary_unit

    return primary_unit, None


def _format_unit_display(unit: Units | None, subunit: Units | None = None) -> str:
    if unit is None and subunit is None:
        return "-"
    if unit is not None and subunit is not None:
        return f"{getattr(unit, 'ten_don_vi', '') or '-'} / {getattr(subunit, 'ten_don_vi', '') or '-'}"
    return getattr(unit or subunit, "ten_don_vi", None) or "-"


def _get_role_codes(db: Session, user_id: str) -> set[str]:
    rows = (
        db.query(Roles.code)
        .join(UserRoles, UserRoles.role_id == Roles.id)
        .filter(UserRoles.user_id == user_id)
        .all()
    )
    return {str(getattr(code, "value", code)).upper() for (code,) in rows if code is not None}


def _is_admin_or_leader(role_codes: set[str]) -> bool:
    return ("ROLE_ADMIN" in role_codes) or ("ROLE_LANH_DAO" in role_codes)


def _is_truong_phong(role_codes: set[str]) -> bool:
    return "ROLE_TRUONG_PHONG" in role_codes


def _is_pho_phong(role_codes: set[str]) -> bool:
    return "ROLE_PHO_PHONG" in role_codes


def _is_truong_khoa(role_codes: set[str]) -> bool:
    return "ROLE_TRUONG_KHOA" in role_codes


def _is_pho_truong_khoa(role_codes: set[str]) -> bool:
    return "ROLE_PHO_TRUONG_KHOA" in role_codes


def _is_manager_like(role_codes: set[str]) -> bool:
    return bool(
        _is_admin_or_leader(role_codes)
        or _is_truong_phong(role_codes)
        or _is_pho_phong(role_codes)
        or _is_truong_khoa(role_codes)
        or _is_pho_truong_khoa(role_codes)
    )


def _is_hc_manager(role_codes: set[str]) -> bool:
    return bool(role_codes & {
        "ROLE_TRUONG_PHONG",
        "ROLE_PHO_PHONG",
        "ROLE_TO_TRUONG",
        "ROLE_PHO_TO",
    })


def _is_cm_manager(role_codes: set[str]) -> bool:
    return bool(role_codes & {
        "ROLE_TRUONG_KHOA",
        "ROLE_PHO_TRUONG_KHOA",
        "ROLE_DIEU_DUONG_TRUONG",
        "ROLE_KY_THUAT_VIEN_TRUONG",
        "ROLE_TRUONG_DON_VI",
        "ROLE_PHO_DON_VI",
        "ROLE_DIEU_DUONG_TRUONG_DON_VI",
        "ROLE_KY_THUAT_VIEN_TRUONG_DON_VI",
    })


def _can_manage_block(role_codes: set[str], block_code: str) -> bool:
    if _is_admin_or_leader(role_codes):
        return True
    if block_code == "HANH_CHINH":
        return _is_hc_manager(role_codes)
    if block_code == "CHUYEN_MON":
        return _is_cm_manager(role_codes)
    return False


def _can_approve_block(role_codes: set[str], block_code: str) -> bool:
    if _is_admin_or_leader(role_codes):
        return True
    if block_code == "HANH_CHINH":
        return bool(role_codes & {
            "ROLE_TRUONG_PHONG",
            "ROLE_PHO_PHONG",
            "ROLE_TO_TRUONG",
            "ROLE_PHO_TO",
        })
    if block_code == "CHUYEN_MON":
        return bool(role_codes & {
            "ROLE_TRUONG_KHOA",
            "ROLE_PHO_TRUONG_KHOA",
            "ROLE_DIEU_DUONG_TRUONG",
            "ROLE_KY_THUAT_VIEN_TRUONG",
            "ROLE_TRUONG_DON_VI",
            "ROLE_PHO_DON_VI",
            "ROLE_DIEU_DUONG_TRUONG_DON_VI",
            "ROLE_KY_THUAT_VIEN_TRUONG_DON_VI",
            "ROLE_TRUONG_NHOM",
            "ROLE_PHO_NHOM",
        })
    return False

def _default_block_for_user(db: Session, user: Users) -> str:
    scope_unit, scope_subunit = _resolve_unit_scope(db, user.id)
    return _unit_block_value(scope_subunit or scope_unit)


def _block_label(block_code: str) -> str:
    if block_code == "CHUYEN_MON":
        return "Khối chuyên môn"
    if block_code == "HANH_CHINH":
        return "Khối hành chính"
    return "Chưa xác định"

def _month_base_from_str(month_str: str) -> date:
    try:
        year, month = [int(x) for x in str(month_str).split("-")]
        return date(year, month, 1)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Tháng không hợp lệ.") from exc


def _week_of_month(value: date) -> int:
    return ((value.day - 1) // 7) + 1


def _is_role_doctor(role_code: str) -> bool:
    code = str(role_code or "").upper()
    return ("BAC_SI" in code) or ("BS" == code)


def _is_role_bgd(role_code: str) -> bool:
    code = str(role_code or "").upper()
    return any(x in code for x in ["BGD", "GIAM_DOC", "PHO_GIAM_DOC"])


def _is_role_nursing_or_tech(role_code: str) -> bool:
    code = str(role_code or "").upper()
    return any(x in code for x in ["DIEU_DUONG", "KY_THUAT_VIEN", "KTV", "DUOC_SI", "HO_LY"])


def _csv_role_codes(role_codes: set[str]) -> str:
    return ",".join(sorted([str(x).upper() for x in role_codes if str(x).strip()]))

def _scope_payload(db: Session, user: Users) -> dict[str, Any]:
    role_codes = _get_role_codes(db, user.id)
    scope_unit, scope_subunit = _resolve_unit_scope(db, user.id)
    block_code = _default_block_for_user(db, user)

    return {
        "user_id": str(user.id),
        "user_name": user.full_name or user.username or "",
        "role_codes": sorted(role_codes),
        "is_admin_or_leader": _is_admin_or_leader(role_codes),
        "is_manager_like": _is_manager_like(role_codes),
        "block_code": block_code,
        "block_label": _block_label(block_code),
        "scope_unit": scope_unit,
        "scope_subunit": scope_subunit,
        "scope_unit_id": getattr(scope_unit, "id", None),
        "scope_subunit_id": getattr(scope_subunit, "id", None),
        "scope_unit_name": getattr(scope_unit, "ten_don_vi", None),
        "scope_subunit_name": getattr(scope_subunit, "ten_don_vi", None),
        "scope_display": _format_unit_display(scope_unit, scope_subunit),
    }


def _unit_member_user_ids(db: Session, unit_ids: list[str]) -> list[str]:
    if not unit_ids:
        return []
    rows = (
        db.query(UserUnitMemberships.user_id)
        .filter(UserUnitMemberships.unit_id.in_(unit_ids))
        .distinct()
        .all()
    )
    return [str(user_id) for (user_id,) in rows if user_id]


def _leave_notify_payload(event_type: str, **kwargs) -> dict[str, Any]:
    payload = {
        "module": "leave_schedule",
        "type": event_type,
        "sent_at": _dt_to_str(_now_vn()),
    }
    payload.update(kwargs)
    return payload


def _emit_leave_schedule_realtime_sync(user_ids: list[str], payload: dict[str, Any]) -> None:
    """
    Phát realtime trong các route đang khai báo def.
    Không đổi route sang async def để tránh ảnh hưởng SQLAlchemy session sync hiện tại.
    """
    clean_ids: list[str] = []
    for raw_user_id in user_ids or []:
        uid = str(raw_user_id or "").strip()
        if uid and uid not in clean_ids:
            clean_ids.append(uid)

    if not clean_ids:
        return

    try:
        from_thread.run(manager.notify_users_json, clean_ids, payload)
    except Exception as exc:
        logger.exception("[leave_schedule] Lỗi phát realtime leave_schedule: %s", exc)


async def _notify_leave_users(db: Session, user_ids: list[str], payload: dict[str, Any]) -> None:
    target_ids = [str(x) for x in user_ids if str(x).strip()]
    if not target_ids:
        return
    await manager.notify_users_json(target_ids, payload)


async def _notify_leave_scope(
    db: Session,
    *,
    scope_unit_id: str | None = None,
    scope_subunit_id: str | None = None,
    payload: dict[str, Any],
) -> None:
    unit_ids = [x for x in [scope_unit_id, scope_subunit_id] if x]
    user_ids = _unit_member_user_ids(db, unit_ids)
    await _notify_leave_users(db, user_ids, payload)


def _load_unread_notice_count(db: Session, user_id: str) -> int:
    row = db.execute(
        text(
            """
            SELECT COUNT(*) AS total_count
            FROM leave_schedule_notices
            WHERE user_id = :user_id AND is_read = 0
            """
        ),
        {"user_id": str(user_id)},
    ).fetchone()
    if not row:
        return 0
    return int(_row_to_dict(row).get("total_count") or 0)



def _load_unread_notice_count_by_types(
    db: Session,
    user_id: str,
    notice_types: list[str],
) -> int:
    clean_types = [str(x).strip() for x in notice_types if str(x).strip()]
    if not clean_types:
        return 0

    placeholders = ",".join([f":notice_type_{idx}" for idx in range(len(clean_types))])
    params: dict[str, Any] = {"user_id": str(user_id)}
    for idx, notice_type in enumerate(clean_types):
        params[f"notice_type_{idx}"] = notice_type

    row = db.execute(
        text(
            f"""
            SELECT COUNT(*) AS total_count
            FROM leave_schedule_notices
            WHERE user_id = :user_id
              AND is_read = 0
              AND notice_type IN ({placeholders})
            """
        ),
        params,
    ).fetchone()

    if not row:
        return 0
    return int(_row_to_dict(row).get("total_count") or 0)


def _load_leave_notice_unread_count(db: Session, user_id: str) -> int:
    return _load_unread_notice_count_by_types(
        db,
        user_id,
        [
            "LEAVE_REQUEST_APPROVED",
            "LEAVE_REQUEST_REJECTED",
            "LEAVE_SWAP_APPROVED",
        ],
    )


def _load_roster_notice_unread_count(db: Session, user_id: str) -> int:
    return _load_unread_notice_count_by_types(
        db,
        user_id,
        [
            "ROSTER_APPROVED",
            "ROSTER_ADJUST_APPROVED",
        ],
    )

def _load_roster_notices_for_user(
    db: Session,
    user_id: str,
    block_code: str = "",
    month_str: str = "",
    week_no: str = "",
    limit: int = 100,
) -> list[dict[str, Any]]:
    params: dict[str, Any] = {
        "user_id": str(user_id),
        "notice_type_1": "ROSTER_APPROVED",
        "notice_type_2": "ROSTER_ADJUST_APPROVED",
        "limit": int(limit),
    }

    sql = """
        SELECT *
        FROM leave_schedule_notices
        WHERE user_id = :user_id
          AND notice_type IN (:notice_type_1, :notice_type_2)
    """

    if block_code:
        sql += " AND COALESCE(block_code, '') = :block_code "
        params["block_code"] = block_code

    if month_str:
        sql += " AND substr(COALESCE(created_at, ''), 1, 7) = :month_str "
        params["month_str"] = month_str

    sql += " ORDER BY is_read ASC, created_at DESC LIMIT :limit "

    rows = db.execute(text(sql), params).fetchall()
    items = [_row_to_dict(r) for r in rows]
    notify_user_ids: list[str] = []
    
    if week_no and str(week_no).isdigit():
        wanted_week = int(week_no)
        filtered: list[dict[str, Any]] = []
        for item in items:
            created_at_str = str(item.get("created_at") or "")
            try:
                created_dt = datetime.strptime(created_at_str, "%Y-%m-%d %H:%M:%S").date()
            except Exception:
                continue
            if _week_of_month(created_dt) == wanted_week:
                filtered.append(item)
        items = filtered

    return items


def _load_approved_leave_rows_for_user(
    db: Session,
    user_id: str,
    target_month: str,
) -> list[dict[str, Any]]:
    month_base = _month_base_from_str(target_month)
    if month_base.month == 12:
        next_month = date(month_base.year + 1, 1, 1)
    else:
        next_month = date(month_base.year, month_base.month + 1, 1)
    month_end = next_month - timedelta(days=1)

    rows = db.execute(
        text(
            """
            SELECT *
            FROM leave_requests
            WHERE user_id = :user_id
              AND status = 'APPROVED'
              AND NOT (end_date < :month_start OR start_date > :month_end)
            ORDER BY start_date ASC, created_at ASC
            """
        ),
        {
            "user_id": str(user_id),
            "month_start": month_base.isoformat(),
            "month_end": month_end.isoformat(),
        },
    ).fetchall()
    return [_row_to_dict(r) for r in rows]


def _roster_forbidden_dates_for_user(
    db: Session,
    user_id: str,
    target_month: str,
) -> dict[str, str]:
    leave_rows = _load_approved_leave_rows_for_user(db, user_id, target_month)
    out: dict[str, str] = {}
    for item in leave_rows:
        start_dt = _parse_date(str(item.get("start_date") or ""), "Ngày nghỉ")
        end_dt = _parse_date(str(item.get("end_date") or ""), "Ngày nghỉ")
        current = start_dt
        while current <= end_dt:
            out[current.isoformat()] = "Trùng ngày nghỉ đã được phê duyệt."
            prev_day = current - timedelta(days=1)
            if prev_day.isoformat() not in out:
                out[prev_day.isoformat()] = "Liền kề trước ngày nghỉ đã được phê duyệt."
            current += timedelta(days=1)
    return out


def _validate_roster_work_date(
    db: Session,
    *,
    user_id: str,
    target_month: str,
    work_date_str: str,
) -> None:
    work_date = _parse_date(work_date_str, "Ngày trực")
    forbidden_map = _roster_forbidden_dates_for_user(db, user_id, target_month)
    reason = forbidden_map.get(work_date.isoformat())
    if reason:
        raise HTTPException(status_code=400, detail=f"Ngày {work_date.isoformat()} không hợp lệ: {reason}")

def _create_leave_notice(
    db: Session,
    *,
    user_id: str,
    block_code: str,
    notice_type: str,
    title: str,
    message: str,
    related_request_id: str | None = None,
) -> None:
    db.execute(
        text(
            """
            INSERT INTO leave_schedule_notices (
                id, user_id, block_code, notice_type, title, message, related_request_id, is_read, created_at
            ) VALUES (
                :id, :user_id, :block_code, :notice_type, :title, :message, :related_request_id, 0, :created_at
            )
            """
        ),
        {
            "id": str(uuid.uuid4()),
            "user_id": str(user_id),
            "block_code": block_code or "",
            "notice_type": notice_type,
            "title": title,
            "message": message,
            "related_request_id": related_request_id,
            "created_at": _dt_to_str(_now_vn()),
        },
    )

def _group_defs_for_block(block_code: str) -> list[dict[str, Any]]:
    return CM_GROUP_DEFS if block_code == "CHUYEN_MON" else HC_GROUP_DEFS


def _group_def_map(block_code: str) -> dict[str, dict[str, Any]]:
    return {item["type"]: item for item in _group_defs_for_block(block_code)}

def _all_scope_members_with_roles(
    db: Session,
    *,
    scope_unit_id: str | None,
    scope_subunit_id: str | None,
) -> list[dict[str, Any]]:
    unit_ids = [x for x in [scope_unit_id, scope_subunit_id] if x]
    if not unit_ids:
        return []

    rows = (
        db.query(Users, UserUnitMemberships.unit_id)
        .join(UserUnitMemberships, UserUnitMemberships.user_id == Users.id)
        .filter(UserUnitMemberships.unit_id.in_(unit_ids))
        .all()
    )

    seen: dict[str, dict[str, Any]] = {}
    for user_obj, unit_id in rows:
        uid = str(user_obj.id)
        if uid not in seen:
            seen[uid] = {
                "id": uid,
                "name": user_obj.full_name or user_obj.username or uid,
                "unit_id": str(unit_id or ""),
                "role_codes": sorted(_get_role_codes(db, uid)),
            }
    return list(seen.values())


def _belongs_to_any_subunit(db: Session, user_id: str, parent_unit_id: str) -> bool:
    rows = (
        db.query(Units.id)
        .join(UserUnitMemberships, UserUnitMemberships.unit_id == Units.id)
        .filter(UserUnitMemberships.user_id == user_id)
        .filter(Units.parent_id == parent_unit_id)
        .filter(Units.trang_thai == UnitStatus.ACTIVE)
        .all()
    )
    return bool(rows)


def _resolve_roster_candidates(
    db: Session,
    *,
    user: Users,
    scope_info: dict[str, Any],
    block_code: str,
) -> list[dict[str, Any]]:
    role_codes = _get_role_codes(db, user.id)
    scope_unit_id = scope_info.get("scope_unit_id")
    scope_subunit_id = scope_info.get("scope_subunit_id")
    members = _all_scope_members_with_roles(
        db,
        scope_unit_id=scope_unit_id,
        scope_subunit_id=scope_subunit_id,
    )

    bgd_role_codes = {
        "ROLE_LANH_DAO",
        "ROLE_TONG_GIAM_DOC",
        "ROLE_PHO_TONG_GIAM_DOC_THUONG_TRUC",
        "ROLE_PHO_TONG_GIAM_DOC",
        "ROLE_GIAM_DOC",
        "ROLE_PHO_GIAM_DOC_TRUC",
        "ROLE_PHO_GIAM_DOC",
        "ROLE_BGD",
    }

    khoa_leader_role_codes = {
        "ROLE_TRUONG_KHOA",
        "ROLE_PHO_TRUONG_KHOA",
    }

    khoa_nursing_tech_leader_role_codes = {
        "ROLE_DIEU_DUONG_TRUONG",
        "ROLE_KY_THUAT_VIEN_TRUONG",
    }

    subunit_leader_role_codes = {
        "ROLE_TRUONG_DON_VI",
        "ROLE_PHO_DON_VI",
    }

    subunit_nursing_tech_leader_role_codes = {
        "ROLE_DIEU_DUONG_TRUONG_DON_VI",
        "ROLE_KY_THUAT_VIEN_TRUONG_DON_VI",
    }

    out: list[dict[str, Any]] = []

    def _append_global_bgd_users() -> None:
        bgd_users = db.query(Users).all()
        for bgd_user in bgd_users:
            uid = str(getattr(bgd_user, "id", "") or "")
            if not uid:
                continue
            bgd_user_roles = _get_role_codes(db, uid)
            if not (bgd_user_roles & bgd_role_codes):
                continue
            out.append({
                "id": uid,
                "name": bgd_user.full_name or bgd_user.username or uid,
                "unit_id": "",
                "role_codes": sorted(bgd_user_roles),
            })

    if block_code == "HANH_CHINH":
        if bool(role_codes & {"ROLE_TRUONG_PHONG", "ROLE_PHO_PHONG"}):
            for item in members:
                item_roles = set(item.get("role_codes") or [])
                if item["unit_id"] != str(scope_unit_id or ""):
                    continue
                if _belongs_to_any_subunit(db, item["id"], str(scope_unit_id or "")):
                    continue

                # Giữ phạm vi nhân sự trực thuộc phòng như hiện có,
                # đồng thời bảo đảm Trưởng phòng/Phó phòng luôn nằm trong danh sách.
                out.append(item)

        elif bool(role_codes & {"ROLE_TO_TRUONG", "ROLE_PHO_TO"}):
            for item in members:
                item_roles = set(item.get("role_codes") or [])
                if item["unit_id"] != str(scope_subunit_id or ""):
                    continue

                # Giữ phạm vi nhân sự trong tổ như hiện có,
                # đồng thời bảo đảm Tổ trưởng/Tổ phó luôn nằm trong danh sách.
                out.append(item)

    elif block_code == "CHUYEN_MON":
        if bool(role_codes & {"ROLE_TRUONG_KHOA", "ROLE_PHO_TRUONG_KHOA"}):
            for item in members:
                item_roles = set(item.get("role_codes") or [])
                if item["unit_id"] not in {str(scope_unit_id or ""), str(scope_subunit_id or "")}:
                    continue

                if (
                    any(_is_role_doctor(x) for x in item_roles)
                    or any(_is_role_bgd(x) for x in item_roles)
                    or bool(item_roles & khoa_leader_role_codes)
                ):
                    out.append(item)

            # Bổ sung BGĐ dù không thuộc membership trực tiếp của khoa.
            _append_global_bgd_users()

        elif bool(role_codes & {"ROLE_DIEU_DUONG_TRUONG", "ROLE_KY_THUAT_VIEN_TRUONG"}):
            for item in members:
                item_roles = set(item.get("role_codes") or [])
                if item["unit_id"] not in {str(scope_unit_id or ""), str(scope_subunit_id or "")}:
                    continue

                if (
                    any(_is_role_nursing_or_tech(x) for x in item_roles)
                    or bool(item_roles & khoa_nursing_tech_leader_role_codes)
                ):
                    out.append(item)

        elif bool(role_codes & {"ROLE_TRUONG_DON_VI", "ROLE_PHO_DON_VI"}):
            for item in members:
                item_roles = set(item.get("role_codes") or [])
                if item["unit_id"] != str(scope_subunit_id or ""):
                    continue

                if (
                    any(_is_role_doctor(x) for x in item_roles)
                    or bool(item_roles & subunit_leader_role_codes)
                ):
                    out.append(item)

        elif bool(role_codes & {"ROLE_DIEU_DUONG_TRUONG_DON_VI", "ROLE_KY_THUAT_VIEN_TRUONG_DON_VI"}):
            for item in members:
                item_roles = set(item.get("role_codes") or [])
                if item["unit_id"] != str(scope_subunit_id or ""):
                    continue

                if (
                    any(_is_role_nursing_or_tech(x) for x in item_roles)
                    or bool(item_roles & subunit_nursing_tech_leader_role_codes)
                    or bool(item_roles & khoa_nursing_tech_leader_role_codes)
                ):
                    out.append(item)

    seen = set()
    dedup = []
    for item in out:
        if item["id"] in seen:
            continue
        seen.add(item["id"])
        dedup.append(item)

    dedup.sort(key=lambda x: (str(x.get("name") or "").lower(), str(x.get("id") or "")))
    return dedup

def _enrich_schedule_summary_group_filters(
    db: Session,
    *,
    scope_info: dict[str, Any],
    items: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    if not items:
        return items

    block_code = str(scope_info.get("block_code") or "").upper()
    group_defs = _group_def_map(block_code)

    group_keys = []
    seen_keys = set()
    for item in items:
        raw_key = str(item.get("group_key") or "").strip()
        key = raw_key
        if raw_key.startswith("GROUP::"):
            key = raw_key.split("GROUP::", 1)[1].strip()
        if key and key not in seen_keys:
            seen_keys.add(key)
            group_keys.append(key)

    config_map: dict[str, dict[str, Any]] = {}
    parent_ids: list[str] = []

    if group_keys:
        placeholders = []
        params: dict[str, Any] = {}
        for idx, key in enumerate(group_keys):
            p = f"gk_{idx}"
            placeholders.append(f":{p}")
            params[p] = key

        cfg_rows = db.execute(
            text(
                f"""
                SELECT id, parent_group_id, group_type, group_name
                FROM leave_group_configs
                WHERE id IN ({", ".join(placeholders)})
                """
            ),
            params,
        ).fetchall()

        for row in cfg_rows:
            item = _row_to_dict(row)
            config_map[str(item.get("id") or "")] = item
            parent_id = str(item.get("parent_group_id") or "").strip()
            if parent_id:
                parent_ids.append(parent_id)

    parent_map: dict[str, dict[str, Any]] = {}
    parent_ids = sorted(set([x for x in parent_ids if x]))
    if parent_ids:
        placeholders = []
        params = {}
        for idx, key in enumerate(parent_ids):
            p = f"pg_{idx}"
            placeholders.append(f":{p}")
            params[p] = key

        parent_rows = db.execute(
            text(
                f"""
                SELECT id, group_name, group_type
                FROM leave_group_configs
                WHERE id IN ({", ".join(placeholders)})
                """
            ),
            params,
        ).fetchall()

        for row in parent_rows:
            item = _row_to_dict(row)
            parent_map[str(item.get("id") or "")] = item

    for item in items:
        raw_group_key = str(item.get("group_key") or "").strip()
        group_key = raw_group_key
        if raw_group_key.startswith("GROUP::"):
            group_key = raw_group_key.split("GROUP::", 1)[1].strip()

        cfg = config_map.get(group_key)

        item["filter_group_type_code"] = ""
        item["filter_group_type_label"] = ""
        item["filter_child_group_id"] = ""
        item["filter_child_group_name"] = ""

        if not cfg:
            continue

        cfg_type = str(cfg.get("group_type") or "").strip()
        cfg_name = str(cfg.get("group_name") or "").strip()

        if cfg_type == PRIORITY_CHILD_GROUP_TYPE:
            parent_id = str(cfg.get("parent_group_id") or "").strip()
            parent_cfg = parent_map.get(parent_id, {})
            item["filter_group_type_code"] = PRIORITY_PARENT_GROUP_TYPE
            item["filter_group_type_label"] = "Nhóm luân phiên"
            item["filter_child_group_id"] = group_key
            item["filter_child_group_name"] = cfg_name
            item["filter_parent_group_name"] = str(parent_cfg.get("group_name") or "").strip()
        else:
            item["filter_group_type_code"] = cfg_type
            item["filter_group_type_label"] = (
                group_defs.get(cfg_type, {}).get("label")
                or ("Nhóm luân phiên" if cfg_type == PRIORITY_PARENT_GROUP_TYPE else cfg_name or cfg_type)
            )
            item["filter_child_group_id"] = ""
            item["filter_child_group_name"] = ""
            item["filter_parent_group_name"] = ""

    return items

def _month_offset_from_created_at(created_at_str: str, target_month: date) -> int:
    try:
        created_dt = datetime.strptime(str(created_at_str), "%Y-%m-%d %H:%M:%S").date()
    except Exception:
        created_dt = target_month
    return max(0, (target_month.year - created_dt.year) * 12 + (target_month.month - created_dt.month))


def _effective_priority_order(base_order: int, total_groups: int, month_offset: int) -> int:
    if total_groups <= 0:
        return 0
    base_order = max(1, int(base_order or 1))
    return ((base_order - 1 - month_offset) % total_groups) + 1


def _priority_deadline_day(priority_base_day: int, effective_order: int) -> int:
    base_day = max(1, int(priority_base_day or 1))
    return base_day + max(0, int(effective_order or 1) - 1)

def _effective_member_order(base_order: int, total_members: int, month_offset: int) -> int:
    if total_members <= 0:
        return 0
    base_order = max(1, int(base_order or 1))
    return ((base_order - 1 - month_offset) % total_members) + 1


def _normalize_member_orders(members: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for idx, item in enumerate(members, start=1):
        row = dict(item)
        row["member_order"] = int(row.get("member_order") or 0) or idx
        out.append(row)
    return out


def _build_priority_member_contexts(
    *,
    members: list[dict[str, Any]],
    month_offset: int,
) -> list[dict[str, Any]]:
    normalized = _normalize_member_orders(members)
    total_members = len(normalized)
    for row in normalized:
        row["effective_member_order"] = _effective_member_order(
            int(row.get("member_order") or 0),
            total_members,
            month_offset,
        )
    return sorted(
        normalized,
        key=lambda x: (
            int(x.get("effective_member_order") or 0),
            int(x.get("member_order") or 0),
            str(x.get("user_name") or "").lower(),
        ),
    )

def _load_priority_parent_groups(
    db: Session,
    *,
    block_code: str,
    scope_unit_id: str | None,
    scope_subunit_id: str | None,
) -> list[dict[str, Any]]:
    rows = db.execute(
        text(
            """
            SELECT *
            FROM leave_group_configs
            WHERE block_code = :block_code
              AND COALESCE(scope_unit_id, '') = :scope_unit_id
              AND COALESCE(scope_subunit_id, '') = :scope_subunit_id
              AND group_type = :group_type
            ORDER BY is_active DESC, created_at ASC, group_name ASC
            """
        ),
        {
            "block_code": block_code or "",
            "scope_unit_id": scope_unit_id or "",
            "scope_subunit_id": scope_subunit_id or "",
            "group_type": PRIORITY_PARENT_GROUP_TYPE,
        },
    ).fetchall()
    return [_row_to_dict(r) for r in rows]


def _load_priority_children(db: Session, parent_group_id: str) -> list[dict[str, Any]]:
    rows = db.execute(
        text(
            """
            SELECT *
            FROM leave_group_configs
            WHERE parent_group_id = :parent_group_id
              AND group_type = :group_type
            ORDER BY priority_order ASC, created_at ASC, group_name ASC
            """
        ),
        {
            "parent_group_id": str(parent_group_id),
            "group_type": PRIORITY_CHILD_GROUP_TYPE,
        },
    ).fetchall()
    children = [_row_to_dict(r) for r in rows]
    for child in children:
        mem_rows = db.execute(
            text(
                """
                SELECT *
                FROM leave_group_members
                WHERE group_id = :group_id
                ORDER BY
                    CASE WHEN COALESCE(member_order, 0) <= 0 THEN 999999 ELSE member_order END ASC,
                    created_at ASC,
                    user_name ASC
                """
            ),
            {"group_id": str(child["id"])},
        ).fetchall()
        child["members"] = _normalize_member_orders([_row_to_dict(r) for r in mem_rows])
    return children


def _build_priority_parents_context(
    db: Session,
    *,
    block_code: str,
    scope_unit_id: str | None,
    scope_subunit_id: str | None,
    current_month: date,
) -> list[dict[str, Any]]:
    parents = _load_priority_parent_groups(
        db,
        block_code=block_code,
        scope_unit_id=scope_unit_id,
        scope_subunit_id=scope_subunit_id,
    )
    for parent in parents:
        children = _load_priority_children(db, str(parent["id"]))
        total_children = len(children)
        month_offset = _month_offset_from_created_at(parent.get("created_at") or "", current_month)
        for child in children:
            base_order = int(child.get("priority_order") or 0)
            child["effective_priority_order"] = _effective_priority_order(base_order, total_children, month_offset)
            default_deadline_day = _priority_deadline_day(
                int(parent.get("priority_base_day") or 0),
                int(child["effective_priority_order"] or 0),
            )
            child["deadline_day"] = int(child.get("custom_deadline_day") or 0) or default_deadline_day
            child["effective_max_off_per_day"] = int(child.get("custom_max_off_per_day") or 0) or int(parent.get("max_off_per_day") or 0)
            child["members"] = _build_priority_member_contexts(
                members=child.get("members", []),
                month_offset=month_offset,
            )
        children.sort(key=lambda x: (int(x.get("effective_priority_order") or 0), int(x.get("priority_order") or 0)))
        parent["children"] = children
    return parents


def _scope_unit_ids_for_groups(scope_info: dict[str, Any]) -> list[str]:
    out = []
    if scope_info.get("scope_unit_id"):
        out.append(str(scope_info["scope_unit_id"]))
    if scope_info.get("scope_subunit_id"):
        out.append(str(scope_info["scope_subunit_id"]))
    return out


def _load_user_roles_map(db: Session, user_ids: list[str]) -> dict[str, set[str]]:
    out: dict[str, set[str]] = {str(uid): set() for uid in user_ids}
    if not user_ids:
        return out

    rows = (
        db.query(UserRoles.user_id, Roles.code)
        .join(Roles, Roles.id == UserRoles.role_id)
        .filter(UserRoles.user_id.in_(user_ids))
        .all()
    )
    for user_id, code in rows:
        uid = str(user_id)
        out.setdefault(uid, set()).add(str(getattr(code, "value", code)).upper())
    return out


def _load_scope_users_for_groups(db: Session, scope_info: dict[str, Any]) -> list[dict[str, Any]]:
    unit_ids = _scope_unit_ids_for_groups(scope_info)
    block_code = str(scope_info.get("block_code") or "").upper()

    user_rows = []
    if unit_ids:
        user_rows = (
            db.query(Users.id, Users.full_name, Users.username)
            .join(UserUnitMemberships, UserUnitMemberships.user_id == Users.id)
            .filter(UserUnitMemberships.unit_id.in_(unit_ids))
            .distinct()
            .all()
        )

    base_map: dict[str, tuple[Any, Any, Any]] = {}
    for user_id, full_name, username in user_rows:
        base_map[str(user_id)] = (user_id, full_name, username)

    if block_code == "CHUYEN_MON":
        bgd_rows = (
            db.query(Users.id, Users.full_name, Users.username)
            .join(UserRoles, UserRoles.user_id == Users.id)
            .join(Roles, Roles.id == UserRoles.role_id)
            .filter(Roles.code.in_(["ROLE_LANH_DAO", "ROLE_BGD"]))
            .distinct()
            .all()
        )
        for user_id, full_name, username in bgd_rows:
            base_map[str(user_id)] = (user_id, full_name, username)

    merged_rows = list(base_map.values())
    user_ids = [str(row[0]) for row in merged_rows]
    roles_map = _load_user_roles_map(db, user_ids)

    out: list[dict[str, Any]] = []
    for user_id, full_name, username in merged_rows:
        uid = str(user_id)
        out.append(
            {
                "id": uid,
                "name": full_name or username or uid,
                "role_codes": sorted(list(roles_map.get(uid, set()))),
            }
        )
    out.sort(key=lambda x: (x["name"] or "").lower())
    return out


def _is_user_eligible_for_group(group_type: str, role_codes: set[str]) -> bool:
    if group_type == "MANAGER_KHOA":
        return bool(role_codes & CM_MANAGER_KHOA_ROLES)
    if group_type == "MANAGER_SUBUNIT":
        return bool(role_codes & CM_MANAGER_SUBUNIT_ROLES)
    if group_type == "MANAGER_PHONG":
        return bool(role_codes & HC_MANAGER_PHONG_ROLES)
    if group_type == "MANAGER_TO":
        return bool(role_codes & HC_MANAGER_TO_ROLES)
    return True

def _is_manager_group_type(group_type: str) -> bool:
    return group_type in {"MANAGER_PHONG", "MANAGER_TO", "MANAGER_KHOA", "MANAGER_SUBUNIT"}


def _manager_bucket_roles(group_type: str) -> set[str]:
    if group_type == "MANAGER_KHOA":
        return {"ROLE_TRUONG_KHOA", "ROLE_PHO_TRUONG_KHOA"}
    if group_type == "MANAGER_SUBUNIT":
        return {"ROLE_TRUONG_DON_VI", "ROLE_PHO_DON_VI"}
    return set()


def _load_group_members_with_roles(db: Session, group_id: str) -> list[dict[str, Any]]:
    rows = db.execute(
        text(
            """
            SELECT id, group_id, user_id, user_name, created_at, member_order
            FROM leave_group_members
            WHERE group_id = :group_id
            ORDER BY
                CASE WHEN COALESCE(member_order, 0) <= 0 THEN 999999 ELSE member_order END ASC,
                created_at ASC,
                user_name ASC
            """
        ),
        {"group_id": str(group_id)},
    ).fetchall()

    out: list[dict[str, Any]] = []
    for row in rows:
        item = _row_to_dict(row)
        item["role_codes"] = _get_role_codes(db, str(item.get("user_id") or ""))
        out.append(item)
    return out


def _manager_group_runtime_rules(
    *,
    group_type: str,
    members: list[dict[str, Any]],
) -> dict[str, Any]:
    total_members = len(members)
    general_limit = max(total_members - 1, 0)

    role_constraint_label = ""
    bucket_limit = 0
    bucket_roles = _manager_bucket_roles(group_type)

    if group_type in {"MANAGER_KHOA", "MANAGER_SUBUNIT"} and bucket_roles:
        bucket_count = sum(
            1
            for m in members
            if set(m.get("role_codes") or []) & bucket_roles
        )
        if bucket_count >= 2:
            bucket_limit = max(bucket_count - 1, 0)
            if group_type == "MANAGER_KHOA":
                role_constraint_label = f"Ràng buộc riêng Trưởng/Phó khoa: tối đa {bucket_limit} vị trí/ngày"
            else:
                role_constraint_label = f"Ràng buộc riêng Trưởng/Phó đơn vị: tối đa {bucket_limit} vị trí/ngày"

    if total_members <= 0:
        display_limit_label = "Chưa có thành viên"
    elif general_limit > 0:
        display_limit_label = f"{general_limit} người/ngày"
    else:
        display_limit_label = "0 người/ngày"

    return {
        "general_limit": general_limit,
        "bucket_roles": bucket_roles,
        "bucket_limit": bucket_limit,
        "display_limit_label": display_limit_label,
        "role_constraint_label": role_constraint_label,
        "total_members": total_members,
    }


def _sync_manager_group_max_off_per_day(db: Session, group_id: str) -> None:
    row = db.execute(
        text("SELECT id, group_type FROM leave_group_configs WHERE id = :id"),
        {"id": str(group_id)},
    ).fetchone()
    if not row:
        return

    group_item = _row_to_dict(row)
    group_type = str(group_item.get("group_type") or "")
    if not _is_manager_group_type(group_type):
        return

    members = _load_group_members_with_roles(db, str(group_id))
    rules = _manager_group_runtime_rules(group_type=group_type, members=members)

    db.execute(
        text(
            """
            UPDATE leave_group_configs
            SET max_off_per_day = :max_off_per_day,
                updated_at = :updated_at
            WHERE id = :id
            """
        ),
        {
            "id": str(group_id),
            "max_off_per_day": int(rules["general_limit"]),
            "updated_at": _dt_to_str(_now_vn()),
        },
    )


def _load_groups_for_scope(
    db: Session,
    *,
    block_code: str,
    scope_unit_id: str | None,
    scope_subunit_id: str | None,
) -> list[dict[str, Any]]:
    rows = db.execute(
        text(
            """
            SELECT *
            FROM leave_group_configs
            WHERE block_code = :block_code
              AND COALESCE(scope_unit_id, '') = :scope_unit_id
              AND COALESCE(scope_subunit_id, '') = :scope_subunit_id
            ORDER BY is_active DESC, created_at ASC, group_name ASC
            """
        ),
        {
            "block_code": block_code or "",
            "scope_unit_id": scope_unit_id or "",
            "scope_subunit_id": scope_subunit_id or "",
        },
    ).fetchall()

    groups = [_row_to_dict(r) for r in rows]
    for group in groups:
        mem_rows = db.execute(
            text(
                """
                SELECT *
                FROM leave_group_members
                WHERE group_id = :group_id
                ORDER BY created_at ASC, user_name ASC
                """
            ),
            {"group_id": str(group["id"])},
        ).fetchall()
        group["members"] = [_row_to_dict(r) for r in mem_rows]
    return groups


def _assigned_user_ids_in_scope(
    db: Session,
    *,
    block_code: str,
    scope_unit_id: str | None,
    scope_subunit_id: str | None,
) -> set[str]:
    rows = db.execute(
        text(
            """
            SELECT DISTINCT m.user_id
            FROM leave_group_members m
            JOIN leave_group_configs g ON g.id = m.group_id
            WHERE g.block_code = :block_code
              AND COALESCE(g.scope_unit_id, '') = :scope_unit_id
              AND COALESCE(g.scope_subunit_id, '') = :scope_subunit_id
              AND g.is_active = 1
            """
        ),
        {
            "block_code": block_code or "",
            "scope_unit_id": scope_unit_id or "",
            "scope_subunit_id": scope_subunit_id or "",
        },
    ).fetchall()
    return {str(r[0]) for r in rows if r and r[0]}


def _build_groups_tab_context(db: Session, user: Users, scope_info: dict[str, Any], block_code: str) -> dict[str, Any]:
    scope_unit_id = scope_info.get("scope_unit_id")
    scope_subunit_id = scope_info.get("scope_subunit_id")
    role_codes = _get_role_codes(db, user.id)

    groups = _load_groups_for_scope(
        db,
        block_code=block_code,
        scope_unit_id=scope_unit_id,
        scope_subunit_id=scope_subunit_id,
    )

    scope_users = _load_scope_users_for_groups(db, scope_info)
    assigned_user_ids = _assigned_user_ids_in_scope(
        db,
        block_code=block_code,
        scope_unit_id=scope_unit_id,
        scope_subunit_id=scope_subunit_id,
    )

    for group in groups:
        group_type = str(group.get("group_type") or "")
        member_user_ids = {str(m.get("user_id")) for m in group.get("members", [])}
        assignable_users = []
        for u in scope_users:
            user_role_codes = set(u.get("role_codes") or [])
            if not _is_user_eligible_for_group(group_type, user_role_codes):
                continue
            uid = str(u["id"])
            if uid in assigned_user_ids and uid not in member_user_ids:
                continue
            if uid in member_user_ids:
                continue
            assignable_users.append(u)
        group["assignable_users"] = assignable_users

        if _is_manager_group_type(group_type):
            members_with_roles = []
            for m in group.get("members", []):
                item = dict(m)
                item["role_codes"] = _get_role_codes(db, str(item.get("user_id") or ""))
                members_with_roles.append(item)

            rules = _manager_group_runtime_rules(
                group_type=group_type,
                members=members_with_roles,
            )
            group["computed_max_off_per_day"] = int(rules["general_limit"])
            group["display_limit_label"] = rules["display_limit_label"]
            group["role_constraint_label"] = rules["role_constraint_label"]
        else:
            group["computed_max_off_per_day"] = int(group.get("max_off_per_day") or 0)
            if group["computed_max_off_per_day"] > 0:
                group["display_limit_label"] = f"{group['computed_max_off_per_day']} người/ngày"
            elif group_type == "ROTATION":
                group["display_limit_label"] = "Luân phiên"
            else:
                group["display_limit_label"] = "Không hạn mức"
            group["role_constraint_label"] = ""

    current_month_base = _today_vn().replace(day=1)
    if current_month_base.month == 12:
        next_month_base = date(current_month_base.year + 1, 1, 1)
    else:
        next_month_base = date(current_month_base.year, current_month_base.month + 1, 1)

    priority_parents_current = _build_priority_parents_context(
        db,
        block_code=block_code,
        scope_unit_id=scope_unit_id,
        scope_subunit_id=scope_subunit_id,
        current_month=current_month_base,
    )

    priority_parents_next = _build_priority_parents_context(
        db,
        block_code=block_code,
        scope_unit_id=scope_unit_id,
        scope_subunit_id=scope_subunit_id,
        current_month=next_month_base,
    )

    normal_groups = [
        g for g in groups
        if str(g.get("group_type") or "") not in {PRIORITY_PARENT_GROUP_TYPE, PRIORITY_CHILD_GROUP_TYPE}
    ]

    return {
        "group_defs": _group_defs_for_block(block_code),
        "groups_data": normal_groups,
        "priority_parents_data": priority_parents_current,
        "priority_parents_next_data": priority_parents_next,
        "priority_current_month": current_month_base.strftime("%Y-%m"),
        "priority_next_month": next_month_base.strftime("%Y-%m"),
        "can_manage_groups": _can_manage_block(role_codes, block_code),
    }

def _get_active_group_membership_for_user(
    db: Session,
    *,
    user_id: str,
    block_code: str,
    scope_unit_id: str | None,
    scope_subunit_id: str | None,
) -> dict[str, Any] | None:
    row = db.execute(
        text(
            """
            SELECT
                g.id AS group_id,
                g.group_type,
                g.group_name,
                g.max_off_per_day,
                g.scope_unit_id,
                g.scope_unit_name,
                g.scope_subunit_id,
                g.scope_subunit_name,
                m.user_id
            FROM leave_group_members m
            JOIN leave_group_configs g ON g.id = m.group_id
            WHERE m.user_id = :user_id
              AND g.block_code = :block_code
              AND COALESCE(g.scope_unit_id, '') = :scope_unit_id
              AND COALESCE(g.scope_subunit_id, '') = :scope_subunit_id
              AND g.is_active = 1
            LIMIT 1
            """
        ),
        {
            "user_id": str(user_id),
            "block_code": block_code or "",
            "scope_unit_id": scope_unit_id or "",
            "scope_subunit_id": scope_subunit_id or "",
        },
    ).fetchone()
    if not row:
        return None
    return _row_to_dict(row)

def _resolve_priority_user_context_for_date(
    db: Session,
    *,
    user_id: str,
    block_code: str,
    scope_unit_id: str | None,
    scope_subunit_id: str | None,
    target_date: date,
) -> dict[str, Any] | None:
    month_base = target_date.replace(day=1)
    membership = _get_priority_child_membership_for_user(
        db,
        user_id=str(user_id),
        block_code=block_code,
        scope_unit_id=scope_unit_id,
        scope_subunit_id=scope_subunit_id,
    )
    if not membership:
        return None

    child_rows = db.execute(
        text(
            """
            SELECT COUNT(*) AS total_count
            FROM leave_group_configs
            WHERE parent_group_id = :parent_group_id
              AND group_type = :group_type
              AND is_active = 1
            """
        ),
        {
            "parent_group_id": str(membership["parent_group_id"]),
            "group_type": PRIORITY_CHILD_GROUP_TYPE,
        },
    ).fetchone()
    total_children = int(_row_to_dict(child_rows).get("total_count") or 0) if child_rows else 0

    member_rows = db.execute(
        text(
            """
            SELECT *
            FROM leave_group_members
            WHERE group_id = :group_id
            ORDER BY
                CASE WHEN COALESCE(member_order, 0) <= 0 THEN 999999 ELSE member_order END ASC,
                created_at ASC,
                user_name ASC
            """
        ),
        {"group_id": str(membership["child_group_id"])},
    ).fetchall()
    members = _normalize_member_orders([_row_to_dict(r) for r in member_rows])
    total_members = len(members)

    month_offset = _month_offset_from_created_at(membership.get("parent_created_at") or "", month_base)
    effective_priority_order = _effective_priority_order(
        int(membership.get("child_priority_order") or 0),
        total_children,
        month_offset,
    )

    base_member_order = int(membership.get("child_member_order") or 0)
    if base_member_order <= 0:
        for idx, row in enumerate(members, start=1):
            if str(row.get("user_id")) == str(user_id):
                base_member_order = idx
                break
    effective_member_order = _effective_member_order(base_member_order, total_members, month_offset)

    return {
        "parent_group_id": str(membership["parent_group_id"]),
        "parent_group_name": membership.get("parent_group_name") or "",
        "child_group_id": str(membership["child_group_id"]),
        "child_group_name": membership.get("child_group_name") or "",
        "effective_priority_order": effective_priority_order,
        "member_order": base_member_order,
        "effective_member_order": effective_member_order,
        "max_off_per_day": int(membership.get("max_off_per_day") or 0),
        "deadline_day": _priority_deadline_day(
            int(membership.get("priority_base_day") or 0),
            effective_priority_order,
        ),
    }


def _is_priority_registration_day_open(
    *,
    target_date: date,
    deadline_day: int | None,
) -> tuple[bool, str]:
    if not deadline_day:
        return True, ""

    today = _today_vn()
    current_month_base = date(today.year, today.month, 1)

    if current_month_base.month == 12:
        next_month_base = date(current_month_base.year + 1, 1, 1)
    else:
        next_month_base = date(current_month_base.year, current_month_base.month + 1, 1)

    target_month_base = target_date.replace(day=1)
    if target_month_base != next_month_base:
        return False, "Chỉ được đăng ký cho tháng kế tiếp."

    cutoff_day = _month_cutoff(current_month_base)

    if int(deadline_day) > int(cutoff_day):
        return False, (
            f"Nhóm của bạn không còn ngày đăng ký hợp lệ trong tháng hiện tại "
            f"vì ngày ưu tiên ({deadline_day}) rơi sau mốc khóa tháng ({cutoff_day})."
        )

    if today.day < int(deadline_day):
        return False, (
            f"Chưa đến ngày đăng ký của nhóm bạn. "
            f"Nhóm hiện tại chỉ được đăng ký trong ngày {deadline_day} của tháng hiện tại."
        )

    if today.day > int(deadline_day):
        return False, (
            f"Đã hết ngày đăng ký của nhóm bạn. "
            f"Nhóm hiện tại chỉ được đăng ký trong ngày {deadline_day} của tháng hiện tại."
        )

    if today.day > int(cutoff_day):
        return False, f"Đã qua mốc khóa đăng ký tháng sau trong tháng hiện tại (ngày {cutoff_day})."

    return True, ""

def _find_blocking_higher_priority_member(
    db: Session,
    *,
    user: Users,
    target_date: date,
) -> dict[str, Any] | None:
    scope_info = _scope_payload(db, user)
    current_ctx = _resolve_priority_user_context_for_date(
        db,
        user_id=str(user.id),
        block_code=scope_info.get("block_code") or "",
        scope_unit_id=scope_info.get("scope_unit_id"),
        scope_subunit_id=scope_info.get("scope_subunit_id"),
        target_date=target_date,
    )
    if not current_ctx:
        return None

    current_member_order = int(current_ctx.get("effective_member_order") or 0)
    if current_member_order <= 1:
        return None

    member_rows = db.execute(
        text(
            """
            SELECT m.user_id, m.user_name
            FROM leave_group_members m
            WHERE m.group_id = :group_id
            ORDER BY
                CASE WHEN COALESCE(m.member_order, 0) <= 0 THEN 999999 ELSE m.member_order END ASC,
                m.created_at ASC,
                m.user_name ASC
            """
        ),
        {"group_id": str(current_ctx["child_group_id"])},
    ).fetchall()
    members = _normalize_member_orders([_row_to_dict(r) for r in member_rows])
    month_offset = _month_offset_from_created_at(
        db.execute(
            text("SELECT created_at FROM leave_group_configs WHERE id = :id"),
            {"id": str(current_ctx["parent_group_id"])},
        ).fetchone()._mapping["created_at"],
        target_date.replace(day=1),
    )

    higher_user_ids: list[str] = []
    higher_user_names: list[str] = []

    for row in members:
        effective_order = _effective_member_order(
            int(row.get("member_order") or 0),
            len(members),
            month_offset,
        )
        if effective_order < current_member_order:
            higher_user_ids.append(str(row.get("user_id") or ""))
            higher_user_names.append(str(row.get("user_name") or ""))

    if not higher_user_ids:
        return None

    year_month = target_date.strftime("%Y-%m")
    done_rows = db.execute(
        text(
            """
            SELECT DISTINCT user_id
            FROM leave_requests
            WHERE group_key = :group_key
              AND status IN ('PENDING', 'APPROVED')
              AND substr(start_date, 1, 7) = :year_month
            """
        ),
        {
            "group_key": f"GROUP::{current_ctx['parent_group_id']}",
            "year_month": year_month,
        },
    ).fetchall()
    done_user_ids = {str(r[0]) for r in done_rows if r and r[0]}

    for uid, uname in zip(higher_user_ids, higher_user_names):
        if uid not in done_user_ids:
            return {
                "user_id": uid,
                "user_name": uname,
                "message": "Chưa đến lượt, bạn phải chờ người có vị trí ưu tiên cao hơn bạn hoàn thành đăng ký.",
            }

    return None

def _get_priority_child_membership_for_user(
    db: Session,
    *,
    user_id: str,
    block_code: str,
    scope_unit_id: str | None,
    scope_subunit_id: str | None,
) -> dict[str, Any] | None:
    row = db.execute(
        text(
            """
            SELECT
                child.id AS child_group_id,
                child.group_name AS child_group_name,
                child.priority_order AS child_priority_order,
                m.member_order AS child_member_order,                
                parent.id AS parent_group_id,
                parent.group_name AS parent_group_name,
                parent.priority_base_day,
                parent.max_off_per_day,
                parent.created_at AS parent_created_at,
                parent.scope_unit_id,
                parent.scope_subunit_id
            FROM leave_group_members m
            JOIN leave_group_configs child ON child.id = m.group_id
            JOIN leave_group_configs parent ON parent.id = child.parent_group_id
            WHERE m.user_id = :user_id
              AND child.group_type = :child_type
              AND parent.group_type = :parent_type
              AND parent.block_code = :block_code
              AND COALESCE(parent.scope_unit_id, '') = :scope_unit_id
              AND COALESCE(parent.scope_subunit_id, '') = :scope_subunit_id
              AND child.is_active = 1
              AND parent.is_active = 1
            LIMIT 1
            """
        ),
        {
            "user_id": str(user_id),
            "child_type": PRIORITY_CHILD_GROUP_TYPE,
            "parent_type": PRIORITY_PARENT_GROUP_TYPE,
            "block_code": block_code or "",
            "scope_unit_id": scope_unit_id or "",
            "scope_subunit_id": scope_subunit_id or "",
        },
    ).fetchone()
    if not row:
        return None
    return _row_to_dict(row)

def _rotation_assigned_user_ids_for_dates(
    db: Session,
    *,
    group_id: str,
    start_date: date,
    end_date: date,
) -> dict[str, str]:
    rows = db.execute(
        text(
            """
            SELECT id, user_id, user_name, created_at
            FROM leave_group_members
            WHERE group_id = :group_id
            ORDER BY created_at ASC, user_name ASC
            """
        ),
        {"group_id": str(group_id)},
    ).fetchall()

    members = [_row_to_dict(r) for r in rows]
    if not members:
        return {}

    group_row = db.execute(
        text("SELECT created_at FROM leave_group_configs WHERE id = :group_id"),
        {"group_id": str(group_id)},
    ).fetchone()
    if not group_row:
        return {}

    group_created_at = _row_to_dict(group_row).get("created_at") or ""
    try:
        base_date = datetime.strptime(str(group_created_at), "%Y-%m-%d %H:%M:%S").date()
    except Exception:
        base_date = start_date

    workdays = list(_iter_workdays(start_date, end_date))
    result: dict[str, str] = {}
    total = len(members)

    for d in workdays:
        offset = (d - base_date).days
        idx = offset % total if total > 0 else 0
        result[d.isoformat()] = str(members[idx]["user_id"])
    return result

def _resolve_effective_leave_group(
    db: Session,
    user: Users,
    target_month: date | None = None,
) -> tuple[str, str, Units | None, int, int | None]:
    scope_info = _scope_payload(db, user)
    block_code = scope_info.get("block_code") or ""
    scope_unit_id = scope_info.get("scope_unit_id")
    scope_subunit_id = scope_info.get("scope_subunit_id")
    month_base = (target_month or _today_vn()).replace(day=1)

    priority_membership = _get_priority_child_membership_for_user(
        db,
        user_id=str(user.id),
        block_code=block_code,
        scope_unit_id=scope_unit_id,
        scope_subunit_id=scope_subunit_id,
    )
    if priority_membership:
        child_rows = db.execute(
            text(
                """
                SELECT COUNT(*) AS total_count
                FROM leave_group_configs
                WHERE parent_group_id = :parent_group_id
                  AND group_type = :group_type
                  AND is_active = 1
                """
            ),
            {
                "parent_group_id": str(priority_membership["parent_group_id"]),
                "group_type": PRIORITY_CHILD_GROUP_TYPE,
            },
        ).fetchone()
        total_children = int(_row_to_dict(child_rows).get("total_count") or 0) if child_rows else 0
        month_offset = _month_offset_from_created_at(priority_membership.get("parent_created_at") or "", month_base)
        effective_order = _effective_priority_order(
            int(priority_membership.get("child_priority_order") or 0),
            total_children,
            month_offset,
        )
        deadline_day = _priority_deadline_day(
            int(priority_membership.get("priority_base_day") or 0),
            effective_order,
        )

        target_unit = None
        if priority_membership.get("scope_subunit_id"):
            target_unit = db.get(Units, priority_membership.get("scope_subunit_id"))
        elif priority_membership.get("scope_unit_id"):
            target_unit = db.get(Units, priority_membership.get("scope_unit_id"))

        return (
            f"GROUP::{priority_membership['parent_group_id']}",
            f"{priority_membership['parent_group_name']} / {priority_membership['child_group_name']}",
            target_unit,
            int(priority_membership.get("max_off_per_day") or 0),
            deadline_day,
        )

    membership = _get_active_group_membership_for_user(
        db,
        user_id=str(user.id),
        block_code=block_code,
        scope_unit_id=scope_unit_id,
        scope_subunit_id=scope_subunit_id,
    )
    if membership:
        group_id = str(membership.get("group_id"))
        group_name = str(membership.get("group_name") or "Nhóm nghỉ")
        group_type = str(membership.get("group_type") or "")
        max_off_per_day = int(membership.get("max_off_per_day") or 0)

        if _is_manager_group_type(group_type):
            members_with_roles = _load_group_members_with_roles(db, group_id)
            rules = _manager_group_runtime_rules(
                group_type=group_type,
                members=members_with_roles,
            )
            max_off_per_day = int(rules["general_limit"])

        target_unit = None
        if membership.get("scope_subunit_id"):
            target_unit = db.get(Units, membership.get("scope_subunit_id"))
        elif membership.get("scope_unit_id"):
            target_unit = db.get(Units, membership.get("scope_unit_id"))

        return (
            f"GROUP::{group_id}",
            group_name,
            target_unit,
            max_off_per_day,
            None,
        )

    group_key, group_label, unit = _resolve_group_bucket(db, user)
    return group_key, group_label, unit, 1, None
    

@router.get("/api/unread-count")
def leave_schedule_unread_count_api(
    request: Request,
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)
    count = _load_unread_notice_count(db, user.id)
    return {"ok": True, "unread_leave_schedule_count": count}


@router.post("/notices/mark-read")
def mark_leave_schedule_notices_read(
    request: Request,
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)
    db.execute(
        text(
            """
            UPDATE leave_schedule_notices
            SET is_read = 1
            WHERE user_id = :user_id AND is_read = 0
            """
        ),
        {"user_id": str(user.id)},
    )
    db.commit()
    return RedirectResponse(url="/leave-schedule", status_code=302)

def _resolve_group_bucket(db: Session, user: Users) -> tuple[str, str, Units | None]:
    scope_unit, scope_subunit = _resolve_unit_scope(db, user.id)

    if scope_subunit is not None:
        return (
            f"UNIT::{scope_subunit.id}",
            getattr(scope_subunit, "ten_don_vi", None) or "Đơn vị trực thuộc",
            scope_subunit,
        )

    if scope_unit is not None:
        return (
            f"UNIT::{scope_unit.id}",
            getattr(scope_unit, "ten_don_vi", None) or "Đơn vị",
            scope_unit,
        )

    return (
        f"USER::{user.id}",
        user.full_name or user.username or "Người dùng",
        None,
    )


def _can_admin_extra_days(db: Session, user: Users) -> bool:
    return ROLE_ADMIN in _get_role_codes(db, user.id)


def _can_adjust_after_cutoff(db: Session, user: Users) -> bool:
    return bool(_get_role_codes(db, user.id) & MANAGER_ROLE_CODES)


def _can_approve(db: Session, user: Users, block_code: str | None = None) -> bool:
    role_codes = _get_role_codes(db, user.id)
    if block_code:
        return _can_approve_block(role_codes, block_code)
    return bool(role_codes & APPROVER_ROLE_CODES)


def _can_access_leave_request_scope(db: Session, user: Users, item: dict[str, Any]) -> bool:
    role_codes = _get_role_codes(db, user.id)
    if _is_admin_or_leader(role_codes):
        return True

    scope_info = _scope_payload(db, user)
    scope_block_code = str(scope_info.get("block_code") or "").strip().upper()
    request_block_code = str(item.get("block_code") or "").strip().upper()
    if request_block_code and scope_block_code and request_block_code != scope_block_code:
        return False

    request_unit_id = str(item.get("unit_id") or "").strip()
    if not request_unit_id:
        return False

    allowed_unit_ids = _scope_unit_ids_for_view(scope_info)
    return request_unit_id in allowed_unit_ids


def _role_label_for_user(db: Session, user: Users) -> str:
    role_codes = _get_role_codes(db, user.id)
    if ROLE_TRUONG_KHOA in role_codes:
        return "Trưởng khoa"
    if ROLE_PHO_TRUONG_KHOA in role_codes:
        return "Phó khoa"
    if ROLE_KTV_TRUONG in role_codes:
        return "Kỹ thuật viên trưởng"
    if ROLE_ADMIN in role_codes:
        return "Admin"
    return "Người dùng"


def _is_standard_group_registration_open(
    db: Session,
    user: Users,
    *,
    target_date: date,
) -> tuple[bool, str]:
    today = _today_vn()
    current_month_base = date(today.year, today.month, 1)

    if current_month_base.month == 12:
        next_month_base = date(current_month_base.year + 1, 1, 1)
    else:
        next_month_base = date(current_month_base.year, current_month_base.month + 1, 1)

    target_month_base = target_date.replace(day=1)
    if target_month_base != next_month_base:
        return False, "Chỉ được đăng ký cho tháng kế tiếp."

    cutoff_day = _month_cutoff(current_month_base)

    if today.day > cutoff_day:
        if _can_adjust_after_cutoff(db, user):
            return True, ""
        return False, f"Đã qua thời hạn đăng ký tháng sau trong tháng hiện tại (sau ngày {cutoff_day})."

    return True, ""


def _sum_extra_days(db: Session, user_id: str, leave_year: int) -> int:
    row = db.execute(
        text(
            """
            SELECT COALESCE(SUM(extra_days), 0) AS total_extra
            FROM leave_year_adjustments
            WHERE user_id = :user_id AND leave_year = :leave_year
            """
        ),
        {"user_id": user_id, "leave_year": leave_year},
    ).fetchone()
    if not row:
        return 0
    return int(_row_to_dict(row).get("total_extra") or 0)


def _sum_days_by_year(db: Session, user_id: str, leave_type: str, leave_year: int, statuses: tuple[str, ...]) -> float:
    row = db.execute(
        text(
            f"""
            SELECT COALESCE(SUM(day_count), 0) AS total_days
            FROM leave_requests
            WHERE user_id = :user_id
              AND leave_type = :leave_type
              AND status IN ({','.join([f':st{i}' for i in range(len(statuses))])})
              AND substr(start_date, 1, 4) = :leave_year
            """
        ),
        {
            "user_id": user_id,
            "leave_type": leave_type,
            "leave_year": str(leave_year),
            **{f"st{i}": statuses[i] for i in range(len(statuses))},
        },
    ).fetchone()
    if not row:
        return 0.0
    return float(_row_to_dict(row).get("total_days") or 0)


def _sum_days_by_month(db: Session, user_id: str, leave_type: str, year_month: str, statuses: tuple[str, ...]) -> float:
    row = db.execute(
        text(
            f"""
            SELECT COALESCE(SUM(day_count), 0) AS total_days
            FROM leave_requests
            WHERE user_id = :user_id
              AND leave_type = :leave_type
              AND status IN ({','.join([f':st{i}' for i in range(len(statuses))])})
              AND substr(start_date, 1, 7) = :year_month
            """
        ),
        {
            "user_id": user_id,
            "leave_type": leave_type,
            "year_month": year_month,
            **{f"st{i}": statuses[i] for i in range(len(statuses))},
        },
    ).fetchone()
    if not row:
        return 0.0
    return float(_row_to_dict(row).get("total_days") or 0)


def _validate_quota(db: Session, user: Users, leave_type: str, start_date: date, end_date: date, requested_days: int) -> None:
    meta = LEAVE_TYPE_META[leave_type]
    statuses = ("PENDING", "APPROVED")

    if leave_type == "F":
        leave_year = start_date.year
        total_extra = _sum_extra_days(db, user.id, leave_year)
        used_days = _sum_days_by_year(db, user.id, leave_type, leave_year, statuses)
        max_days = int(meta["max_year"] or 0) + total_extra
        if used_days + requested_days > max_days:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Nghỉ phép năm vượt giới hạn. Đã dùng/đang chờ: {used_days:g} ngày; "
                    f"được phép: {max_days:g} ngày."
                ),
            )

    if leave_type == "NL":
        leave_year = start_date.year
        used_days = _sum_days_by_year(db, user.id, leave_type, leave_year, statuses)
        max_days = int(meta["max_year"] or 0)
        if used_days + requested_days > max_days:
            raise HTTPException(
                status_code=400,
                detail=f"Nghỉ lễ, tết vượt mức cấu hình {max_days:g} ngày/năm.",
            )

    if leave_type == "P":
        month_map: dict[str, int] = {}
        for d in _iter_workdays(start_date, end_date):
            ym = d.strftime("%Y-%m")
            month_map[ym] = month_map.get(ym, 0) + 1
        for ym, add_days in month_map.items():
            used_days = _sum_days_by_month(db, user.id, leave_type, ym, statuses)
            max_days = int(meta["max_month"] or 0)
            if used_days + add_days > max_days:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"Nghỉ phép tháng vượt giới hạn 4 ngày công/tháng. "
                        f"Tháng {ym} đã dùng/đang chờ: {used_days:g} ngày."
                    ),
                )


def _find_group_conflict(
    db: Session,
    user: Users,
    group_key: str,
    leave_type: str,
    start_date: date,
    end_date: date,
    max_off_per_day: int = 1,
) -> dict[str, Any] | None:
    rows = db.execute(
        text(
            """
            SELECT *
            FROM leave_requests
            WHERE group_key = :group_key
              AND status IN ('PENDING', 'APPROVED')
              AND NOT (end_date < :start_date OR start_date > :end_date)
            ORDER BY start_date ASC, created_at ASC
            """
        ),
        {
            "group_key": group_key,
            "start_date": _date_to_str(start_date),
            "end_date": _date_to_str(end_date),
        },
    ).fetchall()

    wanted_dates = {d.isoformat() for d in _iter_workdays(start_date, end_date)}

                
    existing_count_by_date: dict[str, int] = {}
    for row in rows:
        item = _row_to_dict(row)

        if item.get("user_id") == user.id:
            return {
                "message": "Khoảng thời gian này đã có đăng ký nghỉ trước đó của chính người dùng.",
                "row": item,
            }

        for d in _iter_workdays(
            _parse_date(item["start_date"], "Ngày bắt đầu"),
            _parse_date(item["end_date"], "Ngày kết thúc"),
        ):
            ds = d.isoformat()
            if ds in wanted_dates:
                existing_count_by_date[ds] = existing_count_by_date.get(ds, 0) + 1

    if max_off_per_day and max_off_per_day > 0:
        for ds in sorted(wanted_dates):
            if existing_count_by_date.get(ds, 0) >= int(max_off_per_day):
                return {
                    "message": f"Ngày {ds} đã đạt hạn mức {int(max_off_per_day)} người nghỉ trong nhóm.",
                    "row": None,
                }

    return None


def _format_request_row(item: dict[str, Any]) -> dict[str, Any]:
    meta = LEAVE_TYPE_META.get(item.get("leave_type") or "", {})
    status = item.get("status") or ""
    status_label = {
        "PENDING": "Chờ duyệt",
        "APPROVED": "Đã duyệt",
        "CANCELLED": "Đã hủy",
        "REJECTED": "Đã từ chối",
    }.get(status, status)

    item = dict(item)
    item["leave_type_label"] = meta.get("label") or item.get("leave_type")
    item["status_label"] = status_label
    item["can_cancel"] = status == "PENDING"

    start_date = str(item.get("start_date") or "")
    end_date = str(item.get("end_date") or "")
    item["date_label"] = start_date if start_date == end_date else f"{start_date} đến {end_date}"

    item["reject_reason"] = (item.get("reject_reason") or "").strip()
    item["can_reject"] = status == "PENDING"
    return item

def _load_requests(
    db: Session,
    user: Users,
    viewer_is_manager: bool,
    list_mode: str,
    request_month: str = "",
) -> list[dict[str, Any]]:
    if viewer_is_manager:
        sql = "SELECT * FROM leave_requests ORDER BY start_date DESC, created_at DESC"
        rows = db.execute(text(sql)).fetchall()
    else:
        if request_month:
            rows = db.execute(
                text(
                    """
                    SELECT *
                    FROM leave_requests
                    WHERE user_id = :user_id
                      AND substr(start_date, 1, 7) = :request_month
                    ORDER BY start_date DESC, created_at DESC
                    """
                ),
                {
                    "user_id": user.id,
                    "request_month": request_month,
                },
            ).fetchall()
        else:
            rows = db.execute(
                text(
                    "SELECT * FROM leave_requests WHERE user_id = :user_id ORDER BY start_date DESC, created_at DESC"
                ),
                {"user_id": user.id},
            ).fetchall()

    items = [_format_request_row(_row_to_dict(r)) for r in rows]
    if list_mode == "pending":
        return [x for x in items if x.get("status") == "PENDING"]
    if list_mode == "approved":
        return [x for x in items if x.get("status") == "APPROVED"]
    return items


def _build_balance_cards(db: Session, user: Users) -> list[dict[str, Any]]:
    current_year = _today_vn().year
    current_month = _today_vn().strftime("%Y-%m")
    extra_f = _sum_extra_days(db, user.id, current_year)
    used_f = _sum_days_by_year(db, user.id, "F", current_year, ("PENDING", "APPROVED"))
    used_p = _sum_days_by_month(db, user.id, "P", current_month, ("PENDING", "APPROVED"))
    used_nl = _sum_days_by_year(db, user.id, "NL", current_year, ("PENDING", "APPROVED"))
    return [
        {
            "code": "F",
            "label": "Nghỉ phép năm",
            "used": used_f,
            "max": 12 + extra_f,
            "note": f"Gồm 12 ngày chuẩn + {extra_f} ngày admin cộng thêm" if extra_f else "Mặc định 12 ngày làm việc/năm",
        },
        {
            "code": "P",
            "label": "Nghỉ phép tháng",
            "used": used_p,
            "max": 4,
            "note": f"Áp dụng cho tháng {current_month}",
        },
        {
            "code": "NL",
            "label": "Nghỉ lễ, tết",
            "used": used_nl,
            "max": 12,
            "note": "Theo mức cấu hình hiện tại của khoa",
        },
        {
            "code": "DB",
            "label": "Phép đặc biệt",
            "used": 0,
            "max": None,
            "note": "Không giới hạn trong module này và không bị ràng buộc bởi quy tắc tạo nhóm",
        },        
        {
            "code": "H",
            "label": "Đi học",
            "used": 0,
            "max": None,
            "note": "Không giới hạn số ngày trong module này",
        },
        {
            "code": "CT",
            "label": "Đi công tác",
            "used": 0,
            "max": None,
            "note": "Không giới hạn số ngày trong module này",
        },
    ]

def _scope_unit_ids_for_view(scope_info: dict[str, Any]) -> list[str]:
    out: list[str] = []
    if scope_info.get("scope_unit_id"):
        out.append(str(scope_info["scope_unit_id"]))
    if scope_info.get("scope_subunit_id"):
        out.append(str(scope_info["scope_subunit_id"]))
    return out


def _load_exhausted_dates_for_user(db: Session, user: Users, month_str: str) -> list[dict[str, Any]]:
    year, month = [int(x) for x in month_str.split("-")]
    first_day = date(year, month, 1)
    last_day = (date(year + (1 if month == 12 else 0), 1 if month == 12 else month + 1, 1) - timedelta(days=1))

    rows = db.execute(
        text(
            """
            SELECT start_date, end_date, leave_type, group_key, group_label
            FROM leave_requests
            WHERE user_id = :user_id
              AND status IN ('PENDING', 'APPROVED')
              AND start_date <= :last_day
              AND end_date >= :first_day
            ORDER BY start_date ASC
            """
        ),
        {
            "user_id": str(user.id),
            "first_day": first_day.isoformat(),
            "last_day": last_day.isoformat(),
        },
    ).fetchall()

    used_days_by_type: dict[str, int] = {}
    exhausted: list[dict[str, Any]] = []

    for row in rows:
        item = _row_to_dict(row)
        s = _parse_date(item["start_date"], "start_date")
        e = _parse_date(item["end_date"], "end_date")
        leave_type = str(item.get("leave_type") or "").upper()

        for d in _iter_workdays(max(s, first_day), min(e, last_day)):
            used_days_by_type.setdefault(leave_type, 0)
            used_days_by_type[leave_type] += 1

            if leave_type == "P" and used_days_by_type[leave_type] >= 4:
                exhausted.append(
                    {
                        "date": d.isoformat(),
                        "reason": "Đã đạt hạn mức nghỉ phép tháng (P) trong tháng.",
                    }
                )

    dedup: dict[str, dict[str, Any]] = {}
    for item in exhausted:
        dedup[item["date"]] = item
    return list(dedup.values())


def _load_scope_leave_calendar_summary(
    db: Session,
    *,
    scope_info: dict[str, Any],
    month_str: str,
) -> list[dict[str, Any]]:
    unit_ids = _scope_unit_ids_for_view(scope_info)
    if not unit_ids:
        return []

    year, month = [int(x) for x in month_str.split("-")]
    first_day = date(year, month, 1)
    last_day = (
        date(year + (1 if month == 12 else 0), 1 if month == 12 else month + 1, 1)
        - timedelta(days=1)
    )

    unit_placeholders = []
    params: dict[str, Any] = {
        "first_day": first_day.isoformat(),
        "last_day": last_day.isoformat(),
    }

    for idx, unit_id in enumerate(unit_ids):
        key = f"unit_id_{idx}"
        unit_placeholders.append(f":{key}")
        params[key] = str(unit_id)

    sql = f"""
        SELECT
            id,
            user_id,
            user_name,
            unit_id,
            unit_name,
            group_key,
            group_label,
            leave_type,
            symbol,
            start_date,
            end_date,
            day_count,
            reason,
            status,
            reject_reason,
            approved_by_name,
            approved_role
        FROM leave_requests
        WHERE unit_id IN ({', '.join(unit_placeholders)})
          AND start_date <= :last_day
          AND end_date >= :first_day
        ORDER BY user_name ASC, start_date ASC
    """

    rows = db.execute(text(sql), params).fetchall()

    out: list[dict[str, Any]] = []
    for row in rows:
        item = _format_request_row(_row_to_dict(row))
        item["can_swap"] = item.get("status") == "PENDING"
        out.append(item)

    out = _enrich_schedule_summary_group_filters(
        db,
        scope_info=scope_info,
        items=out,
    )
    return out
    
    
def _load_scope_pending_approval_summary(
    db: Session,
    *,
    scope_info: dict[str, Any],
    month_str: str,
) -> list[dict[str, Any]]:
    unit_ids = _scope_unit_ids_for_view(scope_info)
    if not unit_ids:
        return []

    year, month = [int(x) for x in month_str.split("-")]
    first_day = date(year, month, 1)
    last_day = (
        date(year + (1 if month == 12 else 0), 1 if month == 12 else month + 1, 1)
        - timedelta(days=1)
    )

    unit_placeholders = []
    params: dict[str, Any] = {
        "first_day": first_day.isoformat(),
        "last_day": last_day.isoformat(),
    }

    for idx, unit_id in enumerate(unit_ids):
        key = f"unit_id_{idx}"
        unit_placeholders.append(f":{key}")
        params[key] = str(unit_id)

    sql = f"""
        SELECT
            user_id,
            user_name,
            unit_name,
            group_label,
            leave_type,
            SUM(COALESCE(day_count, 0)) AS total_days,
            COUNT(*) AS request_count
        FROM leave_requests
        WHERE unit_id IN ({', '.join(unit_placeholders)})
          AND status = 'PENDING'
          AND start_date <= :last_day
          AND end_date >= :first_day
        GROUP BY user_id, user_name, unit_name, group_label, leave_type
        ORDER BY user_name ASC, group_label ASC
    """

    rows = db.execute(text(sql), params).fetchall()

    summary_map: dict[str, dict[str, Any]] = {}
    for row in rows:
        item = _row_to_dict(row)
        user_id = str(item.get("user_id") or "")
        if not user_id:
            continue

        entry = summary_map.setdefault(
            user_id,
            {
                "user_id": user_id,
                "user_name": item.get("user_name") or "-",
                "unit_name": item.get("unit_name") or "-",
                "group_label": item.get("group_label") or "-",
                "P": 0.0,
                "F": 0.0,
                "NL": 0.0,
                "DB": 0.0,
                "CT": 0.0,
                "H": 0.0,
                "request_count": 0,
            },
        )

        leave_type = str(item.get("leave_type") or "").upper()
        total_days = float(item.get("total_days") or 0)
        if leave_type in {"P", "F", "NL", "DB", "CT", "H"}:
            entry[leave_type] += total_days
        entry["request_count"] += int(item.get("request_count") or 0)

    out = list(summary_map.values())
    out.sort(key=lambda x: ((x.get("user_name") or "").lower(), (x.get("group_label") or "").lower()))
    return out    

def _load_leave_notices(
    db: Session,
    user_id: str,
    block_code: str = "",
    month_str: str = "",
    limit: int = 100,
) -> list[dict[str, Any]]:
    params = {
        "user_id": str(user_id),
        "limit": int(limit),
        "roster_notice_1": "ROSTER_APPROVED",
        "roster_notice_2": "ROSTER_ADJUST_APPROVED",
    }
    sql = """
        SELECT *
        FROM leave_schedule_notices
        WHERE user_id = :user_id
          AND notice_type NOT IN (:roster_notice_1, :roster_notice_2)
    """
    if block_code:
        sql += " AND COALESCE(block_code, '') = :block_code "
        params["block_code"] = block_code

    if month_str:
        sql += " AND substr(COALESCE(created_at, ''), 1, 7) = :month_str "
        params["month_str"] = month_str

    sql += " ORDER BY is_read ASC, created_at DESC LIMIT :limit "

    rows = db.execute(text(sql), params).fetchall()
    return [_row_to_dict(r) for r in rows]


def _load_pending_swap_requests(
    db: Session,
    *,
    scope_info: dict[str, Any],
    block_code: str,
) -> list[dict[str, Any]]:
    rows = db.execute(
        text(
            """
            SELECT *
            FROM leave_swap_requests
            WHERE block_code = :block_code
              AND COALESCE(scope_unit_id, '') = :scope_unit_id
              AND COALESCE(scope_subunit_id, '') = :scope_subunit_id
              AND status = 'PENDING'
            ORDER BY created_at DESC
            """
        ),
        {
            "block_code": block_code or "",
            "scope_unit_id": scope_info.get("scope_unit_id") or "",
            "scope_subunit_id": scope_info.get("scope_subunit_id") or "",
        },
    ).fetchall()
    return [_row_to_dict(r) for r in rows]

def _has_any_approved_leave_for_scope(
    db: Session,
    *,
    scope_info: dict[str, Any],
    block_code: str,
    target_month: str,
) -> bool:
    month_base = _month_base_from_str(target_month)
    if month_base.month == 12:
        next_month = date(month_base.year + 1, 1, 1)
    else:
        next_month = date(month_base.year, month_base.month + 1, 1)
    month_end = next_month - timedelta(days=1)

    row = db.execute(
        text(
            """
            SELECT COUNT(*) AS total_count
            FROM leave_requests
            WHERE status = 'APPROVED'
              AND COALESCE(block_code, '') = :block_code
              AND COALESCE(unit_id, '') IN (:scope_unit_id, :scope_subunit_id)
              AND NOT (end_date < :month_start OR start_date > :month_end)
            """
        ),
        {
            "block_code": block_code or "",
            "scope_unit_id": scope_info.get("scope_unit_id") or "",
            "scope_subunit_id": scope_info.get("scope_subunit_id") or "",
            "month_start": month_base.isoformat(),
            "month_end": month_end.isoformat(),
        },
    ).fetchone()
    return int((_row_to_dict(row).get("total_count") if row else 0) or 0) > 0


def _load_roster_plan_header(
    db: Session,
    *,
    scope_info: dict[str, Any],
    block_code: str,
    target_month: str,
) -> dict[str, Any] | None:
    row = db.execute(
        text(
            """
            SELECT *
            FROM leave_roster_plans
            WHERE block_code = :block_code
              AND COALESCE(scope_unit_id, '') = :scope_unit_id
              AND COALESCE(scope_subunit_id, '') = :scope_subunit_id
              AND target_month = :target_month
            ORDER BY
                CASE WHEN status = 'APPROVED' THEN 0 ELSE 1 END,
                updated_at DESC
            LIMIT 1
            """
        ),
        {
            "block_code": block_code or "",
            "scope_unit_id": scope_info.get("scope_unit_id") or "",
            "scope_subunit_id": scope_info.get("scope_subunit_id") or "",
            "target_month": target_month,
        },
    ).fetchone()
    return _row_to_dict(row) if row else None


def _load_roster_plan_items(db: Session, plan_id: str) -> list[dict[str, Any]]:
    rows = db.execute(
        text(
            """
            SELECT *
            FROM leave_roster_plan_items
            WHERE plan_id = :plan_id
            ORDER BY work_date ASC, user_name ASC
            """
        ),
        {"plan_id": str(plan_id)},
    ).fetchall()
    return [_row_to_dict(r) for r in rows]


def _scope_unit_ids_for_calendar(scope_info: dict[str, Any]) -> list[str]:
    out: list[str] = []
    for raw in [scope_info.get("scope_unit_id"), scope_info.get("scope_subunit_id")]:
        value = str(raw or "").strip()
        if value and value not in out:
            out.append(value)
    return out


def _calendar_month_range(month_str: str) -> tuple[date, date]:
    month_base = _month_base_from_str(month_str)
    if month_base.month == 12:
        next_month = date(month_base.year + 1, 1, 1)
    else:
        next_month = date(month_base.year, month_base.month + 1, 1)
    return month_base, next_month - timedelta(days=1)


def _calendar_year_options(center_month_str: str, years_back: int = 1, years_forward: int = 1) -> list[int]:
    month_base = _month_base_from_str(center_month_str)
    return list(range(month_base.year - years_back, month_base.year + years_forward + 1))


def _calendar_month_options() -> list[dict[str, str]]:
    return [{"value": f"{i:02d}", "label": f"Tháng {i:02d}"} for i in range(1, 13)]


def _calendar_role_label(role_codes: set[str]) -> str:
    ordered = [
        ("ROLE_TRUONG_KHOA", "Trưởng khoa"),
        ("ROLE_PHO_TRUONG_KHOA", "Phó khoa"),
        ("ROLE_DIEU_DUONG_TRUONG", "Điều dưỡng trưởng"),
        ("ROLE_KY_THUAT_VIEN_TRUONG", "KTV trưởng"),
        ("ROLE_TRUONG_DON_VI", "Trưởng đơn vị"),
        ("ROLE_PHO_DON_VI", "Phó đơn vị"),
        ("ROLE_DIEU_DUONG_TRUONG_DON_VI", "ĐDT đơn vị"),
        ("ROLE_KY_THUAT_VIEN_TRUONG_DON_VI", "KTVT đơn vị"),
        ("ROLE_TRUONG_PHONG", "Trưởng phòng"),
        ("ROLE_PHO_PHONG", "Phó phòng"),
        ("ROLE_TO_TRUONG", "Tổ trưởng"),
        ("ROLE_PHO_TO", "Tổ phó"),
        ("ROLE_BAC_SI", "Bác sĩ"),
        ("ROLE_DIEU_DUONG", "Điều dưỡng"),
        ("ROLE_KY_THUAT_VIEN", "KTV"),
        ("ROLE_DUOC_SI", "Dược sĩ"),
        ("ROLE_NHAN_VIEN", "Nhân viên"),
    ]
    for code, label in ordered:
        if code in role_codes:
            return label
    return ""


def _load_scope_calendar_users(
    db: Session,
    *,
    scope_info: dict[str, Any],
) -> list[dict[str, Any]]:
    unit_ids = _scope_unit_ids_for_calendar(scope_info)
    if not unit_ids:
        return []

    rows = (
        db.query(
            Users.id,
            Users.full_name,
            Users.username,
            Units.ten_don_vi,
        )
        .join(UserUnitMemberships, UserUnitMemberships.user_id == Users.id)
        .join(Units, Units.id == UserUnitMemberships.unit_id)
        .filter(UserUnitMemberships.unit_id.in_(unit_ids))
        .distinct()
        .all()
    )

    user_ids = [str(user_id) for user_id, _, _, _ in rows if user_id]
    roles_map = _load_user_roles_map(db, user_ids)

    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for user_id, full_name, username, unit_name in rows:
        uid = str(user_id or "")
        if not uid or uid in seen:
            continue
        seen.add(uid)
        role_codes = set(roles_map.get(uid, set()))
        out.append(
            {
                "user_id": uid,
                "user_name": full_name or username or uid,
                "unit_name": unit_name or "",
                "role_label": _calendar_role_label(role_codes),
            }
        )

    out.sort(key=lambda x: str(x.get("user_name") or "").lower())
    return out


def _load_scope_calendar_leave_rows(
    db: Session,
    *,
    scope_info: dict[str, Any],
    block_code: str,
    month_str: str,
) -> list[dict[str, Any]]:
    unit_ids = _scope_unit_ids_for_calendar(scope_info)
    if not unit_ids:
        return []

    month_start, month_end = _calendar_month_range(month_str)

    placeholders = []
    params: dict[str, Any] = {
        "block_code": block_code or "",
        "month_start": month_start.isoformat(),
        "month_end": month_end.isoformat(),
    }
    for idx, unit_id in enumerate(unit_ids):
        key = f"unit_id_{idx}"
        placeholders.append(f":{key}")
        params[key] = unit_id

    sql = f"""
        SELECT
            id,
            user_id,
            user_name,
            unit_id,
            unit_name,
            leave_type,
            symbol,
            start_date,
            end_date,
            status
        FROM leave_requests
        WHERE status = 'APPROVED'
          AND COALESCE(block_code, '') = :block_code
          AND COALESCE(unit_id, '') IN ({", ".join(placeholders)})
          AND NOT (end_date < :month_start OR start_date > :month_end)
        ORDER BY user_name ASC, start_date ASC
    """
    rows = db.execute(text(sql), params).fetchall()
    return [_row_to_dict(r) for r in rows]


def _load_scope_calendar_roster_rows(
    db: Session,
    *,
    scope_info: dict[str, Any],
    block_code: str,
    month_str: str,
) -> list[dict[str, Any]]:
    plan_header = _load_roster_plan_header(
        db,
        scope_info=scope_info,
        block_code=block_code,
        target_month=month_str,
    )
    if not plan_header:
        return []

    return _load_roster_plan_items(db, str(plan_header["id"]))


def _calendar_leave_cell_style(symbol: str) -> dict[str, str]:
    symbol = str(symbol or "").upper().strip()
    if symbol in {"P", "F"}:
        return {"bg": "#fee2e2", "text": "#991b1b"}
    if symbol == "NL":
        return {"bg": "#fed7aa", "text": "#9a3412"}
    if symbol in {"CT", "H", "ĐB", "DB"}:
        return {"bg": "#ede9fe", "text": "#5b21b6"}
    return {"bg": "#e0f2fe", "text": "#0c4a6e"}


def _build_view_calendar_context(
    db: Session,
    *,
    scope_info: dict[str, Any],
    block_code: str,
    month_str: str,
    view_mode: str,
) -> dict[str, Any]:
    users = _load_scope_calendar_users(db, scope_info=scope_info)
    day_defs = _build_calendar_days(month_str)
    leave_rows = _load_scope_calendar_leave_rows(
        db,
        scope_info=scope_info,
        block_code=block_code,
        month_str=month_str,
    )
    roster_rows = _load_scope_calendar_roster_rows(
        db,
        scope_info=scope_info,
        block_code=block_code,
        month_str=month_str,
    )

    day_keys = [str(x.get("date") or "") for x in day_defs]
    matrix: list[dict[str, Any]] = []

    leave_map: dict[tuple[str, str], dict[str, Any]] = {}
    month_start, month_end = _calendar_month_range(month_str)

    for row in leave_rows:
        row_start = _parse_date(str(row.get("start_date") or ""), "Ngày bắt đầu")
        row_end = _parse_date(str(row.get("end_date") or ""), "Ngày kết thúc")
        actual_start = max(row_start, month_start)
        actual_end = min(row_end, month_end)
        for d in _iter_workdays(actual_start, actual_end):
            day_key = d.isoformat()
            leave_map[(str(row.get("user_id") or ""), day_key)] = {
                "symbol": str(row.get("symbol") or ""),
                "leave_type": str(row.get("leave_type") or ""),
                "style": _calendar_leave_cell_style(str(row.get("symbol") or "")),
            }

    roster_map: dict[tuple[str, str], dict[str, Any]] = {}
    post_roster_rest_map: dict[tuple[str, str], dict[str, Any]] = {}

    for row in roster_rows:
        user_id = str(row.get("user_id") or "")
        work_date_str = str(row.get("work_date") or "")
        if not user_id or not work_date_str:
            continue

        roster_map[(user_id, work_date_str)] = {
            "shift_code": str(row.get("shift_code") or "TRUC"),
            "note": str(row.get("note") or ""),
        }

        try:
            work_date = _parse_date(work_date_str, "Ngày trực")
        except Exception:
            continue

        next_rest_date = _next_workday(work_date)
        if month_start <= next_rest_date <= month_end:
            next_rest_key = next_rest_date.isoformat()
            post_roster_rest_map[(user_id, next_rest_key)] = {
                "symbol": "NT",
                "title": "Nghỉ trực",
            }

    for user in users:
        uid = str(user.get("user_id") or "")
        cells: list[dict[str, Any]] = []
        for item in day_defs:
            day_key = str(item.get("date") or "")
            leave_item = leave_map.get((uid, day_key))
            roster_item = roster_map.get((uid, day_key))
            post_roster_rest_item = post_roster_rest_map.get((uid, day_key))

            has_leave = leave_item is not None
            has_roster = roster_item is not None
            has_post_roster_rest = (post_roster_rest_item is not None) and (not has_leave)

            cell = {
                "date": day_key,
                "day": item.get("day"),
                "weekday": item.get("weekday"),
                "is_sunday": bool(item.get("is_sunday")),
                "leave_symbol": leave_item.get("symbol") if leave_item else "",
                "leave_type": leave_item.get("leave_type") if leave_item else "",
                "has_roster": has_roster,
                "roster_symbol": "TR" if has_roster else "",
                "has_post_roster_rest": has_post_roster_rest,
                "post_roster_rest_symbol": "NT" if has_post_roster_rest else "",
                "bg_color": "",
                "text_color": "",
                "title": "",
            }

            if view_mode in {"combined", "leave"} and has_leave:
                cell["bg_color"] = leave_item["style"]["bg"]
                cell["text_color"] = leave_item["style"]["text"]
                cell["title"] = f"Nghỉ: {cell['leave_symbol']}"
            elif view_mode in {"combined", "leave"} and has_post_roster_rest:
                cell["bg_color"] = "#ffedd5"
                cell["text_color"] = "#9a3412"
                cell["title"] = "Nghỉ trực"
            elif view_mode in {"combined", "roster"} and has_roster:
                cell["bg_color"] = "#dbeafe"
                cell["text_color"] = "#1d4ed8"
                cell["title"] = "Trực"

            cells.append(cell)

        matrix.append(
            {
                "user_id": uid,
                "user_name": user.get("user_name") or "",
                "unit_name": user.get("unit_name") or "",
                "role_label": user.get("role_label") or "",
                "cells": cells,
            }
        )

    month_base = _month_base_from_str(month_str)
    return {
        "target_month": month_str,
        "target_year": str(month_base.year),
        "target_month_only": f"{month_base.month:02d}",
        "year_options": _calendar_year_options(month_str),
        "month_options": _calendar_month_options(),
        "view_mode": view_mode if view_mode in {"combined", "leave", "roster"} else "combined",
        "day_defs": day_defs,
        "matrix_rows": matrix,
        "leave_count": len(leave_rows),
        "roster_count": len(roster_rows),
        "post_roster_rest_count": len(post_roster_rest_map),
    }


def _render_calendar_export_html(
    *,
    title: str,
    view_calendar_context: dict[str, Any],
) -> str:
    day_defs = view_calendar_context.get("day_defs") or []
    rows = view_calendar_context.get("matrix_rows") or []

    parts: list[str] = []
    parts.append("<html><head><meta charset='utf-8'></head><body>")
    parts.append(f"<h2>{title}</h2>")
    parts.append("<table border='1' cellspacing='0' cellpadding='4'>")
    parts.append("<thead>")
    parts.append("<tr><th rowspan='2'>Nhân sự</th>")
    for day in day_defs:
        parts.append(f"<th>{day.get('day')}</th>")
    parts.append("</tr>")
    parts.append("<tr>")
    for day in day_defs:
        weekday = int(day.get("weekday") or 0)
        label = ["T2", "T3", "T4", "T5", "T6", "T7", "CN"][weekday]
        parts.append(f"<th>{label}</th>")
    parts.append("</tr>")
    parts.append("</thead><tbody>")

    for row in rows:
        role_label = str(row.get("role_label") or "").strip()
        unit_name = str(row.get("unit_name") or "").strip()
        left_text = str(row.get("user_name") or "")
        extra = " - ".join([x for x in [role_label, unit_name] if x])
        if extra:
            left_text = f"{left_text} ({extra})"

        parts.append("<tr>")
        parts.append(f"<td>{left_text}</td>")
        for cell in row.get("cells") or []:
            bg = str(cell.get("bg_color") or "")
            text_color = str(cell.get("text_color") or "")
            content = ""
            if cell.get("leave_symbol"):
                content = str(cell.get("leave_symbol") or "")
            elif cell.get("has_roster"):
                content = "TR"
            style_parts = []
            if bg:
                style_parts.append(f"background:{bg}")
            if text_color:
                style_parts.append(f"color:{text_color}")
            style_attr = f" style=\"{';'.join(style_parts)}\"" if style_parts else ""
            parts.append(f"<td{style_attr}>{content}</td>")
        parts.append("</tr>")

    parts.append("</tbody></table></body></html>")
    return "".join(parts)

def _load_roster_adjustments(
    db: Session,
    plan_id: str,
    *,
    status: str | None = None,
) -> list[dict[str, Any]]:
    sql = """
        SELECT *
        FROM leave_roster_adjustments
        WHERE plan_id = :plan_id
    """
    params = {"plan_id": str(plan_id)}
    if status:
        sql += " AND status = :status "
        params["status"] = status
    sql += " ORDER BY created_at DESC "
    rows = db.execute(text(sql), params).fetchall()
    return [_row_to_dict(r) for r in rows]


def _build_roster_context(
    db: Session,
    *,
    user: Users,
    scope_info: dict[str, Any],
    block_code: str,
    target_month: str,
) -> dict[str, Any]:
    candidates = _resolve_roster_candidates(
        db,
        user=user,
        scope_info=scope_info,
        block_code=block_code,
    )
    can_open_roster = _has_any_approved_leave_for_scope(
        db,
        scope_info=scope_info,
        block_code=block_code,
        target_month=target_month,
    )
    plan_header = _load_roster_plan_header(
        db,
        scope_info=scope_info,
        block_code=block_code,
        target_month=target_month,
    )
    plan_items = _load_roster_plan_items(db, str(plan_header["id"])) if plan_header else []
    pending_adjustments = _load_roster_adjustments(db, str(plan_header["id"]), status="PENDING") if plan_header else []

    forbidden_map_by_user: dict[str, dict[str, str]] = {}
    for item in candidates:
        forbidden_map_by_user[item["id"]] = _roster_forbidden_dates_for_user(db, item["id"], target_month)

    return {
        "can_open_roster": can_open_roster,
        "candidates": candidates,
        "plan_header": plan_header or {},
        "plan_items": plan_items,
        "pending_adjustments": pending_adjustments,
        "forbidden_map_by_user": forbidden_map_by_user,
    }


def _build_roster_notice_context(
    db: Session,
    *,
    user: Users,
    scope_info: dict[str, Any],
    block_code: str,
    target_month: str,
    week_no: str = "",
) -> dict[str, Any]:
    unit_ids = [str(x) for x in [scope_info.get("scope_unit_id"), scope_info.get("scope_subunit_id")] if x]
    user_ids = _unit_member_user_ids(db, unit_ids)

    plan_header = _load_roster_plan_header(
        db,
        scope_info=scope_info,
        block_code=block_code,
        target_month=target_month,
    )
    if not plan_header:
        return {"items": [], "week_no": week_no}

    items = _load_roster_plan_items(db, str(plan_header["id"]))
    items = [x for x in items if str(x.get("user_id") or "") in {str(uid) for uid in user_ids}]

    if week_no and str(week_no).isdigit():
        wanted_week = int(week_no)
        items = [
            x for x in items
            if _week_of_month(_parse_date(str(x.get("work_date") or ""), "Ngày trực")) == wanted_week
        ]

    return {
        "items": items,
        "week_no": week_no,
        "plan_header": plan_header,
    }

def _is_cutoff_locked(db: Session, user: Users, target_month: date | None = None) -> bool:
    today = _today_vn()
    current_month_base = date(today.year, today.month, 1)

    if current_month_base.month == 12:
        next_month_base = date(current_month_base.year + 1, 1, 1)
    else:
        next_month_base = date(current_month_base.year, current_month_base.month + 1, 1)

    target_base = (target_month or next_month_base).replace(day=1)

    if target_base != next_month_base:
        return True

    cutoff = _month_cutoff(current_month_base)
    if today.day <= cutoff:
        return False
    return not _can_adjust_after_cutoff(db, user)

@router.get("")
def leave_schedule_index(
    request: Request,
    list_mode: str = "all",
    month: str = "",
    request_month: str = "",
    block_tab: str = "",
    sub_tab: str = "",
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    scope_info = _scope_payload(db, user)
    role_codes = _get_role_codes(db, user.id)

    current_block_tab = block_tab if block_tab in {"HANH_CHINH", "CHUYEN_MON"} else ""
    allowed_sub_tabs = {"register", "roster", "view_calendar", "groups", "leave_notice", "roster_notice"}
    current_sub_tab = sub_tab if (current_block_tab and sub_tab in allowed_sub_tabs) else ""

    effective_block_code = current_block_tab or (scope_info.get("block_code") or "")
    effective_block_label = _block_label(effective_block_code) if effective_block_code else ""

    viewer_is_manager = _can_manage_block(role_codes, effective_block_code) if effective_block_code else False
    approver = _can_approve_block(role_codes, effective_block_code) if effective_block_code else False

    if month:
        current_month = month
    else:
        today = _today_vn()
        if today.month == 12:
            current_month = f"{today.year + 1}-01"
        else:
            current_month = f"{today.year}-{today.month + 1:02d}"

    balances = _build_balance_cards(db, user)
    effective_request_month = request_month or current_month

    requests = _load_requests(
        db,
        user,
        viewer_is_manager or approver or _can_admin_extra_days(db, user),
        list_mode,
        effective_request_month,
    )
    can_submit, can_submit_message = _can_show_register_button(
        db,
        user,
        target_month_str=current_month,
    )
    group_key, group_label, primary_unit = _resolve_group_bucket(db, user)

    exhausted_dates = _load_exhausted_dates_for_user(db, user, current_month)
    calendar_days = _build_calendar_days(current_month)
    group_rule_blocked_dates = _load_group_rule_blocked_dates(db, user, current_month)
    schedule_summary = []
    approval_summary = []
    pending_swaps = []
    leave_notices = []
    roster_context = {}
    roster_notice_context = {}
    view_calendar_context = {}

    msg = request.query_params.get("msg", "")
    error = request.query_params.get("error", "")
    week_no = request.query_params.get("week", "")
    view_mode = request.query_params.get("view_mode", "combined")
    unread_notice_count = _load_unread_notice_count(db, user.id)
    unread_leave_notice_count = _load_leave_notice_unread_count(db, user.id)
    unread_roster_notice_count = _load_roster_notice_unread_count(db, user.id)

    groups_context = {}
    if current_block_tab and current_sub_tab == "groups":
        groups_context = _build_groups_tab_context(db, user, scope_info, current_block_tab)

    if current_block_tab and current_sub_tab == "roster":
        roster_context = _build_roster_context(
            db,
            user=user,
            scope_info=scope_info,
            block_code=effective_block_code,
            target_month=current_month,
        )
    if current_block_tab and current_sub_tab == "view_calendar":
        view_calendar_context = _build_view_calendar_context(
            db,
            scope_info=scope_info,
            block_code=effective_block_code,
            month_str=current_month,
            view_mode=view_mode,
        )

    if current_block_tab and current_sub_tab == "roster_notice":
        roster_notice_context = {
            "items": _load_roster_notices_for_user(
                db,
                user.id,
                effective_block_code,
                month_str=current_month,
                week_no=week_no,
                limit=100,
            )
        }

    if current_block_tab and current_sub_tab == "register" and (viewer_is_manager or approver):
        schedule_summary = _load_scope_leave_calendar_summary(
            db,
            scope_info=scope_info,
            month_str=current_month,
        )
        approval_summary = _load_scope_pending_approval_summary(
            db,
            scope_info=scope_info,
            month_str=current_month,
        )
        pending_swaps = _load_pending_swap_requests(
            db,
            scope_info=scope_info,
            block_code=effective_block_code,
        )

    if current_block_tab and current_sub_tab == "leave_notice":
        leave_notices = _load_leave_notices(
            db,
            user.id,
            effective_block_code,
            month_str=current_month,
            limit=100,
        )

    return templates.TemplateResponse(
        "leave_schedule.html",
        {
            "request": request,
            "app_name": settings.APP_NAME,
            "company_name": settings.COMPANY_NAME,
            "block_tab": current_block_tab,
            "sub_tab": current_sub_tab,
            "current_list_mode": list_mode,
            "request_month": effective_request_month,
            "requests": requests,
            "balances": balances,
            "can_submit": can_submit,
            "can_approve": approver,
            "can_admin_adjust": _can_admin_extra_days(db, user),
            "is_manager": viewer_is_manager,
            "group_key": group_key,
            "group_label": group_label,
            "primary_unit_name": getattr(primary_unit, "ten_don_vi", "") if primary_unit else "",
            "leave_type_meta": LEAVE_TYPE_META,
            "today_str": _today_vn().isoformat(),
            "current_month": current_month,
            "cutoff_day": _month_cutoff(_today_vn()),
            "message": msg,
            "error": error,
            "scope_info": scope_info,
            "scope_display": scope_info["scope_display"],
            "block_code": effective_block_code,
            "block_label": effective_block_label,
            "unread_notice_count": unread_notice_count,
            "unread_leave_notice_count": unread_leave_notice_count,
            "unread_roster_notice_count": unread_roster_notice_count,
            "group_defs": groups_context.get("group_defs", []),
            "priority_parents_data": groups_context.get("priority_parents_data", []),
            "priority_parents_next_data": groups_context.get("priority_parents_next_data", []),
            "priority_current_month": groups_context.get("priority_current_month", ""),
            "priority_next_month": groups_context.get("priority_next_month", ""),
            "groups_data": groups_context.get("groups_data", []),
            "can_manage_groups": groups_context.get("can_manage_groups", False),
            "exhausted_dates": exhausted_dates,
            "schedule_summary": schedule_summary,
            "approval_summary": approval_summary,
            "pending_swaps": pending_swaps,
            "leave_notices": leave_notices,
            "roster_context": roster_context,
            "roster_notice_context": roster_notice_context,
            "view_calendar_context": view_calendar_context,
            "view_mode": view_mode,
            "week_no": week_no,
            "calendar_days": calendar_days,
            "group_rule_blocked_dates": group_rule_blocked_dates,
            "can_submit_message": can_submit_message,
        },
    )

@router.post("/roster/create-draft")
def create_roster_draft(
    request: Request,
    block_tab: str = Form(...),
    month: str = Form(...),
    selected_user_ids: list[str] = Form([]),
    selected_work_dates: list[str] = Form([]),
    notes: list[str] = Form([]),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    scope_info = _scope_payload(db, user)
    role_codes = _get_role_codes(db, user.id)
    effective_block = block_tab or (scope_info.get("block_code") or "")
    if not _can_manage_block(role_codes, effective_block):
        raise HTTPException(status_code=403, detail="Bạn không có quyền lập lịch trực.")

    if not _has_any_approved_leave_for_scope(
        db,
        scope_info=scope_info,
        block_code=effective_block,
        target_month=month,
    ):
        raise HTTPException(status_code=400, detail="Chưa có kết quả phê duyệt nghỉ, chưa thể lập lịch trực.")

    candidates = _resolve_roster_candidates(
        db,
        user=user,
        scope_info=scope_info,
        block_code=effective_block,
    )
    allowed_ids = {str(x["id"]) for x in candidates}

    if not (len(selected_user_ids) == len(selected_work_dates) == len(notes)):
        raise HTTPException(status_code=400, detail="Dữ liệu lịch trực không đồng nhất.")

    target_rows: list[dict[str, Any]] = []
    for idx in range(len(selected_user_ids)):
        uid = str(selected_user_ids[idx] or "").strip()
        work_date = str(selected_work_dates[idx] or "").strip()
        note = str(notes[idx] or "").strip()
        if not uid or not work_date:
            continue
        if uid not in allowed_ids:
            raise HTTPException(status_code=400, detail="Có người dùng ngoài phạm vi được phép phân trực.")
        _validate_roster_work_date(
            db,
            user_id=uid,
            target_month=month,
            work_date_str=work_date,
        )
        target_rows.append(
            {
                "user_id": uid,
                "user_name": next((x["name"] for x in candidates if str(x["id"]) == uid), uid),
                "work_date": work_date,
                "note": note,
            }
        )

    if not target_rows:
        raise HTTPException(
            status_code=400,
            detail="Chưa có ngày trực hợp lệ. Mỗi nhân sự phải chọn ít nhất 1 ngày trực hợp lệ trên lịch."
        )

    now_str = _dt_to_str(_now_vn())
    old_plan = _load_roster_plan_header(
        db,
        scope_info=scope_info,
        block_code=effective_block,
        target_month=month,
    )
    if old_plan and old_plan.get("status") == "APPROVED":
        raise HTTPException(status_code=400, detail="Lịch trực tháng này đã được phê duyệt, không thể ghi đè nháp.")

    if old_plan:
        db.execute(text("DELETE FROM leave_roster_plan_items WHERE plan_id = :plan_id"), {"plan_id": str(old_plan["id"])})
        db.execute(
            text(
                """
                UPDATE leave_roster_plans
                SET planner_user_id = :planner_user_id,
                    planner_user_name = :planner_user_name,
                    planner_role_codes = :planner_role_codes,
                    status = 'DRAFT',
                    updated_at = :updated_at
                WHERE id = :id
                """
            ),
            {
                "id": str(old_plan["id"]),
                "planner_user_id": str(user.id),
                "planner_user_name": user.full_name or user.username or "",
                "planner_role_codes": _csv_role_codes(role_codes),
                "updated_at": now_str,
            },
        )
        plan_id = str(old_plan["id"])
    else:
        plan_id = str(uuid.uuid4())
        db.execute(
            text(
                """
                INSERT INTO leave_roster_plans (
                    id, block_code, scope_unit_id, scope_unit_name, scope_subunit_id, scope_subunit_name,
                    target_month, planner_user_id, planner_user_name, planner_role_codes,
                    status, created_at, updated_at
                ) VALUES (
                    :id, :block_code, :scope_unit_id, :scope_unit_name, :scope_subunit_id, :scope_subunit_name,
                    :target_month, :planner_user_id, :planner_user_name, :planner_role_codes,
                    'DRAFT', :created_at, :updated_at
                )
                """
            ),
            {
                "id": plan_id,
                "block_code": effective_block,
                "scope_unit_id": scope_info.get("scope_unit_id") or "",
                "scope_unit_name": scope_info.get("scope_unit_name") or "",
                "scope_subunit_id": scope_info.get("scope_subunit_id") or "",
                "scope_subunit_name": scope_info.get("scope_subunit_name") or "",
                "target_month": month,
                "planner_user_id": str(user.id),
                "planner_user_name": user.full_name or user.username or "",
                "planner_role_codes": _csv_role_codes(role_codes),
                "created_at": now_str,
                "updated_at": now_str,
            },
        )

    for item in target_rows:
        db.execute(
            text(
                """
                INSERT INTO leave_roster_plan_items (
                    id, plan_id, user_id, user_name, work_date, shift_code, note, status, created_at, updated_at
                ) VALUES (
                    :id, :plan_id, :user_id, :user_name, :work_date, 'TRUC', :note, 'PLANNED', :created_at, :updated_at
                )
                """
            ),
            {
                "id": str(uuid.uuid4()),
                "plan_id": plan_id,
                "user_id": item["user_id"],
                "user_name": item["user_name"],
                "work_date": item["work_date"],
                "note": item["note"],
                "created_at": now_str,
                "updated_at": now_str,
            },
        )

    db.commit()

    return RedirectResponse(
        url=f"/leave-schedule?block_tab={effective_block}&sub_tab=roster&month={month}&msg=Đã lưu nháp lịch trực.",
        status_code=303,
    )


@router.post("/roster/approve")
async def approve_roster_plan(
    request: Request,
    block_tab: str = Form(...),
    month: str = Form(...),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    scope_info = _scope_payload(db, user)
    role_codes = _get_role_codes(db, user.id)
    effective_block = block_tab or (scope_info.get("block_code") or "")
    if not _can_approve_block(role_codes, effective_block):
        raise HTTPException(status_code=403, detail="Bạn không có quyền phê duyệt lịch trực.")

    plan_header = _load_roster_plan_header(
        db,
        scope_info=scope_info,
        block_code=effective_block,
        target_month=month,
    )
    if not plan_header:
        raise HTTPException(status_code=404, detail="Chưa có nháp lịch trực để phê duyệt.")
    if str(plan_header.get("status") or "") == "APPROVED":
        raise HTTPException(status_code=400, detail="Lịch trực đã được phê duyệt trước đó.")

    now_str = _dt_to_str(_now_vn())
    db.execute(
        text(
            """
            UPDATE leave_roster_plans
            SET status = 'APPROVED',
                approved_by_id = :approved_by_id,
                approved_by_name = :approved_by_name,
                approved_at = :approved_at,
                updated_at = :updated_at
            WHERE id = :id
            """
        ),
        {
            "id": str(plan_header["id"]),
            "approved_by_id": str(user.id),
            "approved_by_name": user.full_name or user.username or "",
            "approved_at": now_str,
            "updated_at": now_str,
        },
    )

    plan_items = _load_roster_plan_items(db, str(plan_header["id"]))
    target_user_ids = sorted({str(x.get("user_id") or "") for x in plan_items if str(x.get("user_id") or "").strip()})
    for uid in target_user_ids:
        user_items = [x for x in plan_items if str(x.get("user_id") or "") == uid]
        date_list = ", ".join([str(x.get("work_date") or "") for x in user_items])
        _create_leave_notice(
            db,
            user_id=uid,
            block_code=effective_block,
            notice_type="ROSTER_APPROVED",
            title=f"Lịch trực tháng {month} đã được phê duyệt",
            message=f"Lịch trực/làm việc của bạn trong tháng {month}: {date_list}",
            related_request_id=str(plan_header["id"]),
        )

    db.commit()

    await _notify_leave_users(
        db,
        target_user_ids,
        _leave_notify_payload(
            "roster_approved",
            block_code=effective_block,
            target_month=month,
        ),
    )

    return RedirectResponse(
        url=f"/leave-schedule?block_tab={effective_block}&sub_tab=roster&month={month}&msg=Đã phê duyệt lịch trực.",
        status_code=303,
    )

@router.post("/roster/request-adjust")
def request_roster_adjustment(
    request: Request,
    block_tab: str = Form(...),
    month: str = Form(...),
    roster_item_id: str = Form(...),
    new_work_date: str = Form(...),
    reason: str = Form(""),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    scope_info = _scope_payload(db, user)
    effective_block = block_tab or (scope_info.get("block_code") or "")

    plan_header = _load_roster_plan_header(
        db,
        scope_info=scope_info,
        block_code=effective_block,
        target_month=month,
    )
    if not plan_header or str(plan_header.get("status") or "") != "APPROVED":
        raise HTTPException(status_code=400, detail="Chỉ được điều chỉnh khi lịch trực đã được phê duyệt.")

    row = db.execute(
        text(
            """
            SELECT *
            FROM leave_roster_plan_items
            WHERE id = :id AND plan_id = :plan_id
            """
        ),
        {"id": str(roster_item_id), "plan_id": str(plan_header["id"])},
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Không tìm thấy dòng lịch trực cần điều chỉnh.")

    item = _row_to_dict(row)
    _validate_roster_work_date(
        db,
        user_id=str(item.get("user_id") or ""),
        target_month=month,
        work_date_str=new_work_date,
    )

    now_str = _dt_to_str(_now_vn())
    db.execute(
        text(
            """
            INSERT INTO leave_roster_adjustments (
                id, plan_id, roster_item_id, user_id, user_name, old_work_date, new_work_date, reason,
                status, requested_by_id, requested_by_name, created_at, updated_at
            ) VALUES (
                :id, :plan_id, :roster_item_id, :user_id, :user_name, :old_work_date, :new_work_date, :reason,
                'PENDING', :requested_by_id, :requested_by_name, :created_at, :updated_at
            )
            """
        ),
        {
            "id": str(uuid.uuid4()),
            "plan_id": str(plan_header["id"]),
            "roster_item_id": str(item.get("id") or ""),
            "user_id": str(item.get("user_id") or ""),
            "user_name": str(item.get("user_name") or ""),
            "old_work_date": str(item.get("work_date") or ""),
            "new_work_date": new_work_date,
            "reason": reason or "",
            "requested_by_id": str(user.id),
            "requested_by_name": user.full_name or user.username or "",
            "created_at": now_str,
            "updated_at": now_str,
        },
    )
    db.commit()

    return RedirectResponse(
        url=f"/leave-schedule?block_tab={effective_block}&sub_tab=roster&month={month}&msg=Đã gửi yêu cầu điều chỉnh lịch trực.",
        status_code=303,
    )
    
@router.post("/roster/approve-adjust")
async def approve_roster_adjustment(
    request: Request,
    block_tab: str = Form(...),
    month: str = Form(...),
    adjustment_id: str = Form(...),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    scope_info = _scope_payload(db, user)
    role_codes = _get_role_codes(db, user.id)
    effective_block = block_tab or (scope_info.get("block_code") or "")
    if not _can_approve_block(role_codes, effective_block):
        raise HTTPException(status_code=403, detail="Bạn không có quyền phê duyệt điều chỉnh lịch trực.")

    row = db.execute(
        text(
            """
            SELECT *
            FROM leave_roster_adjustments
            WHERE id = :id AND status = 'PENDING'
            """
        ),
        {"id": str(adjustment_id)},
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Không tìm thấy yêu cầu điều chỉnh đang chờ duyệt.")

    item = _row_to_dict(row)
    _validate_roster_work_date(
        db,
        user_id=str(item.get("user_id") or ""),
        target_month=month,
        work_date_str=str(item.get("new_work_date") or ""),
    )

    now_str = _dt_to_str(_now_vn())
    db.execute(
        text(
            """
            UPDATE leave_roster_plan_items
            SET work_date = :work_date,
                note = COALESCE(note, ''),
                updated_at = :updated_at
            WHERE id = :id
            """
        ),
        {
            "id": str(item.get("roster_item_id") or ""),
            "work_date": str(item.get("new_work_date") or ""),
            "updated_at": now_str,
        },
    )
    db.execute(
        text(
            """
            UPDATE leave_roster_adjustments
            SET status = 'APPROVED',
                approved_by_id = :approved_by_id,
                approved_by_name = :approved_by_name,
                approved_at = :approved_at,
                updated_at = :updated_at
            WHERE id = :id
            """
        ),
        {
            "id": str(item.get("id") or ""),
            "approved_by_id": str(user.id),
            "approved_by_name": user.full_name or user.username or "",
            "approved_at": now_str,
            "updated_at": now_str,
        },
    )

    _create_leave_notice(
        db,
        user_id=str(item.get("user_id") or ""),
        block_code=effective_block,
        notice_type="ROSTER_ADJUST_APPROVED",
        title=f"Điều chỉnh lịch trực tháng {month} đã được phê duyệt",
        message=(
            f"Lịch trực của bạn được điều chỉnh từ {item.get('old_work_date') or '-'} "
            f"sang {item.get('new_work_date') or '-'}."
        ),
        related_request_id=str(item.get("plan_id") or ""),
    )
    db.commit()

    await _notify_leave_users(
        db,
        [str(item.get("user_id") or "")],
        _leave_notify_payload(
            "roster_adjust_approved",
            block_code=effective_block,
            target_month=month,
        ),
    )

    return RedirectResponse(
        url=f"/leave-schedule?block_tab={effective_block}&sub_tab=roster&month={month}&msg=Đã phê duyệt điều chỉnh lịch trực.",
        status_code=303,
    )    
    

@router.post("/add")
async def create_leave_request(
    request: Request,
    leave_type: str = Form(...),
    selected_dates: str = Form(""),
    start_date: str = Form(""),
    end_date: str = Form(""),
    reason: str = Form(""),
    block_tab: str = Form(""),
    month: str = Form(""),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    effective_block = block_tab or _default_block_for_user(db, user)
    fallback_month = month or (
        f"{_today_vn().year + 1}-01"
        if _today_vn().month == 12
        else f"{_today_vn().year}-{_today_vn().month + 1:02d}"
    )
    redirect_base = (
        f"/leave-schedule?block_tab={effective_block}&sub_tab=register&month={fallback_month}"
    )

    leave_type = (leave_type or "").strip().upper()
    if leave_type not in LEAVE_TYPE_META:
        raise HTTPException(status_code=400, detail="Loại nghỉ không hợp lệ.")

    selected_list = _parse_selected_dates_csv(selected_dates)

    # ===== Nhánh mới: đăng ký theo nhiều ngày chọn trên lịch =====
    if selected_list:
        first_date = selected_list[0]
        month_key = first_date.strftime("%Y-%m")

        if any(d.strftime("%Y-%m") != month_key for d in selected_list):
            return RedirectResponse(
                url=f"{redirect_base}&error=Chỉ được đăng ký trong cùng một tháng.",
                status_code=302,
            )

        if _is_cutoff_locked(db, user, first_date):
            return RedirectResponse(
                url=f"{redirect_base}&error=Đã qua thời hạn đăng ký trong tháng. Chỉ người có quyền mới được điều chỉnh.",
                status_code=302,
            )

        group_key, group_label, unit, max_off_per_day, deadline_day = _resolve_effective_leave_group(
            db,
            user,
            first_date,
        )

        if leave_type not in LEAVE_TYPES_BYPASS_GROUP_RULES:
            if deadline_day is None:
                is_open, open_message = _is_standard_group_registration_open(
                    db,
                    user,
                    target_date=first_date,
                )
                if not is_open:
                    return RedirectResponse(
                        url=f"{redirect_base}&error={open_message}",
                        status_code=302,
                    )
            else:
                is_open, open_message = _is_priority_registration_day_open(
                    target_date=first_date,
                    deadline_day=deadline_day,
                )
                if not is_open:
                    return RedirectResponse(
                        url=f"{redirect_base}&error={open_message}",
                        status_code=302,
                    )

                blocking_member = _find_blocking_higher_priority_member(
                    db,
                    user=user,
                    target_date=first_date,
                )
                if blocking_member:
                    return RedirectResponse(
                        url=f"{redirect_base}&error={blocking_member['message']}",
                        status_code=302,
                    )

            _validate_quota_for_selected_dates(db, user, leave_type, selected_list)

            conflict = _find_group_conflict_for_selected_dates(
                db,
                user,
                group_key,
                selected_list,
                max_off_per_day=max_off_per_day,
            )
            if conflict:
                row = conflict.get("row") or {}
                detail = conflict["message"]
                if row:
                    detail += (
                        f" Người đã đăng ký: {row.get('user_name') or '-'}; "
                        f"thời gian: {row.get('start_date')} đến {row.get('end_date')}"
                    )
                return RedirectResponse(
                    url=f"{redirect_base}&error={detail}",
                    status_code=302,
                )
        else:
            _validate_quota_for_selected_dates(db, user, leave_type, selected_list)

        now_str = _dt_to_str(_now_vn())
        block_code = _default_block_for_user(db, user)
        symbol = _leave_symbol(leave_type)

        for d in selected_list:
            request_id = str(uuid.uuid4())
            db.execute(
                text(
                    """
                    INSERT INTO leave_requests (
                        id, user_id, user_name, unit_id, unit_name, group_key, group_label,
                        leave_type, symbol, start_date, end_date, day_count, reason,
                        status, block_code, created_at, updated_at
                    ) VALUES (
                        :id, :user_id, :user_name, :unit_id, :unit_name, :group_key, :group_label,
                        :leave_type, :symbol, :start_date, :end_date, :day_count, :reason,
                        'PENDING', :block_code, :created_at, :updated_at
                    )
                    """
                ),
                {
                    "id": request_id,
                    "user_id": user.id,
                    "user_name": user.full_name or user.username,
                    "unit_id": getattr(unit, "id", None),
                    "unit_name": getattr(unit, "ten_don_vi", None),
                    "group_key": group_key,
                    "group_label": group_label,
                    "leave_type": leave_type,
                    "symbol": symbol,
                    "start_date": _date_to_str(d),
                    "end_date": _date_to_str(d),
                    "day_count": 1,
                    "reason": (reason or "").strip(),
                    "block_code": block_code,
                    "created_at": now_str,
                    "updated_at": now_str,
                },
            )

        _create_leave_notice(
            db,
            user_id=str(user.id),
            block_code=block_code,
            notice_type="LEAVE_SUBMITTED",
            title="Đăng ký nghỉ đã được ghi nhận",
            message=f"Bạn đã đăng ký {len(selected_list)} ngày nghỉ loại {_leave_symbol(leave_type)} trong tháng {month_key}.",
        )

        db.commit()

        try:
            notify_user_ids = [str(user.id)]
            notify_user_ids.extend(
                _manager_user_ids_for_leave_scope(
                    db,
                    getattr(unit, "id", None),
                    block_code=block_code,
                )
            )

            await _notify_leave_schedule_users(
                notify_user_ids,
                _leave_schedule_payload(
                    "leave_request_submitted",
                    request_id=request_id,
                    changed_user_id=str(user.id),
                    block_code=block_code,
                    notice_type="LEAVE_SUBMITTED",
                ),
            )
        except Exception as ex:
            logger.exception("[leave_schedule] Lỗi phát realtime sau đăng ký nghỉ nhiều ngày: %s", ex)

        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_code}&sub_tab=register&month={month_key}&msg=Đã ghi nhận đăng ký nghỉ.",
            status_code=302,
        )

    # ===== Nhánh cũ: giữ tương thích nếu form cũ còn gửi start_date/end_date =====
    if not start_date or not end_date:
        return RedirectResponse(
            url=f"{redirect_base}&error=Chưa chọn ngày nghỉ.",
            status_code=302,
        )

    start_dt = _parse_date(start_date, "Ngày bắt đầu")
    end_dt = _parse_date(end_date, "Ngày kết thúc")

    if end_dt < start_dt:
        return RedirectResponse(
            url=f"{redirect_base}&error=Ngày kết thúc không được nhỏ hơn ngày bắt đầu.",
            status_code=302,
        )

    requested_days = _count_workdays(start_dt, end_dt)
    if requested_days <= 0:
        return RedirectResponse(
            url=f"{redirect_base}&error=Khoảng thời gian đăng ký không có ngày công hợp lệ.",
            status_code=302,
        )

    month_key = start_dt.strftime("%Y-%m")

    if _is_cutoff_locked(db, user, start_dt):
        return RedirectResponse(
            url=f"{redirect_base}&error=Đã qua thời hạn đăng ký trong tháng. Chỉ người có quyền mới được điều chỉnh.",
            status_code=302,
        )

    group_key, group_label, unit, max_off_per_day, deadline_day = _resolve_effective_leave_group(
        db,
        user,
        start_dt,
    )

    if leave_type not in LEAVE_TYPES_BYPASS_GROUP_RULES:
        if deadline_day is None:
            is_open, open_message = _is_standard_group_registration_open(
                db,
                user,
                target_date=start_dt,
            )
            if not is_open:
                return RedirectResponse(
                    url=f"{redirect_base}&error={open_message}",
                    status_code=302,
                )
        else:
            is_open, open_message = _is_priority_registration_day_open(
                target_date=start_dt,
                deadline_day=deadline_day,
            )
            if not is_open:
                return RedirectResponse(
                    url=f"{redirect_base}&error={open_message}",
                    status_code=302,
                )

            blocking_member = _find_blocking_higher_priority_member(
                db,
                user=user,
                target_date=start_dt,
            )
            if blocking_member:
                return RedirectResponse(
                    url=f"{redirect_base}&error={blocking_member['message']}",
                    status_code=302,
                )

        _validate_quota(db, user, leave_type, start_dt, end_dt, requested_days)

        conflict = _find_group_conflict(
            db,
            user,
            group_key,
            leave_type,
            start_dt,
            end_dt,
            max_off_per_day=max_off_per_day,
        )
        if conflict:
            row = conflict.get("row") or {}
            detail = conflict["message"]
            if row:
                detail += (
                    f" Người đã đăng ký: {row.get('user_name') or '-'}; "
                    f"thời gian: {row.get('start_date')} đến {row.get('end_date')}"
                )
            return RedirectResponse(
                url=f"{redirect_base}&error={detail}",
                status_code=302,
            )
    else:
        _validate_quota(db, user, leave_type, start_dt, end_dt, requested_days)

    now_str = _dt_to_str(_now_vn())
    request_id = str(uuid.uuid4())
    block_code = _default_block_for_user(db, user)

    db.execute(
        text(
            """
            INSERT INTO leave_requests (
                id, user_id, user_name, unit_id, unit_name, group_key, group_label,
                leave_type, symbol, start_date, end_date, day_count, reason,
                status, block_code, created_at, updated_at
            ) VALUES (
                :id, :user_id, :user_name, :unit_id, :unit_name, :group_key, :group_label,
                :leave_type, :symbol, :start_date, :end_date, :day_count, :reason,
                'PENDING', :block_code, :created_at, :updated_at
            )
            """
        ),
        {
            "id": request_id,
            "user_id": user.id,
            "user_name": user.full_name or user.username,
            "unit_id": getattr(unit, "id", None),
            "unit_name": getattr(unit, "ten_don_vi", None),
            "group_key": group_key,
            "group_label": group_label,
            "leave_type": leave_type,
            "symbol": _leave_symbol(leave_type),
            "start_date": _date_to_str(start_dt),
            "end_date": _date_to_str(end_dt),
            "day_count": requested_days,
            "reason": (reason or "").strip(),
            "block_code": block_code,
            "created_at": now_str,
            "updated_at": now_str,
        },
    )

    _create_leave_notice(
        db,
        user_id=str(user.id),
        block_code=block_code,
        notice_type="LEAVE_SUBMITTED",
        title="Đăng ký nghỉ đã được ghi nhận",
        message=f"Bạn đã đăng ký nghỉ từ {_date_to_str(start_dt)} đến {_date_to_str(end_dt)}.",
        related_request_id=request_id,
    )

    db.commit()

    try:
        notify_user_ids = [str(user.id)]
        notify_user_ids.extend(
            _manager_user_ids_for_leave_scope(
                db,
                getattr(unit, "id", None),
                block_code=block_code,
            )
        )

        await _notify_leave_schedule_users(
            notify_user_ids,
            _leave_schedule_payload(
                "leave_request_submitted",
                request_id=request_id,
                changed_user_id=str(user.id),
                block_code=block_code,
                notice_type="LEAVE_SUBMITTED",
            ),
        )
    except Exception as ex:
        logger.exception("[leave_schedule] Lỗi phát realtime sau đăng ký nghỉ: %s", ex)

    return RedirectResponse(
        url=f"/leave-schedule?block_tab={block_code}&sub_tab=register&month={month_key}&msg=Đã ghi nhận đăng ký nghỉ.",
        status_code=302,
    )

@router.post("/approve")
def approve_leave_request(
    request: Request,
    request_id: str = Form(...),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)
    row = db.execute(text("SELECT * FROM leave_requests WHERE id = :id"), {"id": request_id}).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Không tìm thấy đăng ký nghỉ.")

    item = _row_to_dict(row)
    request_block_code = str(item.get("block_code") or "").strip() or _default_block_for_user(db, user)

    if not _can_approve(db, user, request_block_code):
        raise HTTPException(status_code=403, detail="Bạn không có quyền phê duyệt.")
    if not _can_access_leave_request_scope(db, user, item):
        raise HTTPException(status_code=403, detail="Phiếu nghỉ không thuộc phạm vi đơn vị được phép phê duyệt.")

    if item.get("status") != "PENDING":
        return RedirectResponse(url="/leave-schedule?error=Phiếu này đã được xử lý.", status_code=302)

    now_str = _dt_to_str(_now_vn())
    db.execute(
        text(
            """
            UPDATE leave_requests
            SET status = 'APPROVED',
                approved_by_id = :approved_by_id,
                approved_by_name = :approved_by_name,
                approved_role = :approved_role,
                approved_at = :approved_at,
                updated_at = :updated_at
            WHERE id = :id AND status = 'PENDING'
            """
        ),
        {
            "id": request_id,
            "approved_by_id": user.id,
            "approved_by_name": user.full_name or user.username,
            "approved_role": _role_label_for_user(db, user),
            "approved_at": now_str,
            "updated_at": now_str,
        },
    )

    requester_id = str(item.get("user_id") or "")
    block_code = item.get("block_code") or _default_block_for_user(db, user)

    if requester_id:
        _create_leave_notice(
            db,
            user_id=requester_id,
            block_code=block_code,
            notice_type="LEAVE_REQUEST_APPROVED",
            title="Đăng ký nghỉ đã được duyệt",
            message=(
                f"Phiếu nghỉ từ {item.get('start_date')} đến {item.get('end_date')} "
                f"đã được {user.full_name or user.username} phê duyệt."
            ),
            related_request_id=request_id,
        )

    db.commit()

    payload = _leave_notify_payload(
        "leave_request_approved",
        request_id=request_id,
        block_code=block_code,
        approved_by=user.full_name or user.username,
        requester_id=requester_id,
    )

    return RedirectResponse(
        url=f"/leave-schedule?block_tab={block_code}&sub_tab=register&list_mode=pending&msg=Đã phê duyệt đăng ký nghỉ.",
        status_code=302,
    )

@router.post("/reject")
def reject_leave_request(
    request: Request,
    request_id: str = Form(...),
    reject_reason: str = Form(""),
    block_tab: str = Form(""),
    month: str = Form(""),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    row = db.execute(
        text("SELECT * FROM leave_requests WHERE id = :id"),
        {"id": request_id},
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Không tìm thấy đăng ký nghỉ.")

    item = _row_to_dict(row)
    request_block_code = (
        block_tab
        or str(item.get("block_code") or "").strip()
        or _default_block_for_user(db, user)
    ).strip()

    fallback_month = (
        f"{_today_vn().year + 1}-01"
        if _today_vn().month == 12
        else f"{_today_vn().year}-{_today_vn().month + 1:02d}"
    )

    if not _can_approve(db, user, request_block_code):
        raise HTTPException(status_code=403, detail="Bạn không có quyền từ chối phiếu nghỉ này.")
    if not _can_access_leave_request_scope(db, user, item):
        raise HTTPException(status_code=403, detail="Phiếu nghỉ không thuộc phạm vi đơn vị được phép từ chối.")

    if item.get("status") != "PENDING":
        return RedirectResponse(
            url=(
                f"/leave-schedule?block_tab={request_block_code}"
                f"&sub_tab=register"
                f"&month={month or fallback_month}"
                f"&error=Chỉ được từ chối phiếu đang ở trạng thái Chờ duyệt."
            ),
            status_code=302,
        )

    reject_reason = (reject_reason or "").strip()
    now_str = _dt_to_str(_now_vn())

    db.execute(
        text(
            """
            UPDATE leave_requests
            SET status = 'REJECTED',
                reject_reason = :reject_reason,
                approved_by_id = :approved_by_id,
                approved_by_name = :approved_by_name,
                approved_role = :approved_role,
                approved_at = :approved_at,
                updated_at = :updated_at
            WHERE id = :id
            """
        ),
        {
            "id": request_id,
            "reject_reason": reject_reason,
            "approved_by_id": user.id,
            "approved_by_name": user.full_name or user.username,
            "approved_role": _role_label_for_user(db, user),
            "approved_at": now_str,
            "updated_at": now_str,
        },
    )

    requester_id = str(item.get("user_id") or "")
    block_code = request_block_code

    if requester_id:
        date_label = str(item.get("start_date") or "")
        if str(item.get("start_date") or "") != str(item.get("end_date") or ""):
            date_label = f"{item.get('start_date')} đến {item.get('end_date')}"
        detail_reason = f" Lý do: {reject_reason}." if reject_reason else ""

        _create_leave_notice(
            db,
            user_id=requester_id,
            block_code=block_code,
            notice_type="LEAVE_REQUEST_REJECTED",
            title="Đăng ký nghỉ bị từ chối",
            message=(
                f"Phiếu nghỉ ngày {date_label} đã bị "
                f"{user.full_name or user.username} từ chối.{detail_reason}"
            ),
            related_request_id=request_id,
        )

    db.commit()

    return RedirectResponse(
        url=(
            f"/leave-schedule?block_tab={block_code}"
            f"&sub_tab=register"
            f"&month={month or fallback_month}"
            f"&msg=Đã từ chối đăng ký nghỉ."
        ),
        status_code=302,
    )


@router.post("/approve-bulk")
def approve_leave_requests_bulk(
    request: Request,
    block_tab: str = Form(...),
    month: str = Form(""),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    if block_tab not in {"HANH_CHINH", "CHUYEN_MON"}:
        raise HTTPException(status_code=400, detail="Khối không hợp lệ.")
    if not _can_approve(db, user, block_tab):
        raise HTTPException(status_code=403, detail="Bạn không có quyền phê duyệt hàng loạt.")

    scope_info = _scope_payload(db, user)
    unit_ids = _scope_unit_ids_for_view(scope_info)
    if not unit_ids:
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=register&month={month or _today_vn().strftime('%Y-%m')}&error=Không xác định được phạm vi đơn vị để phê duyệt.",
            status_code=302,
        )

    if month:
        month_key = month
    else:
        today = _today_vn()
        if today.month == 12:
            month_key = f"{today.year + 1}-01"
        else:
            month_key = f"{today.year}-{today.month + 1:02d}"
            
    year, month_num = [int(x) for x in month_key.split("-")]
    first_day = date(year, month_num, 1)
    last_day = (
        date(year + (1 if month_num == 12 else 0), 1 if month_num == 12 else month_num + 1, 1)
        - timedelta(days=1)
    )

    unit_placeholders = []
    params: dict[str, Any] = {
        "first_day": first_day.isoformat(),
        "last_day": last_day.isoformat(),
        "approved_by_id": user.id,
        "approved_by_name": user.full_name or user.username,
        "approved_role": _role_label_for_user(db, user),
        "approved_at": _dt_to_str(_now_vn()),
        "updated_at": _dt_to_str(_now_vn()),
    }

    for idx, unit_id in enumerate(unit_ids):
        key = f"unit_id_{idx}"
        unit_placeholders.append(f":{key}")
        params[key] = str(unit_id)

    rows = db.execute(
        text(
            f"""
            SELECT id, user_id, start_date, end_date, block_code
            FROM leave_requests
            WHERE unit_id IN ({', '.join(unit_placeholders)})
              AND status = 'PENDING'
              AND start_date <= :last_day
              AND end_date >= :first_day
            ORDER BY user_name ASC, start_date ASC, created_at ASC
            """
        ),
        params,
    ).fetchall()

    items = [_row_to_dict(r) for r in rows]
    if not items:
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=register&month={month_key}&error=Không có phiếu chờ duyệt để phê duyệt chung.",
            status_code=302,
        )

    notify_user_ids: list[str] = []

    for item in items:
        db.execute(
            text(
                """
                UPDATE leave_requests
                SET status = 'APPROVED',
                    approved_by_id = :approved_by_id,
                    approved_by_name = :approved_by_name,
                    approved_role = :approved_role,
                    approved_at = :approved_at,
                    updated_at = :updated_at
                WHERE id = :id AND status = 'PENDING'
                """
            ),
            {
                "id": item["id"],
                "approved_by_id": params["approved_by_id"],
                "approved_by_name": params["approved_by_name"],
                "approved_role": params["approved_role"],
                "approved_at": params["approved_at"],
                "updated_at": params["updated_at"],
            },
        )

        requester_id = str(item.get("user_id") or "")
        if requester_id:
            row_block_code = str(item.get("block_code") or block_tab)
            _create_leave_notice(
                db,
                user_id=requester_id,
                block_code=row_block_code,
                notice_type="LEAVE_REQUEST_APPROVED",
                title="Đăng ký nghỉ đã được duyệt",
                message=(
                    f"Phiếu nghỉ từ {item.get('start_date')} đến {item.get('end_date')} "
                    f"đã được {user.full_name or user.username} phê duyệt."
                ),
                related_request_id=item["id"],
            )
            if requester_id not in notify_user_ids:
                notify_user_ids.append(requester_id)
    db.commit()

    if notify_user_ids:
        _emit_leave_schedule_realtime_sync(
            notify_user_ids,
            _leave_notify_payload(
                "leave_requests_approved",
                block_code=block_tab,
                target_month=month_key,
            ),
        )

    return RedirectResponse(
        url=f"/leave-schedule?block_tab={block_tab}&sub_tab=register&month={month_key}&msg=Đã phê duyệt chung các phiếu nghỉ chờ duyệt trong tháng.",
        status_code=302,
    )
    

@router.post("/cancel")
def cancel_leave_request(
    request: Request,
    request_id: str = Form(...),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)
    row = db.execute(text("SELECT * FROM leave_requests WHERE id = :id"), {"id": request_id}).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Không tìm thấy đăng ký nghỉ.")

    item = _row_to_dict(row)
    if item.get("status") != "PENDING":
        return RedirectResponse(url="/leave-schedule?error=Chỉ phiếu đang chờ duyệt mới được hủy.", status_code=302)

    can_cancel = item.get("user_id") == user.id or _can_adjust_after_cutoff(db, user) or _can_admin_extra_days(db, user)
    if not can_cancel:
        raise HTTPException(status_code=403, detail="Bạn không có quyền hủy phiếu này.")

    db.execute(
        text("UPDATE leave_requests SET status = 'CANCELLED', updated_at = :updated_at WHERE id = :id"),
        {"id": request_id, "updated_at": _dt_to_str(_now_vn())},
    )

    requester_id = str(item.get("user_id") or "")
    block_code = item.get("block_code") or _default_block_for_user(db, user)

    if requester_id and requester_id != str(user.id):
        _create_leave_notice(
            db,
            user_id=requester_id,
            block_code=block_code,
            notice_type="LEAVE_REQUEST_CANCELLED",
            title="Đăng ký nghỉ đã bị hủy",
            message=(
                f"Phiếu nghỉ từ {item.get('start_date')} đến {item.get('end_date')} đã được hủy."
            ),
            related_request_id=request_id,
        )

    db.commit()

    payload = _leave_notify_payload(
        "leave_request_cancelled",
        request_id=request_id,
        block_code=block_code,
        requester_id=requester_id,
        cancelled_by=user.full_name or user.username,
    )

    return RedirectResponse(
        url=f"/leave-schedule?block_tab={block_code}&sub_tab=register&msg=Đã hủy đăng ký nghỉ.",
        status_code=302,
    )


@router.post("/swap-request")
def create_leave_swap_request(
    request: Request,
    request_a_id: str = Form(...),
    request_b_id: str = Form(...),
    swap_note: str = Form(""),
    block_tab: str = Form(...),
    month: str = Form(""),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    month_key = month or (
        f"{_today_vn().year + 1}-01"
        if _today_vn().month == 12
        else f"{_today_vn().year}-{_today_vn().month + 1:02d}"
    )

    if block_tab not in {"HANH_CHINH", "CHUYEN_MON"}:
        raise HTTPException(status_code=400, detail="Khối không hợp lệ.")

    role_codes = _get_role_codes(db, user.id)
    if not _can_approve_block(role_codes, block_tab):
        raise HTTPException(status_code=403, detail="Bạn không có quyền điều chỉnh lịch nghỉ.")

    row_a = db.execute(
        text("SELECT * FROM leave_requests WHERE id = :id"),
        {"id": request_a_id},
    ).fetchone()
    row_b = db.execute(
        text("SELECT * FROM leave_requests WHERE id = :id"),
        {"id": request_b_id},
    ).fetchone()

    if not row_a or not row_b:
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=register&month={month_key}&error=Không tìm thấy 1 trong 2 phiếu nghỉ cần hoán đổi.",
            status_code=302,
        )

    a = _row_to_dict(row_a)
    b = _row_to_dict(row_b)

    if a.get("status") != "PENDING" or b.get("status") != "PENDING":
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=register&month={month_key}&error=Chỉ được điều chỉnh/hoán đổi các phiếu đang chờ duyệt.",
            status_code=302,
        )
        
    if str(request_a_id) == str(request_b_id):
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=register&month={month_key}&error=Hai phiếu nghỉ dùng để hoán đổi phải là hai phiếu khác nhau.",
            status_code=302,
        )    

    scope_info = _scope_payload(db, user)

    db.execute(
        text(
            """
            INSERT INTO leave_swap_requests (
                id, block_code, scope_unit_id, scope_subunit_id,
                user_a_id, user_a_name, request_a_id,
                user_b_id, user_b_name, request_b_id,
                swap_note, status,
                created_by_id, created_by_name,
                created_at, updated_at
            ) VALUES (
                :id, :block_code, :scope_unit_id, :scope_subunit_id,
                :user_a_id, :user_a_name, :request_a_id,
                :user_b_id, :user_b_name, :request_b_id,
                :swap_note, 'PENDING',
                :created_by_id, :created_by_name,
                :created_at, :updated_at
            )
            """
        ),
        {
            "id": str(uuid.uuid4()),
            "block_code": block_tab,
            "scope_unit_id": scope_info.get("scope_unit_id"),
            "scope_subunit_id": scope_info.get("scope_subunit_id"),
            "user_a_id": a.get("user_id"),
            "user_a_name": a.get("user_name"),
            "request_a_id": request_a_id,
            "user_b_id": b.get("user_id"),
            "user_b_name": b.get("user_name"),
            "request_b_id": request_b_id,
            "swap_note": (swap_note or "").strip(),
            "created_by_id": str(user.id),
            "created_by_name": user.full_name or user.username,
            "created_at": _dt_to_str(_now_vn()),
            "updated_at": _dt_to_str(_now_vn()),
        },
    )
    db.commit()

    return RedirectResponse(
        url=f"/leave-schedule?block_tab={block_tab}&sub_tab=register&month={month_key}&msg=Đã tạo yêu cầu hoán đổi lịch nghỉ, chờ phê duyệt.",
        status_code=302,
    )

@router.post("/groups/create-priority-parent")
def create_priority_parent_group(
    request: Request,
    parent_name: str = Form(...),
    priority_base_day: int = Form(...),
    max_off_per_day: int = Form(...),
    block_tab: str = Form(...),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    if block_tab not in {"HANH_CHINH", "CHUYEN_MON"}:
        raise HTTPException(status_code=400, detail="Khối không hợp lệ.")

    role_codes = _get_role_codes(db, user.id)
    if not _can_manage_block(role_codes, block_tab):
        raise HTTPException(status_code=403, detail="Bạn không có quyền tạo cơ chế nhóm ưu tiên.")

    if not (1 <= int(priority_base_day) <= 31):
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&error=Ngày chốt nhóm ưu tiên 1 phải từ 1 đến 31.",
            status_code=302,
        )

    if int(max_off_per_day) <= 0:
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&error=Số người nghỉ tối đa trong 1 ngày phải lớn hơn 0.",
            status_code=302,
        )

    scope_info = _scope_payload(db, user)
    existing = db.execute(
        text(
            """
            SELECT id
            FROM leave_group_configs
            WHERE block_code = :block_code
              AND COALESCE(scope_unit_id, '') = :scope_unit_id
              AND COALESCE(scope_subunit_id, '') = :scope_subunit_id
              AND group_type = :group_type
              AND group_name = :group_name
            """
        ),
        {
            "block_code": block_tab,
            "scope_unit_id": scope_info.get("scope_unit_id") or "",
            "scope_subunit_id": scope_info.get("scope_subunit_id") or "",
            "group_type": PRIORITY_PARENT_GROUP_TYPE,
            "group_name": (parent_name or "").strip(),
        },
    ).fetchone()
    if existing:
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&error=Nhóm cha này đã tồn tại.",
            status_code=302,
        )

    now_str = _dt_to_str(_now_vn())
    db.execute(
        text(
            """
            INSERT INTO leave_group_configs (
                id, block_code, scope_unit_id, scope_unit_name, scope_subunit_id, scope_subunit_name,
                group_type, group_name, max_off_per_day, is_active, created_by_id, created_by_name,
                created_at, updated_at, parent_group_id, priority_base_day, priority_order
            ) VALUES (
                :id, :block_code, :scope_unit_id, :scope_unit_name, :scope_subunit_id, :scope_subunit_name,
                :group_type, :group_name, :max_off_per_day, 1, :created_by_id, :created_by_name,
                :created_at, :updated_at, NULL, :priority_base_day, 0
            )
            """
        ),
        {
            "id": str(uuid.uuid4()),
            "block_code": block_tab,
            "scope_unit_id": scope_info.get("scope_unit_id"),
            "scope_unit_name": scope_info.get("scope_unit_name"),
            "scope_subunit_id": scope_info.get("scope_subunit_id"),
            "scope_subunit_name": scope_info.get("scope_subunit_name"),
            "group_type": PRIORITY_PARENT_GROUP_TYPE,
            "group_name": (parent_name or "").strip(),
            "max_off_per_day": int(max_off_per_day),
            "created_by_id": user.id,
            "created_by_name": user.full_name or user.username,
            "created_at": now_str,
            "updated_at": now_str,
            "priority_base_day": int(priority_base_day),
        },
    )
    db.commit()

    return RedirectResponse(
        url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&msg=Đã tạo nhóm cha cơ chế ưu tiên.",
        status_code=302,
    )


@router.post("/groups/create-priority-child")
def create_priority_child_group(
    request: Request,
    parent_group_id: str = Form(...),
    child_name: str = Form(...),
    block_tab: str = Form(...),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    if block_tab not in {"HANH_CHINH", "CHUYEN_MON"}:
        raise HTTPException(status_code=400, detail="Khối không hợp lệ.")

    role_codes = _get_role_codes(db, user.id)
    if not _can_manage_block(role_codes, block_tab):
        raise HTTPException(status_code=403, detail="Bạn không có quyền tạo nhóm con ưu tiên.")

    parent_row = db.execute(
        text(
            """
            SELECT *
            FROM leave_group_configs
            WHERE id = :id
              AND group_type = :group_type
              AND block_code = :block_code
            """
        ),
        {"id": parent_group_id, "group_type": PRIORITY_PARENT_GROUP_TYPE, "block_code": block_tab},
    ).fetchone()
    if not parent_row:
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&error=Không tìm thấy nhóm cha cơ chế ưu tiên.",
            status_code=302,
        )
    parent_item = _row_to_dict(parent_row)

    existed = db.execute(
        text(
            """
            SELECT id
            FROM leave_group_configs
            WHERE parent_group_id = :parent_group_id
              AND group_type = :group_type
              AND group_name = :group_name
            """
        ),
        {
            "parent_group_id": parent_group_id,
            "group_type": PRIORITY_CHILD_GROUP_TYPE,
            "group_name": (child_name or "").strip(),
        },
    ).fetchone()
    if existed:
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&error=Tên nhóm con đã tồn tại trong cơ chế này.",
            status_code=302,
        )

    count_row = db.execute(
        text(
            """
            SELECT COUNT(*) AS total_count
            FROM leave_group_configs
            WHERE parent_group_id = :parent_group_id
              AND group_type = :group_type
            """
        ),
        {"parent_group_id": parent_group_id, "group_type": PRIORITY_CHILD_GROUP_TYPE},
    ).fetchone()
    next_order = int(_row_to_dict(count_row).get("total_count") or 0) + 1

    now_str = _dt_to_str(_now_vn())
    db.execute(
        text(
            """
            INSERT INTO leave_group_configs (
                id, block_code, scope_unit_id, scope_unit_name, scope_subunit_id, scope_subunit_name,
                group_type, group_name, max_off_per_day, is_active, created_by_id, created_by_name,
                created_at, updated_at, parent_group_id, priority_base_day, priority_order
            ) VALUES (
                :id, :block_code, :scope_unit_id, :scope_unit_name, :scope_subunit_id, :scope_subunit_name,
                :group_type, :group_name, 0, 1, :created_by_id, :created_by_name,
                :created_at, :updated_at, :parent_group_id, 0, :priority_order
            )
            """
        ),
        {
            "id": str(uuid.uuid4()),
            "block_code": block_tab,
            "scope_unit_id": parent_item.get("scope_unit_id"),
            "scope_unit_name": parent_item.get("scope_unit_name"),
            "scope_subunit_id": parent_item.get("scope_subunit_id"),
            "scope_subunit_name": parent_item.get("scope_subunit_name"),
            "group_type": PRIORITY_CHILD_GROUP_TYPE,
            "group_name": (child_name or "").strip(),
            "created_by_id": user.id,
            "created_by_name": user.full_name or user.username,
            "created_at": now_str,
            "updated_at": now_str,
            "parent_group_id": parent_group_id,
            "priority_order": next_order,
        },
    )
    db.commit()

    return RedirectResponse(
        url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&msg=Đã tạo nhóm con ưu tiên.",
        status_code=302,
    )


@router.post("/groups/update-priority-child")
def update_priority_child_group(
    request: Request,
    child_group_id: str = Form(...),
    block_tab: str = Form(...),
    custom_max_off_per_day: int = Form(...),
    custom_deadline_day: int = Form(...),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    if block_tab not in {"HANH_CHINH", "CHUYEN_MON"}:
        raise HTTPException(status_code=400, detail="Khối không hợp lệ.")

    role_codes = _get_role_codes(db, user.id)
    if not _can_manage_block(role_codes, block_tab):
        raise HTTPException(status_code=403, detail="Bạn không có quyền cập nhật nhóm con.")

    if custom_max_off_per_day < 0:
        custom_max_off_per_day = 0
    if custom_deadline_day < 1 or custom_deadline_day > 31:
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&error=Ngày chốt nhóm con phải từ 1 đến 31.",
            status_code=302,
        )

    row = db.execute(
        text(
            """
            SELECT id
            FROM leave_group_configs
            WHERE id = :id
              AND group_type = :group_type
              AND block_code = :block_code
            """
        ),
        {
            "id": child_group_id,
            "group_type": PRIORITY_CHILD_GROUP_TYPE,
            "block_code": block_tab,
        },
    ).fetchone()
    if not row:
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&error=Không tìm thấy nhóm con.",
            status_code=302,
        )

    db.execute(
        text(
            """
            UPDATE leave_group_configs
            SET custom_max_off_per_day = :custom_max_off_per_day,
                custom_deadline_day = :custom_deadline_day,
                updated_at = :updated_at
            WHERE id = :id
            """
        ),
        {
            "id": child_group_id,
            "custom_max_off_per_day": custom_max_off_per_day,
            "custom_deadline_day": custom_deadline_day,
            "updated_at": _dt_to_str(_now_vn()),
        },
    )
    db.commit()

    return RedirectResponse(
        url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&msg=Đã cập nhật hạn mức nghỉ và thời hạn chốt của nhóm con.",
        status_code=302,
    )

@router.post("/groups/create")
def create_leave_group(
    request: Request,
    group_type: str = Form(...),
    block_tab: str = Form(...),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    if block_tab not in {"HANH_CHINH", "CHUYEN_MON"}:
        raise HTTPException(status_code=400, detail="Khối không hợp lệ.")

    role_codes = _get_role_codes(db, user.id)
    if not _can_manage_block(role_codes, block_tab):
        raise HTTPException(status_code=403, detail="Bạn không có quyền tạo nhóm trong khối này.")

    scope_info = _scope_payload(db, user)
    scope_unit_id = scope_info.get("scope_unit_id")
    scope_subunit_id = scope_info.get("scope_subunit_id")

    def_map = _group_def_map(block_tab)
    group_meta = def_map.get(group_type)
    if not group_meta:
        raise HTTPException(status_code=400, detail="Loại nhóm không hợp lệ.")

    existing = db.execute(
        text(
            """
            SELECT id
            FROM leave_group_configs
            WHERE block_code = :block_code
              AND COALESCE(scope_unit_id, '') = :scope_unit_id
              AND COALESCE(scope_subunit_id, '') = :scope_subunit_id
              AND group_type = :group_type
              AND is_active = 1
            """
        ),
        {
            "block_code": block_tab,
            "scope_unit_id": scope_unit_id or "",
            "scope_subunit_id": scope_subunit_id or "",
            "group_type": group_type,
        },
    ).fetchone()
    if existing:
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&error=Nhóm này đã tồn tại trong phạm vi hiện hành.",
            status_code=302,
        )

    now_str = _dt_to_str(_now_vn())
    new_group_id = str(uuid.uuid4())
    db.execute(
        text(
            """
            INSERT INTO leave_group_configs (
                id, block_code, scope_unit_id, scope_unit_name, scope_subunit_id, scope_subunit_name,
                group_type, group_name, max_off_per_day, is_active,
                created_by_id, created_by_name, created_at, updated_at
            ) VALUES (
                :id, :block_code, :scope_unit_id, :scope_unit_name, :scope_subunit_id, :scope_subunit_name,
                :group_type, :group_name, :max_off_per_day, 1,
                :created_by_id, :created_by_name, :created_at, :updated_at
            )
            """
        ),
        {
            "id": new_group_id,
            "block_code": block_tab,
            "scope_unit_id": scope_unit_id,
            "scope_unit_name": scope_info.get("scope_unit_name"),
            "scope_subunit_id": scope_subunit_id,
            "scope_subunit_name": scope_info.get("scope_subunit_name"),
            "group_type": group_type,
            "group_name": group_meta["label"],
            "max_off_per_day": int(group_meta["max_off_per_day"]),
            "created_by_id": user.id,
            "created_by_name": user.full_name or user.username,
            "created_at": now_str,
            "updated_at": now_str,
        },
    )
    _sync_manager_group_max_off_per_day(db, new_group_id)    
    db.commit()

    return RedirectResponse(
        url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&msg=Đã thành lập nhóm.",
        status_code=302,
    )


@router.post("/groups/add-member")
def add_leave_group_member(
    request: Request,
    group_id: str = Form(...),
    user_id: str = Form(...),
    block_tab: str = Form(...),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    if block_tab not in {"HANH_CHINH", "CHUYEN_MON"}:
        raise HTTPException(status_code=400, detail="Khối không hợp lệ.")

    role_codes = _get_role_codes(db, user.id)
    if not _can_manage_block(role_codes, block_tab):
        raise HTTPException(status_code=403, detail="Bạn không có quyền cập nhật nhóm trong khối này.")

    group_row = db.execute(
        text("SELECT * FROM leave_group_configs WHERE id = :id AND is_active = 1"),
        {"id": group_id},
    ).fetchone()
    if not group_row:
        raise HTTPException(status_code=404, detail="Không tìm thấy nhóm.")

    group_item = _row_to_dict(group_row)
    scope_unit_id = group_item.get("scope_unit_id")
    scope_subunit_id = group_item.get("scope_subunit_id")
    group_type = str(group_item.get("group_type") or "")

    target_user = db.get(Users, user_id)
    if not target_user:
        raise HTTPException(status_code=404, detail="Không tìm thấy người dùng.")

    scope_info = _scope_payload(db, user)
    scope_users = _load_scope_users_for_groups(db, scope_info)
    scope_user_ids = {str(x["id"]) for x in scope_users}
    if str(user_id) not in scope_user_ids:
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&error=Người dùng không thuộc phạm vi đơn vị được phép thêm.",
            status_code=302,
        )

    target_role_codes = _get_role_codes(db, str(user_id))
    if not _is_user_eligible_for_group(group_type, target_role_codes):
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&error=Người dùng không đúng vị trí được phép của nhóm này.",
            status_code=302,
        )

    existed_same_group = db.execute(
        text(
            """
            SELECT id
            FROM leave_group_members
            WHERE group_id = :group_id AND user_id = :user_id
            """
        ),
        {"group_id": group_id, "user_id": str(user_id)},
    ).fetchone()
    if existed_same_group:
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&error=Người dùng đã có trong nhóm này.",
            status_code=302,
        )

    existed_other_group = db.execute(
        text(
            """
            SELECT m.id
            FROM leave_group_members m
            JOIN leave_group_configs g ON g.id = m.group_id
            WHERE m.user_id = :user_id
              AND g.block_code = :block_code
              AND COALESCE(g.scope_unit_id, '') = :scope_unit_id
              AND COALESCE(g.scope_subunit_id, '') = :scope_subunit_id
              AND g.is_active = 1
            """
        ),
        {
            "user_id": str(user_id),
            "block_code": block_tab,
            "scope_unit_id": scope_unit_id or "",
            "scope_subunit_id": scope_subunit_id or "",
        },
    ).fetchone()
    if existed_other_group:
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&error=Mỗi người chỉ được thuộc 1 nhóm trong cùng phạm vi khối này.",
            status_code=302,
        )

    order_row = db.execute(
        text(
            """
            SELECT COALESCE(MAX(member_order), 0) AS max_order
            FROM leave_group_members
            WHERE group_id = :group_id
            """
        ),
        {"group_id": group_id},
    ).fetchone()
    next_member_order = int(_row_to_dict(order_row).get("max_order") or 0) + 1 if order_row else 1

    db.execute(
        text(
            """
            INSERT INTO leave_group_members (id, group_id, user_id, user_name, created_at, member_order)
            VALUES (:id, :group_id, :user_id, :user_name, :created_at, :member_order)
            """
        ),
        {
            "id": str(uuid.uuid4()),
            "group_id": group_id,
            "user_id": str(user_id),
            "user_name": target_user.full_name or target_user.username,
            "created_at": _dt_to_str(_now_vn()),
            "member_order": next_member_order,
        },
    )
    _sync_manager_group_max_off_per_day(db, group_id)
    db.commit()

    return RedirectResponse(
        url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&msg=Đã thêm người dùng vào nhóm.",
        status_code=302,
    )


@router.post("/groups/remove-member")
def remove_leave_group_member(
    request: Request,
    member_id: str = Form(...),
    block_tab: str = Form(...),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    if block_tab not in {"HANH_CHINH", "CHUYEN_MON"}:
        raise HTTPException(status_code=400, detail="Khối không hợp lệ.")

    role_codes = _get_role_codes(db, user.id)
    if not _can_manage_block(role_codes, block_tab):
        raise HTTPException(status_code=403, detail="Bạn không có quyền cập nhật nhóm trong khối này.")

    row = db.execute(
        text(
            """
            SELECT m.id, m.group_id
            FROM leave_group_members m
            JOIN leave_group_configs g ON g.id = m.group_id
            WHERE m.id = :member_id
              AND g.block_code = :block_code
            """
        ),
        {"member_id": member_id, "block_code": block_tab},
    ).fetchone()
    if not row:
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&error=Không tìm thấy thành viên nhóm.",
            status_code=302,
        )

    db.execute(
        text("DELETE FROM leave_group_members WHERE id = :member_id"),
        {"member_id": member_id},
    )
    if row and "group_id" in _row_to_dict(row):
        _sync_manager_group_max_off_per_day(db, _row_to_dict(row).get("group_id"))
    db.commit()

    return RedirectResponse(
        url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&msg=Đã gỡ người dùng khỏi nhóm.",
        status_code=302,
    )


@router.post("/groups/toggle-lock")
def toggle_leave_group_lock(
    request: Request,
    group_id: str = Form(...),
    block_tab: str = Form(...),
    action: str = Form(...),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    if block_tab not in {"HANH_CHINH", "CHUYEN_MON"}:
        raise HTTPException(status_code=400, detail="Khối không hợp lệ.")

    role_codes = _get_role_codes(db, user.id)
    if not _can_manage_block(role_codes, block_tab):
        raise HTTPException(status_code=403, detail="Bạn không có quyền khóa/mở khóa nhóm trong khối này.")

    action = (action or "").strip().lower()
    if action not in {"lock", "unlock"}:
        raise HTTPException(status_code=400, detail="Thao tác không hợp lệ.")

    row = db.execute(
        text("SELECT id, is_active FROM leave_group_configs WHERE id = :group_id AND block_code = :block_code"),
        {"group_id": group_id, "block_code": block_tab},
    ).fetchone()
    if not row:
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&error=Không tìm thấy nhóm.",
            status_code=302,
        )

    is_active = int(_row_to_dict(row).get("is_active") or 0)
    target_active = 0 if action == "lock" else 1

    if is_active == target_active:
        msg = "Nhóm đã ở trạng thái mong muốn."
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&msg={msg}",
            status_code=302,
        )

    db.execute(
        text(
            """
            UPDATE leave_group_configs
            SET is_active = :is_active,
                updated_at = :updated_at
            WHERE id = :group_id
            """
        ),
        {
            "group_id": group_id,
            "is_active": target_active,
            "updated_at": _dt_to_str(_now_vn()),
        },
    )
    
    db.execute(
        text(
            """
            UPDATE leave_group_configs
            SET is_active = :is_active,
                updated_at = :updated_at
            WHERE parent_group_id = :group_id
            """
        ),
        {
            "group_id": group_id,
            "is_active": target_active,
            "updated_at": _dt_to_str(_now_vn()),
        },
    )
    db.commit()

    msg = "Đã khóa nhóm." if action == "lock" else "Đã mở khóa nhóm."
    return RedirectResponse(
        url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&msg={msg}",
        status_code=302,
    )


@router.post("/groups/delete")
def delete_leave_group(
    request: Request,
    group_id: str = Form(...),
    block_tab: str = Form(...),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    if block_tab not in {"HANH_CHINH", "CHUYEN_MON"}:
        raise HTTPException(status_code=400, detail="Khối không hợp lệ.")

    role_codes = _get_role_codes(db, user.id)
    if not _can_manage_block(role_codes, block_tab):
        raise HTTPException(status_code=403, detail="Bạn không có quyền xóa nhóm trong khối này.")

    row = db.execute(
        text("SELECT id FROM leave_group_configs WHERE id = :group_id AND block_code = :block_code"),
        {"group_id": group_id, "block_code": block_tab},
    ).fetchone()
    if not row:
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&error=Không tìm thấy nhóm.",
            status_code=302,
        )

    child_rows = db.execute(
        text("SELECT id FROM leave_group_configs WHERE parent_group_id = :group_id"),
        {"group_id": group_id},
    ).fetchall()
    child_ids = [str(r[0]) for r in child_rows if r and r[0]]

    for child_id in child_ids:
        db.execute(
            text("DELETE FROM leave_group_members WHERE group_id = :group_id"),
            {"group_id": child_id},
        )

    db.execute(text("DELETE FROM leave_group_members WHERE group_id = :group_id"), {"group_id": group_id})
    db.execute(text("DELETE FROM leave_group_configs WHERE parent_group_id = :group_id"), {"group_id": group_id})
    db.execute(text("DELETE FROM leave_group_configs WHERE id = :group_id"), {"group_id": group_id})
    db.commit()

    return RedirectResponse(
        url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&msg=Đã xóa hẳn nhóm.",
        status_code=302,
    )

@router.post("/groups/deactivate")
def deactivate_leave_group(
    request: Request,
    group_id: str = Form(...),
    block_tab: str = Form(...),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    if block_tab not in {"HANH_CHINH", "CHUYEN_MON"}:
        raise HTTPException(status_code=400, detail="Khối không hợp lệ.")

    role_codes = _get_role_codes(db, user.id)
    if not _can_manage_block(role_codes, block_tab):
        raise HTTPException(status_code=403, detail="Bạn không có quyền khóa nhóm trong khối này.")

    row = db.execute(
        text(
            """
            SELECT id
            FROM leave_group_configs
            WHERE id = :group_id
              AND block_code = :block_code
              AND is_active = 1
            """
        ),
        {"group_id": group_id, "block_code": block_tab},
    ).fetchone()
    if not row:
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&error=Không tìm thấy nhóm.",
            status_code=302,
        )

    db.execute(
        text("DELETE FROM leave_group_members WHERE group_id = :group_id"),
        {"group_id": group_id},
    )
    db.execute(
        text(
            """
            UPDATE leave_group_configs
            SET is_active = 0,
                updated_at = :updated_at
            WHERE id = :group_id
            """
        ),
        {"group_id": group_id, "updated_at": _dt_to_str(_now_vn())},
    )
    db.commit()

    return RedirectResponse(
        url=f"/leave-schedule?block_tab={block_tab}&sub_tab=groups&msg=Đã khóa nhóm.",
        status_code=302,
    )

@router.post("/adjust-annual")
def adjust_annual_leave(
    request: Request,
    user_id: str = Form(...),
    leave_year: int = Form(...),
    extra_days: int = Form(...),
    reason: str = Form(""),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)
    if not _can_admin_extra_days(db, user):
        raise HTTPException(status_code=403, detail="Chỉ Admin mới được cộng thêm ngày phép năm.")
    target_user = db.get(Users, user_id)
    if not target_user:
        raise HTTPException(status_code=404, detail="Không tìm thấy người dùng.")
    if extra_days <= 0:
        raise HTTPException(status_code=400, detail="Số ngày cộng thêm phải lớn hơn 0.")

    db.execute(
        text(
            """
            INSERT INTO leave_year_adjustments (id, user_id, leave_year, extra_days, reason, created_by_id, created_at)
            VALUES (:id, :user_id, :leave_year, :extra_days, :reason, :created_by_id, :created_at)
            """
        ),
        {
            "id": str(uuid.uuid4()),
            "user_id": user_id,
            "leave_year": int(leave_year),
            "extra_days": int(extra_days),
            "reason": (reason or "").strip(),
            "created_by_id": user.id,
            "created_at": _dt_to_str(_now_vn()),
        },
    )
    db.commit()
    return RedirectResponse(url="/leave-schedule?msg=Đã cộng thêm ngày phép năm.", status_code=302)



@router.get("/api/nav-badge")
def leave_schedule_nav_badge(request: Request, db: Session = Depends(get_db)):
    _ensure_tables(db)
    user = login_required(request, db)

    leave_notice_count = _load_leave_notice_unread_count(db, user.id)
    roster_notice_count = _load_roster_notice_unread_count(db, user.id)
    total_count = leave_notice_count + roster_notice_count

    return JSONResponse(
        {
            "ok": True,
            "unread_leave_schedule_count": total_count,
            "unread_leave_notice_count": leave_notice_count,
            "unread_roster_notice_count": roster_notice_count,
        }
    )

@router.post("/notice/mark-read")
def mark_leave_notice_read(
    request: Request,
    notice_id: str = Form(...),
    block_tab: str = Form(""),
    return_sub_tab: str = Form("leave_notice"),
    return_month: str = Form(""),
    return_week: str = Form(""),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    db.execute(
        text(
            """
            UPDATE leave_schedule_notices
            SET is_read = 1
            WHERE id = :id AND user_id = :user_id
            """
        ),
        {"id": notice_id, "user_id": str(user.id)},
    )
    db.commit()

    safe_sub_tab = return_sub_tab if return_sub_tab in {"leave_notice", "roster_notice"} else "leave_notice"

    query_parts: list[str] = []
    if block_tab:
        query_parts.append(f"block_tab={block_tab}")
    query_parts.append(f"sub_tab={safe_sub_tab}")
    if return_month:
        query_parts.append(f"month={return_month}")
    if safe_sub_tab == "roster_notice" and return_week:
        query_parts.append(f"week={return_week}")
    query_parts.append("msg=Đã đánh dấu đã đọc.")

    return RedirectResponse(
        url="/leave-schedule?" + "&".join(query_parts),
        status_code=302,
    )

@router.get("/export-calendar")
def export_leave_calendar_matrix(
    request: Request,
    block_tab: str = "",
    month: str = "",
    view_mode: str = "combined",
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    scope_info = _scope_payload(db, user)
    effective_block = block_tab if block_tab in {"HANH_CHINH", "CHUYEN_MON"} else (scope_info.get("block_code") or "")
    if not effective_block:
        raise HTTPException(status_code=400, detail="Khối không hợp lệ.")

    target_month = month or _today_vn().strftime("%Y-%m")
    context = _build_view_calendar_context(
        db,
        scope_info=scope_info,
        block_code=effective_block,
        month_str=target_month,
        view_mode=view_mode,
    )

    mode_label_map = {
        "combined": "Nghỉ-Trực",
        "leave": "Chỉ nghỉ",
        "roster": "Chỉ trực",
    }
    mode_label = mode_label_map.get(view_mode, "Nghỉ-Trực")
    title = f"Ma trận lịch {mode_label} - {effective_block} - {target_month}"
    html = _render_calendar_export_html(
        title=title,
        view_calendar_context=context,
    )

    filename = f"ma_tran_lich_nghi_truc_{effective_block.lower()}_{target_month}_{view_mode}.xls"
    return StreamingResponse(
        BytesIO(html.encode("utf-8")),
        media_type="application/vnd.ms-excel; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export")
def export_leave_requests(
    request: Request,
    month: str = "",
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)
    if not (_can_adjust_after_cutoff(db, user) or _can_approve(db, user) or _can_admin_extra_days(db, user)):
        raise HTTPException(status_code=403, detail="Bạn không có quyền xuất dữ liệu đăng ký nghỉ.")

    target_month = month or _today_vn().strftime("%Y-%m")
    rows = db.execute(
        text(
            """
            SELECT *
            FROM leave_requests
            WHERE substr(start_date, 1, 7) = :target_month
            ORDER BY start_date ASC, group_label ASC, user_name ASC
            """
        ),
        {"target_month": target_month},
    ).fetchall()
    data = [_format_request_row(_row_to_dict(r)) for r in rows]

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, Side
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail="Thiếu thư viện openpyxl. Cần bổ sung openpyxl vào môi trường chạy ứng dụng để xuất Excel.",
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Dang ky nghi"

    ws["A1"] = "BẢNG TỔNG HỢP ĐĂNG KÝ NGHỈ"
    ws["A2"] = f"Tháng: {target_month}"
    ws.merge_cells("A1:K1")
    ws.merge_cells("A2:K2")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"].font = Font(italic=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = [
        "STT", "Họ và tên", "Nhóm/Bộ phận", "Loại nghỉ", "Ký hiệu",
        "Từ ngày", "Đến ngày", "Số ngày công", "Lý do", "Trạng thái", "Người duyệt",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_no = 5
    for idx, item in enumerate(data, start=1):
        values = [
            idx,
            item.get("user_name"),
            item.get("group_label"),
            item.get("leave_type_label"),
            item.get("symbol"),
            item.get("start_date"),
            item.get("end_date"),
            item.get("day_count"),
            item.get("reason"),
            item.get("status_label"),
            item.get("approved_by_name") or "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_no, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        row_no += 1

    widths = {"A": 8, "B": 24, "C": 24, "D": 18, "E": 10, "F": 14, "G": 14, "H": 14, "I": 28, "J": 16, "K": 24}
    for col_name, width in widths.items():
        ws.column_dimensions[col_name].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"dang_ky_nghi_{target_month}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
    
    
@router.post("/swap-approve")
def approve_leave_swap_request(
    request: Request,
    swap_id: str = Form(...),
    block_tab: str = Form(...),
    month: str = Form(""),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    month_key = month or (
        f"{_today_vn().year + 1}-01"
        if _today_vn().month == 12
        else f"{_today_vn().year}-{_today_vn().month + 1:02d}"
    )

    role_codes = _get_role_codes(db, user.id)
    if not _can_approve_block(role_codes, block_tab):
        raise HTTPException(status_code=403, detail="Bạn không có quyền phê duyệt điều chỉnh lịch nghỉ.")

    row = db.execute(
        text("SELECT * FROM leave_swap_requests WHERE id = :id"),
        {"id": swap_id},
    ).fetchone()
    if not row:
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=register&month={month_key}&error=Không tìm thấy yêu cầu hoán đổi.",
            status_code=302,
        )

    item = _row_to_dict(row)
    if item.get("status") != "PENDING":
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=register&month={month_key}&error=Yêu cầu hoán đổi này không còn ở trạng thái chờ duyệt.",
            status_code=302,
        )

    row_a = db.execute(
        text("SELECT * FROM leave_requests WHERE id = :id"),
        {"id": item["request_a_id"]},
    ).fetchone()
    row_b = db.execute(
        text("SELECT * FROM leave_requests WHERE id = :id"),
        {"id": item["request_b_id"]},
    ).fetchone()
    if not row_a or not row_b:
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=register&month={month_key}&error=Phiếu nghỉ dùng để hoán đổi không còn tồn tại.",
            status_code=302,
        )

    a = _row_to_dict(row_a)
    b = _row_to_dict(row_b)

    if a.get("status") not in {"PENDING", "APPROVED"} or b.get("status") not in {"PENDING", "APPROVED"}:
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=register&month={month_key}&error=Chỉ được phê duyệt điều chỉnh với các phiếu còn hiệu lực.",
            status_code=302,
        )
        
    if str(a.get("id") or "") == str(b.get("id") or ""):
        return RedirectResponse(
            url=f"/leave-schedule?block_tab={block_tab}&sub_tab=register&month={month_key}&error=Yêu cầu hoán đổi không hợp lệ vì hai phiếu nghỉ trùng nhau.",
            status_code=302,
        )

    a_old_start = _parse_date(str(a["start_date"]), "Ngày bắt đầu A")
    a_old_end = _parse_date(str(a["end_date"]), "Ngày kết thúc A")
    b_old_start = _parse_date(str(b["start_date"]), "Ngày bắt đầu B")
    b_old_end = _parse_date(str(b["end_date"]), "Ngày kết thúc B")

    now_str = _dt_to_str(_now_vn())

    a_new_start = b_old_start
    a_new_end = b_old_end
    b_new_start = a_old_start
    b_new_end = a_old_end

    a_new_day_count = _count_workdays(a_new_start, a_new_end)
    b_new_day_count = _count_workdays(b_new_start, b_new_end)

    db.execute(
        text(
            """
            UPDATE leave_requests
            SET start_date = :new_start,
                end_date = :new_end,
                day_count = :day_count,
                updated_at = :updated_at
            WHERE id = :id
            """
        ),
        {
            "id": a["id"],
            "new_start": _date_to_str(a_new_start),
            "new_end": _date_to_str(a_new_end),
            "day_count": a_new_day_count,
            "updated_at": now_str,
        },
    )

    db.execute(
        text(
            """
            UPDATE leave_requests
            SET start_date = :new_start,
                end_date = :new_end,
                day_count = :day_count,
                updated_at = :updated_at
            WHERE id = :id
            """
        ),
        {
            "id": b["id"],
            "new_start": _date_to_str(b_new_start),
            "new_end": _date_to_str(b_new_end),
            "day_count": b_new_day_count,
            "updated_at": now_str,
        },
    )

    db.execute(
        text(
            """
            UPDATE leave_swap_requests
            SET status = 'APPROVED',
                approved_by_id = :approved_by_id,
                approved_by_name = :approved_by_name,
                approved_at = :approved_at,
                updated_at = :updated_at
            WHERE id = :id
            """
        ),
        {
            "id": swap_id,
            "approved_by_id": str(user.id),
            "approved_by_name": user.full_name or user.username,
            "approved_at": now_str,
            "updated_at": now_str,
        },
    )

    if a.get("user_id"):
        _create_leave_notice(
            db,
            user_id=str(a.get("user_id")),
            block_code=block_tab,
            notice_type="LEAVE_SWAP_APPROVED",
            title="Điều chỉnh lịch nghỉ đã được duyệt",
            message=(
                f"Phiếu nghỉ của anh/chị đã được điều chỉnh từ "
                f"{a.get('start_date')} - {a.get('end_date')} sang "
                f"{_date_to_str(a_new_start)} - {_date_to_str(a_new_end)}."
            ),
            related_request_id=str(a.get("id")),
        )

    if b.get("user_id"):
        _create_leave_notice(
            db,
            user_id=str(b.get("user_id")),
            block_code=block_tab,
            notice_type="LEAVE_SWAP_APPROVED",
            title="Điều chỉnh lịch nghỉ đã được duyệt",
            message=(
                f"Phiếu nghỉ của anh/chị đã được điều chỉnh từ "
                f"{b.get('start_date')} - {b.get('end_date')} sang "
                f"{_date_to_str(b_new_start)} - {_date_to_str(b_new_end)}."
            ),
            related_request_id=str(b.get("id")),
        )

    db.commit()

    return RedirectResponse(
        url=f"/leave-schedule?block_tab={block_tab}&sub_tab=register&month={month_key}&msg=Đã phê duyệt điều chỉnh lịch nghỉ.",
        status_code=302,
    )