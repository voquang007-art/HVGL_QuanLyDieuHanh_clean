# -*- coding: utf-8 -*-
from __future__ import annotations
import os
import uuid
import logging
import unicodedata
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from io import BytesIO
from typing import Any

from anyio import from_thread

from fastapi import APIRouter, Depends, Form, HTTPException, Request
from sqlalchemy import text
from sqlalchemy.orm import Session
from starlette.responses import JSONResponse, RedirectResponse, Response, StreamingResponse
from starlette.templating import Jinja2Templates

from ..chat.realtime import manager
from ..config import settings
from ..database import get_db
from ..models import (
    BlockCode,
    Roles,
    UnitStatus,
    Units,
    UserRoles,
    UserUnitMemberships,
    Users,
)
from ..security.deps import login_required
from ..work_access import get_work_access_from_session

logger = logging.getLogger("app.work_positions")

router = APIRouter(prefix="/work-positions", tags=["work_positions"])
templates = Jinja2Templates(directory=os.path.join(os.path.dirname(__file__), "..", "templates"))

templates.env.globals["work_access"] = get_work_access_from_session
templates.env.globals["pending_user_approval_count"] = lambda: 0
templates.env.globals["user_unit_names"] = lambda user_id: []


CREATE_WORK_POSITION_CATALOG_SQL = """
CREATE TABLE IF NOT EXISTS work_position_catalog (
    id TEXT PRIMARY KEY,
    owner_unit_id TEXT,
    owner_unit_name TEXT,
    apply_unit_id TEXT,
    apply_unit_name TEXT,
    block_code TEXT,
    position_type TEXT NOT NULL,
    position_name TEXT NOT NULL,
    sort_order INTEGER NOT NULL DEFAULT 0,
    is_active INTEGER NOT NULL DEFAULT 1,
    created_by_id TEXT,
    created_by_name TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
)
"""

CREATE_WORK_POSITION_PLANS_SQL = """
CREATE TABLE IF NOT EXISTS work_position_plans (
    id TEXT PRIMARY KEY,
    block_code TEXT,
    scope_unit_id TEXT,
    scope_unit_name TEXT,
    scope_subunit_id TEXT,
    scope_subunit_name TEXT,
    target_date TEXT NOT NULL,
    shift_code TEXT,
    status TEXT NOT NULL DEFAULT 'DRAFT',
    submitted_by_id TEXT,
    submitted_by_name TEXT,
    submitted_at TEXT,
    approved_by_id TEXT,
    approved_by_name TEXT,
    approved_at TEXT,
    unit_approved_by_id TEXT,
    unit_approved_by_name TEXT,
    unit_approved_at TEXT,
    khth_approved_by_id TEXT,
    khth_approved_by_name TEXT,
    khth_approved_at TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
)
"""

CREATE_WORK_POSITION_PLAN_ITEMS_SQL = """
CREATE TABLE IF NOT EXISTS work_position_plan_items (
    id TEXT PRIMARY KEY,
    plan_id TEXT NOT NULL,
    catalog_id TEXT,
    position_name TEXT NOT NULL,
    assigned_user_id TEXT,
    assigned_user_name TEXT,
    source_contract_type TEXT,
    source_contract_ref TEXT,
    notes TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
)
"""

CREATE_WORK_POSITION_NOTICES_SQL = """
CREATE TABLE IF NOT EXISTS work_position_notices (
    id TEXT PRIMARY KEY,
    user_id TEXT NOT NULL,
    target_date TEXT,
    notice_type TEXT NOT NULL,
    title TEXT NOT NULL,
    message TEXT,
    related_plan_id TEXT,
    is_read INTEGER NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL,
    approved_at TEXT
)
"""

CREATE_WORK_POSITION_ADJUSTMENTS_SQL = """
CREATE TABLE IF NOT EXISTS work_position_adjustments (
    id TEXT PRIMARY KEY,
    original_plan_id TEXT NOT NULL,
    block_code TEXT,
    scope_unit_id TEXT,
    scope_unit_name TEXT,
    scope_subunit_id TEXT,
    scope_subunit_name TEXT,
    target_date TEXT NOT NULL,
    status TEXT NOT NULL DEFAULT 'UNIT_APPROVED',
    adjust_type TEXT NOT NULL,
    reason TEXT,
    requested_by_id TEXT,
    requested_by_name TEXT,
    requested_at TEXT,
    khth_approved_by_id TEXT,
    khth_approved_by_name TEXT,
    khth_approved_at TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
)
"""

CREATE_WORK_POSITION_ADJUSTMENT_ITEMS_SQL = """
CREATE TABLE IF NOT EXISTS work_position_adjustment_items (
    id TEXT PRIMARY KEY,
    adjustment_id TEXT NOT NULL,
    original_plan_id TEXT NOT NULL,
    original_item_id TEXT,
    catalog_id TEXT,
    position_name TEXT,
    work_date TEXT,
    session_code TEXT,
    staff_group TEXT,
    adjust_type TEXT NOT NULL,
    old_assigned_user_id TEXT,
    old_assigned_user_name TEXT,
    new_assigned_user_id TEXT,
    new_assigned_user_name TEXT,
    notes TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
)
"""

CREATE_INDEXES_SQL = [
    "CREATE INDEX IF NOT EXISTS idx_wp_catalog_owner_apply ON work_position_catalog (owner_unit_id, apply_unit_id, is_active)",
    "CREATE INDEX IF NOT EXISTS idx_wp_plans_scope_date ON work_position_plans (scope_unit_id, scope_subunit_id, target_date, status)",
    "CREATE INDEX IF NOT EXISTS idx_wp_plan_items_plan ON work_position_plan_items (plan_id)",
    "CREATE INDEX IF NOT EXISTS idx_wp_notices_user_read ON work_position_notices (user_id, is_read, created_at)",
    "CREATE INDEX IF NOT EXISTS idx_wp_adjustments_plan_status ON work_position_adjustments (original_plan_id, status, target_date)",
    "CREATE INDEX IF NOT EXISTS idx_wp_adjustment_items_adjustment ON work_position_adjustment_items (adjustment_id)",
    "CREATE INDEX IF NOT EXISTS idx_wp_adjustment_items_cell ON work_position_adjustment_items (original_plan_id, catalog_id, work_date, session_code, staff_group)",
]

ASSIGN_APPROVER_ROLE_CODES = {
    "ROLE_ADMIN",
    "ROLE_LANH_DAO",
    "ROLE_TONG_GIAM_DOC",
    "ROLE_PHO_TONG_GIAM_DOC_THUONG_TRUC",
    "ROLE_PHO_TONG_GIAM_DOC",
    "ROLE_GIAM_DOC",
    "ROLE_PHO_GIAM_DOC_TRUC",
    "ROLE_PHO_GIAM_DOC",
    "ROLE_TRUONG_PHONG",
    "ROLE_PHO_PHONG",
    "ROLE_TO_TRUONG",
    "ROLE_PHO_TO",
    "ROLE_TRUONG_KHOA",
    "ROLE_PHO_TRUONG_KHOA",
    "ROLE_DIEU_DUONG_TRUONG",
    "ROLE_KY_THUAT_VIEN_TRUONG",
    "ROLE_TRUONG_DON_VI",
    "ROLE_PHO_DON_VI",
    "ROLE_DIEU_DUONG_TRUONG_DON_VI",
    "ROLE_KY_THUAT_VIEN_TRUONG_DON_VI",
    "ROLE_TRUONG_NHOM",
    "ROLE_PHO_NHOM",
}

KHTH_VIEWER_ROLE_CODES = {"ROLE_TRUONG_PHONG", "ROLE_PHO_PHONG"}

HOSPITAL_POSITION_SUMMARY_BGD_ROLE_CODES = {
    "ROLE_LANH_DAO",
    "ROLE_TONG_GIAM_DOC",
    "ROLE_PHO_TONG_GIAM_DOC_THUONG_TRUC",
    "ROLE_PHO_TONG_GIAM_DOC",
    "ROLE_GIAM_DOC",
    "ROLE_PHO_GIAM_DOC_TRUC",
    "ROLE_PHO_GIAM_DOC",
}

HOSPITAL_POSITION_SUMMARY_KHOA_ROLE_CODES = {
    "ROLE_TRUONG_KHOA",
    "ROLE_PHO_TRUONG_KHOA",
}

HOSPITAL_POSITION_SUMMARY_LIMITED_UNIT_ROLE_CODES = {
    "ROLE_DIEU_DUONG_TRUONG",
    "ROLE_TO_TRUONG",
    "ROLE_PHO_TO",
}


def _now_vn() -> datetime:
    return datetime.utcnow() + timedelta(hours=7)


def _dt_to_str(value: datetime | None) -> str | None:
    return value.strftime("%Y-%m-%d %H:%M:%S") if value else None

def _date_to_str(value) -> str | None:
    return value.isoformat() if value else None


def _parse_week_value(week_value: str) -> datetime:
    raw = str(week_value or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="Tuần không hợp lệ.")
    try:
        return datetime.strptime(raw + "-1", "%G-W%V-%u")
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Tuần không hợp lệ.") from exc


def _week_value_from_date_str(date_str: str) -> str:
    if not date_str:
        return ""
    try:
        dt = datetime.strptime(str(date_str), "%Y-%m-%d")
        iso_year, iso_week, _ = dt.isocalendar()
        return f"{iso_year}-W{iso_week:02d}"
    except Exception:
        return ""


def _week_label_from_date_str(date_str: str) -> str:
    if not date_str:
        return "-"
    try:
        dt = datetime.strptime(str(date_str), "%Y-%m-%d")
        start = dt.date()
        end = start + timedelta(days=6)
        return f"Tuần {start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}"
    except Exception:
        return str(date_str)


def _current_week_value() -> str:
    now = _now_vn()
    iso_year, iso_week, _ = now.isocalendar()
    return f"{iso_year}-W{iso_week:02d}"


def _build_week_days(target_week_monday: str) -> list[dict[str, str]]:
    if not target_week_monday:
        return []
    start_dt = datetime.strptime(target_week_monday, "%Y-%m-%d")
    out: list[dict[str, str]] = []
    weekday_labels = ["T2", "T3", "T4", "T5", "T6", "T7", "CN"]
    for i in range(7):
        d = start_dt + timedelta(days=i)
        out.append({
            "date": d.strftime("%Y-%m-%d"),
            "label": weekday_labels[i],
            "display": d.strftime("%d/%m"),
        })
    return out


def _row_to_dict(row: Any) -> dict[str, Any]:
    if hasattr(row, "_mapping"):
        return dict(row._mapping)
    return dict(row)


def _enum_str(value: Any) -> str:
    return str(getattr(value, "value", value) or "").upper().strip()


def _normalize_text(value: str | None) -> str:
    raw = str(value or "").strip().lower()
    raw = unicodedata.normalize("NFD", raw)
    raw = "".join(ch for ch in raw if unicodedata.category(ch) != "Mn")
    return raw


def _unit_category_value(unit: Units | None) -> str:
    if not unit:
        return ""
    return _enum_str(getattr(unit, "unit_category", None))


def _unit_block_value(unit: Units | None) -> str:
    if not unit:
        return ""
    return _enum_str(getattr(unit, "block_code", None))


def _ensure_column(db: Session, table_name: str, column_name: str, column_sql: str) -> None:
    cols = db.execute(text(f"PRAGMA table_info({table_name})")).fetchall()
    col_names = {str(r[1]) for r in cols if r and len(r) > 1}
    if column_name not in col_names:
        db.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_sql}"))


def _ensure_tables(db: Session) -> None:
    db.execute(text(CREATE_WORK_POSITION_CATALOG_SQL))
    db.execute(text(CREATE_WORK_POSITION_PLANS_SQL))
    db.execute(text(CREATE_WORK_POSITION_PLAN_ITEMS_SQL))
    db.execute(text(CREATE_WORK_POSITION_NOTICES_SQL))
    db.execute(text(CREATE_WORK_POSITION_ADJUSTMENTS_SQL))
    db.execute(text(CREATE_WORK_POSITION_ADJUSTMENT_ITEMS_SQL))

    _ensure_column(db, "work_position_plan_items", "work_date", "TEXT")
    _ensure_column(db, "work_position_plan_items", "session_code", "TEXT")
    _ensure_column(db, "work_position_plan_items", "staff_group", "TEXT")

    _ensure_column(db, "work_position_plans", "unit_approved_by_id", "TEXT")
    _ensure_column(db, "work_position_plans", "unit_approved_by_name", "TEXT")
    _ensure_column(db, "work_position_plans", "unit_approved_at", "TEXT")
    _ensure_column(db, "work_position_plans", "khth_approved_by_id", "TEXT")
    _ensure_column(db, "work_position_plans", "khth_approved_by_name", "TEXT")
    _ensure_column(db, "work_position_plans", "khth_approved_at", "TEXT")

    for sql in CREATE_INDEXES_SQL:
        db.execute(text(sql))
    db.commit()


def _table_exists(db: Session, table_name: str) -> bool:
    row = db.execute(
        text(
            """
            SELECT name
            FROM sqlite_master
            WHERE type = 'table' AND name = :table_name
            """
        ),
        {"table_name": table_name},
    ).fetchone()
    return row is not None


def _get_role_codes(db: Session, user_id: str) -> set[str]:
    rows = (
        db.query(Roles.code)
        .join(UserRoles, UserRoles.role_id == Roles.id)
        .filter(UserRoles.user_id == user_id)
        .all()
    )
    return {str(getattr(code, "value", code)).upper() for (code,) in rows if code is not None}


def _is_admin_or_leader(role_codes: set[str]) -> bool:
    return ("ROLE_ADMIN" in role_codes) or ("ROLE_LANH_DAO" in role_codes)


def _is_truong_phong(role_codes: set[str]) -> bool:
    return "ROLE_TRUONG_PHONG" in role_codes


def _is_pho_phong(role_codes: set[str]) -> bool:
    return "ROLE_PHO_PHONG" in role_codes


def _is_truong_khoa(role_codes: set[str]) -> bool:
    return "ROLE_TRUONG_KHOA" in role_codes


def _is_pho_truong_khoa(role_codes: set[str]) -> bool:
    return "ROLE_PHO_TRUONG_KHOA" in role_codes


def _is_manager_like(role_codes: set[str]) -> bool:
    return bool(
        _is_admin_or_leader(role_codes)
        or _is_truong_phong(role_codes)
        or _is_pho_phong(role_codes)
        or _is_truong_khoa(role_codes)
        or _is_pho_truong_khoa(role_codes)
    )


def _can_assign_positions(role_codes: set[str]) -> bool:
    return bool(ASSIGN_APPROVER_ROLE_CODES & role_codes)


def _primary_membership(db: Session, user_id: str) -> UserUnitMemberships | None:
    q = db.query(UserUnitMemberships).filter(UserUnitMemberships.user_id == user_id)
    if hasattr(UserUnitMemberships, "is_primary"):
        q = q.order_by(UserUnitMemberships.is_primary.desc())
    if hasattr(UserUnitMemberships, "id"):
        q = q.order_by(UserUnitMemberships.id.desc())
    return q.first()


def _primary_unit(db: Session, user_id: str) -> Units | None:
    membership = _primary_membership(db, user_id)
    if not membership:
        return None
    return db.get(Units, getattr(membership, "unit_id", None))


def _membership_units(db: Session, user_id: str) -> list[Units]:
    rows = (
        db.query(Units)
        .join(UserUnitMemberships, UserUnitMemberships.unit_id == Units.id)
        .filter(UserUnitMemberships.user_id == user_id)
        .filter(Units.trang_thai == UnitStatus.ACTIVE)
        .order_by(Units.cap_do.asc(), Units.order_index.asc(), Units.ten_don_vi.asc())
        .all()
    )
    dedup: dict[str, Units] = {}
    for unit in rows:
        if unit and getattr(unit, "id", None):
            dedup[str(unit.id)] = unit
    return list(dedup.values())


def _unit_scope_pair(db: Session, unit: Units | None) -> tuple[Units | None, Units | None]:
    if not unit:
        return None, None

    if _unit_category_value(unit) == "SUBUNIT":
        parent_unit = db.get(Units, getattr(unit, "parent_id", None))
        return parent_unit, unit

    return unit, None


def _first_unit_by_category(
    units: list[Units],
    *,
    block_code: str = "",
    categories: set[str] | None = None,
) -> Units | None:
    wanted_block = str(block_code or "").strip().upper()
    wanted_categories = {str(x or "").strip().upper() for x in (categories or set()) if str(x or "").strip()}

    for unit in units:
        if not unit:
            continue
        if wanted_block and _unit_block_value(unit) != wanted_block:
            continue
        if wanted_categories and _unit_category_value(unit) not in wanted_categories:
            continue
        return unit

    return None


def _resolve_unit_scope(
    db: Session,
    user_id: str,
    role_codes: set[str] | None = None,
) -> tuple[Units | None, Units | None]:
    units = _membership_units(db, user_id)
    if not units:
        return None, None

    role_codes = role_codes or _get_role_codes(db, user_id)

    selected_unit: Units | None = None

    if role_codes & {"ROLE_ADMIN", "ROLE_LANH_DAO", "ROLE_TONG_GIAM_DOC", "ROLE_PHO_TONG_GIAM_DOC_THUONG_TRUC", "ROLE_PHO_TONG_GIAM_DOC"}:
        selected_unit = (
            _first_unit_by_category(units, block_code="HANH_CHINH", categories={"ROOT"})
            or _first_unit_by_category(units, block_code="CHUYEN_MON", categories={"EXECUTIVE"})
        )

    if selected_unit is None and role_codes & {"ROLE_GIAM_DOC", "ROLE_PHO_GIAM_DOC_TRUC", "ROLE_PHO_GIAM_DOC"}:
        selected_unit = _first_unit_by_category(units, block_code="CHUYEN_MON", categories={"EXECUTIVE"})

    if selected_unit is None and role_codes & {"ROLE_TRUONG_KHOA", "ROLE_PHO_TRUONG_KHOA", "ROLE_DIEU_DUONG_TRUONG", "ROLE_KY_THUAT_VIEN_TRUONG"}:
        selected_unit = _first_unit_by_category(units, block_code="CHUYEN_MON", categories={"KHOA"})

    if selected_unit is None and role_codes & {"ROLE_TRUONG_DON_VI", "ROLE_PHO_DON_VI", "ROLE_DIEU_DUONG_TRUONG_DON_VI", "ROLE_KY_THUAT_VIEN_TRUONG_DON_VI", "ROLE_TRUONG_NHOM", "ROLE_PHO_NHOM"}:
        selected_unit = _first_unit_by_category(units, block_code="CHUYEN_MON", categories={"SUBUNIT"})

    if selected_unit is None and role_codes & {"ROLE_TRUONG_PHONG", "ROLE_PHO_PHONG"}:
        selected_unit = _first_unit_by_category(units, block_code="HANH_CHINH", categories={"PHONG"})

    if selected_unit is None and role_codes & {"ROLE_TO_TRUONG", "ROLE_PHO_TO"}:
        selected_unit = _first_unit_by_category(units, block_code="HANH_CHINH", categories={"SUBUNIT"})

    if selected_unit is None:
        selected_unit = _primary_unit(db, user_id) or units[0]

    return _unit_scope_pair(db, selected_unit)


def _format_unit_display(unit: Units | None, subunit: Units | None = None) -> str:
    if unit is None and subunit is None:
        return "-"
    if unit is not None and subunit is not None:
        return f"{getattr(unit, 'ten_don_vi', '') or '-'} / {getattr(subunit, 'ten_don_vi', '') or '-'}"
    return getattr(unit or subunit, "ten_don_vi", None) or "-"


def _block_label(block_code: str) -> str:
    if block_code == "HANH_CHINH":
        return "Khối hành chính"
    if block_code == "CHUYEN_MON":
        return "Khối chuyên môn"
    return "Chưa xác định"


def _position_layout_for_type(position_type: str | None) -> str:
    pos_type = str(position_type or "").strip().upper()
    if pos_type in {"PHONG_KHAM", "PHONG_DIEU_TRI"}:
        return "CLINIC_MATRIX"
    if pos_type == "QUAY":
        return "COUNTER_MULTI"
    if pos_type == "XET_NGHIEM":
        return "LAB_WEEKLY"
    return "SIMPLE_WEEKLY"


def _is_khth_manager(db: Session, user: Users, role_codes: set[str]) -> bool:
    if not (role_codes & KHTH_VIEWER_ROLE_CODES):
        return False
    for unit in _membership_units(db, user.id):
        unit_name_norm = _normalize_text(getattr(unit, "ten_don_vi", ""))
        if "ke hoach tong hop" in unit_name_norm:
            return True
    return False


def _user_has_unit_name_keyword(db: Session, user: Users, keywords: set[str]) -> bool:
    normalized_keywords = {_normalize_text(x) for x in keywords if str(x or "").strip()}
    if not normalized_keywords:
        return False

    for unit in _membership_units(db, user.id):
        unit_name_norm = _normalize_text(getattr(unit, "ten_don_vi", ""))
        if any(keyword in unit_name_norm for keyword in normalized_keywords):
            return True
    return False


def _can_view_hospital_position_summary(db: Session, user: Users, role_codes: set[str]) -> bool:
    if "ROLE_ADMIN" in role_codes:
        return True

    if _is_khth_manager(db, user, role_codes):
        return True

    if role_codes & HOSPITAL_POSITION_SUMMARY_BGD_ROLE_CODES:
        return True

    if role_codes & HOSPITAL_POSITION_SUMMARY_KHOA_ROLE_CODES:
        return True

    if "ROLE_DIEU_DUONG_TRUONG" in role_codes and _user_has_unit_name_keyword(
        db,
        user,
        {"Khoa Khám bệnh"},
    ):
        return True

    if role_codes & {"ROLE_TO_TRUONG", "ROLE_PHO_TO"}:
        if _user_has_unit_name_keyword(db, user, {"Chăm sóc khách hàng", "CSKH"}):
            return True

    return False

def _scope_option_label(unit: Units | None, subunit: Units | None) -> str:
    scope_unit = unit
    scope_subunit = subunit
    active_unit = scope_subunit or scope_unit
    block_code = _unit_block_value(active_unit)
    block_label = _block_label(block_code)

    category = _unit_category_value(active_unit)
    if category == "ROOT":
        role_hint = "HĐTV / Toàn viện"
    elif category == "EXECUTIVE":
        role_hint = "BGĐ"
    elif category == "KHOA":
        role_hint = "Khoa"
    elif category == "PHONG":
        role_hint = "Phòng"
    elif category == "SUBUNIT":
        role_hint = "Đơn vị/Tổ/Nhóm"
    else:
        role_hint = "Phạm vi"

    return f"{role_hint} - {_format_unit_display(scope_unit, scope_subunit)} - {block_label}"


def _scope_options_for_user(
    db: Session,
    user: Users,
    role_codes: set[str],
) -> list[dict[str, Any]]:
    units = _membership_units(db, user.id)
    options: list[dict[str, Any]] = []
    seen_ids: set[str] = set()

    for unit in units:
        scope_unit, scope_subunit = _unit_scope_pair(db, unit)
        active_unit = scope_subunit or scope_unit
        selected_id = str(getattr(active_unit, "id", "") or "").strip()
        if not selected_id or selected_id in seen_ids:
            continue

        block_code = _unit_block_value(active_unit)
        item = {
            "selected_scope_unit_id": selected_id,
            "scope_unit": scope_unit,
            "scope_subunit": scope_subunit,
            "scope_unit_id": getattr(scope_unit, "id", None),
            "scope_subunit_id": getattr(scope_subunit, "id", None),
            "scope_unit_name": getattr(scope_unit, "ten_don_vi", None),
            "scope_subunit_name": getattr(scope_subunit, "ten_don_vi", None),
            "block_code": block_code,
            "block_label": _block_label(block_code),
            "scope_display": _format_unit_display(scope_unit, scope_subunit),
            "option_label": _scope_option_label(scope_unit, scope_subunit),
        }
        options.append(item)
        seen_ids.add(selected_id)

    default_unit, default_subunit = _resolve_unit_scope(db, user.id, role_codes)
    default_active_unit = default_subunit or default_unit
    default_selected_id = str(getattr(default_active_unit, "id", "") or "").strip()
    if default_selected_id and default_selected_id not in seen_ids:
        block_code = _unit_block_value(default_active_unit)
        options.insert(
            0,
            {
                "selected_scope_unit_id": default_selected_id,
                "scope_unit": default_unit,
                "scope_subunit": default_subunit,
                "scope_unit_id": getattr(default_unit, "id", None),
                "scope_subunit_id": getattr(default_subunit, "id", None),
                "scope_unit_name": getattr(default_unit, "ten_don_vi", None),
                "scope_subunit_name": getattr(default_subunit, "ten_don_vi", None),
                "block_code": block_code,
                "block_label": _block_label(block_code),
                "scope_display": _format_unit_display(default_unit, default_subunit),
                "option_label": _scope_option_label(default_unit, default_subunit),
            },
        )

    return options


def _scope_payload(
    db: Session,
    user: Users,
    selected_scope_unit_id: str | None = None,
) -> dict[str, Any]:
    role_codes = _get_role_codes(db, user.id)
    scope_options = _scope_options_for_user(db, user, role_codes)
    selected_scope_unit_id = str(selected_scope_unit_id or "").strip()

    selected_option: dict[str, Any] | None = None
    if selected_scope_unit_id:
        for option in scope_options:
            if str(option.get("selected_scope_unit_id") or "") == selected_scope_unit_id:
                selected_option = option
                break

        if selected_option is None:
            raise HTTPException(status_code=403, detail="Phạm vi thao tác không thuộc quyền của tài khoản hiện tại.")

    if selected_option is None and scope_options:
        default_unit, default_subunit = _resolve_unit_scope(db, user.id, role_codes)
        default_active_unit = default_subunit or default_unit
        default_selected_id = str(getattr(default_active_unit, "id", "") or "").strip()

        for option in scope_options:
            if str(option.get("selected_scope_unit_id") or "") == default_selected_id:
                selected_option = option
                break

        if selected_option is None:
            selected_option = scope_options[0]

    if selected_option:
        scope_unit = selected_option.get("scope_unit")
        scope_subunit = selected_option.get("scope_subunit")
        block_code = str(selected_option.get("block_code") or "")
        current_selected_scope_unit_id = str(selected_option.get("selected_scope_unit_id") or "")
    else:
        scope_unit = None
        scope_subunit = None
        block_code = ""
        current_selected_scope_unit_id = ""

    return {
        "user_id": str(user.id),
        "user_name": user.full_name or user.username or "",
        "role_codes": sorted(role_codes),
        "is_admin_or_leader": _is_admin_or_leader(role_codes),
        "is_manager_like": _is_manager_like(role_codes),
        "can_assign_positions": _can_assign_positions(role_codes),
        "is_khth_manager": _is_khth_manager(db, user, role_codes),
        "block_code": block_code,
        "block_label": _block_label(block_code),
        "scope_unit": scope_unit,
        "scope_subunit": scope_subunit,
        "scope_unit_id": getattr(scope_unit, "id", None),
        "scope_subunit_id": getattr(scope_subunit, "id", None),
        "scope_unit_name": getattr(scope_unit, "ten_don_vi", None),
        "scope_subunit_name": getattr(scope_subunit, "ten_don_vi", None),
        "scope_display": _format_unit_display(scope_unit, scope_subunit),
        "selected_scope_unit_id": current_selected_scope_unit_id,
        "scope_options": scope_options,
    }


def _position_notify_payload(event_type: str, **kwargs) -> dict[str, Any]:
    payload = {
        "module": "work_position",
        "type": event_type,
        "sent_at": _dt_to_str(_now_vn()),
    }
    payload.update(kwargs)
    return payload

async def _notify_work_position_users(user_ids: list[str], payload: dict[str, Any]) -> None:
    clean_ids: list[str] = []
    for raw_user_id in user_ids or []:
        uid = str(raw_user_id or "").strip()
        if uid and uid not in clean_ids:
            clean_ids.append(uid)

    if not clean_ids:
        return

    await manager.notify_users_json(clean_ids, payload)


def _dedupe_user_ids(user_ids: list[str]) -> list[str]:
    out: list[str] = []
    for raw_user_id in user_ids or []:
        uid = str(raw_user_id or "").strip()
        if uid and uid not in out:
            out.append(uid)
    return out


def _hospital_position_summary_viewer_user_ids(db: Session) -> list[str]:
    out: list[str] = []
    for item in db.query(Users).all():
        user_id = str(getattr(item, "id", "") or "").strip()
        if not user_id:
            continue
        role_codes = _get_role_codes(db, user_id)
        if _can_view_hospital_position_summary(db, item, role_codes):
            out.append(user_id)
    return _dedupe_user_ids(out)


def _work_position_scope_realtime_user_ids(
    db: Session,
    *,
    scope_unit_id: str | None = None,
    scope_subunit_id: str | None = None,
    include_khth: bool = False,
    include_hospital_viewers: bool = False,
    extra_user_ids: list[str] | None = None,
) -> list[str]:
    unit_ids = [str(x) for x in [scope_unit_id, scope_subunit_id] if str(x or "").strip()]
    user_ids = _unit_member_user_ids(db, unit_ids)

    if include_khth:
        user_ids.extend(_load_khth_manager_user_ids(db))

    if include_hospital_viewers:
        user_ids.extend(_hospital_position_summary_viewer_user_ids(db))

    if extra_user_ids:
        user_ids.extend([str(x) for x in extra_user_ids if str(x or "").strip()])

    return _dedupe_user_ids(user_ids)


def _emit_work_position_realtime_sync(user_ids: list[str], payload: dict[str, Any]) -> None:
    """
    Phát realtime trong các route đang khai báo def.
    Không đổi hàng loạt route sang async def để tránh phát sinh lỗi ngoài phạm vi.
    """
    clean_ids: list[str] = []
    for raw_user_id in user_ids or []:
        uid = str(raw_user_id or "").strip()
        if uid and uid not in clean_ids:
            clean_ids.append(uid)

    if not clean_ids:
        return

    try:
        from_thread.run(manager.notify_users_json, clean_ids, payload)
    except Exception as ex:
        logger.exception("[work_positions] Lỗi phát realtime work_position: %s", ex)


def _work_position_all_changed_payload(event_type: str, *, actor_user_id: str = "", **kwargs) -> dict[str, Any]:
    payload = _position_notify_payload(
        event_type,
        actor_user_id=str(actor_user_id or ""),
    )
    payload.update(kwargs)
    return payload


def _is_two_step_position_type(block_code: str, position_type: str) -> bool:
    block_val = str(block_code or "").strip().upper()
    pos_type = str(position_type or "").strip().upper()
    if block_val != "CHUYEN_MON":
        return False
    return pos_type in {"PHONG_KHAM", "XET_NGHIEM"}


def _plan_needs_khth_approval(block_code: str, plan_items: list[dict[str, Any]]) -> bool:
    if str(block_code or "").strip().upper() != "CHUYEN_MON":
        return False
    for item in plan_items:
        if _is_two_step_position_type(block_code, str(item.get("position_type") or "")):
            return True
    return False


def _approval_status_label(value: str) -> str:
    raw = str(value or "").strip().upper()
    return {
        "DRAFT": "Chờ duyệt cấp đơn vị",
        "UNIT_APPROVED": "Đã duyệt cấp đơn vị",
        "KHTH_APPROVED": "Đã duyệt KHTH",
        "APPROVED": "Đã duyệt",
    }.get(raw, raw or "-")

def _adjustment_status_label(value: str) -> str:
    raw = str(value or "").strip().upper()
    return {
        "UNIT_APPROVED": "Chờ phê duyệt cấp Bệnh viện",
        "KHTH_APPROVED": "Đã phê duyệt cấp Bệnh viện",
        "REJECTED": "Đã từ chối",
        "CANCELLED": "Đã hủy",
    }.get(raw, raw or "-")


def _adjustment_type_label(value: str) -> str:
    raw = str(value or "").strip().upper()
    return {
        "CLOSE_ROOM": "Không thực hiện / Đóng phòng",
        "CHANGE_BS": "Thay bác sĩ",
        "CHANGE_DD": "Thay điều dưỡng / thư ký y khoa",
        "CHANGE_BOTH": "Thay cả bác sĩ và điều dưỡng / thư ký y khoa",
    }.get(raw, raw or "-")


def _unit_member_user_ids(db: Session, unit_ids: list[str]) -> list[str]:
    if not unit_ids:
        return []
    rows = (
        db.query(UserUnitMemberships.user_id)
        .filter(UserUnitMemberships.unit_id.in_(unit_ids))
        .distinct()
        .all()
    )
    return [str(user_id) for (user_id,) in rows if user_id]


async def _notify_position_users(db: Session, user_ids: list[str], payload: dict[str, Any]) -> None:
    target_ids = [str(x) for x in user_ids if str(x).strip()]
    if not target_ids:
        return
    await manager.notify_users_json(target_ids, payload)


async def _notify_position_scope(
    db: Session,
    *,
    scope_unit_id: str | None = None,
    scope_subunit_id: str | None = None,
    payload: dict[str, Any],
) -> None:
    unit_ids = [x for x in [scope_unit_id, scope_subunit_id] if x]
    user_ids = _unit_member_user_ids(db, unit_ids)
    await _notify_position_users(db, user_ids, payload)


def _load_khth_manager_user_ids(db: Session) -> list[str]:
    khth_user_ids: list[str] = []

    for item in db.query(Users).all():
        user_id = str(getattr(item, "id", "") or "")
        if not user_id:
            continue

        role_codes = _get_role_codes(db, user_id)
        if _is_khth_manager(db, item, role_codes):
            if user_id not in khth_user_ids:
                khth_user_ids.append(user_id)

    return khth_user_ids


def _create_khth_pending_notice(
    db: Session,
    *,
    notice_type: str,
    title: str,
    message: str,
    target_date: str,
    related_plan_id: str,
) -> list[str]:
    created_at = _dt_to_str(_now_vn()) or ""
    khth_user_ids = _load_khth_manager_user_ids(db)

    for user_id in khth_user_ids:
        db.execute(
            text(
                """
                INSERT INTO work_position_notices (
                    id, user_id, target_date, notice_type, title, message,
                    related_plan_id, is_read, created_at, approved_at
                ) VALUES (
                    :id, :user_id, :target_date, :notice_type, :title, :message,
                    :related_plan_id, 0, :created_at, :approved_at
                )
                """
            ),
            {
                "id": str(uuid.uuid4()),
                "user_id": user_id,
                "target_date": str(target_date or ""),
                "notice_type": str(notice_type or ""),
                "title": str(title or ""),
                "message": str(message or ""),
                "related_plan_id": str(related_plan_id or ""),
                "created_at": created_at,
                "approved_at": "",
            },
        )

    return khth_user_ids


def _can_manage_catalog(db: Session, user: Users) -> bool:
    role_codes = _get_role_codes(db, user.id)
    return bool(
        _is_manager_like(role_codes)
        or role_codes & {
            "ROLE_TO_TRUONG",
            "ROLE_PHO_TO",
            "ROLE_DIEU_DUONG_TRUONG",
            "ROLE_KY_THUAT_VIEN_TRUONG",
            "ROLE_DIEU_DUONG_TRUONG_DON_VI",
            "ROLE_KY_THUAT_VIEN_TRUONG_DON_VI",
        }
    )


def _load_catalog_rows(db: Session, scope_info: dict[str, Any], block_tab: str, *, full_block_scope: bool = False) -> list[dict[str, Any]]:
    conditions = ["is_active = 1"]
    params: dict[str, Any] = {}

    if block_tab in {"HANH_CHINH", "CHUYEN_MON"}:
        conditions.append("block_code = :block_code")
        params["block_code"] = block_tab

    if not full_block_scope:
        if scope_info.get("scope_subunit_id"):
            conditions.append("(apply_unit_id = :apply_unit_id OR owner_unit_id = :owner_unit_id)")
            params["apply_unit_id"] = scope_info["scope_subunit_id"]
            params["owner_unit_id"] = scope_info["scope_subunit_id"]
        elif scope_info.get("scope_unit_id"):
            conditions.append("(apply_unit_id = :apply_unit_id OR owner_unit_id = :owner_unit_id)")
            params["apply_unit_id"] = scope_info["scope_unit_id"]
            params["owner_unit_id"] = scope_info["scope_unit_id"]

    sql = f"""
        SELECT *
        FROM work_position_catalog
        WHERE {' AND '.join(conditions)}
        ORDER BY apply_unit_name ASC, sort_order ASC, position_name ASC, created_at DESC
    """
    rows = db.execute(text(sql), params).fetchall()
    items = [_row_to_dict(r) for r in rows]
    for item in items:
        item["position_layout"] = _position_layout_for_type(item.get("position_type"))
    return items


def _load_position_plan_rows(
    db: Session,
    scope_info: dict[str, Any],
    block_tab: str,
    *,
    full_block_scope: bool = False,
    target_week_monday: str = "",
    history_year: str = "",
    history_month: str = "",
) -> list[dict[str, Any]]:
    conditions = ["1=1"]
    params: dict[str, Any] = {}

    if block_tab in {"HANH_CHINH", "CHUYEN_MON"}:
        conditions.append("block_code = :block_code")
        params["block_code"] = block_tab

    if not full_block_scope:
        if scope_info.get("scope_subunit_id"):
            conditions.append("(scope_subunit_id = :scope_subunit_id OR scope_unit_id = :scope_unit_id)")
            params["scope_subunit_id"] = scope_info["scope_subunit_id"]
            params["scope_unit_id"] = scope_info.get("scope_unit_id")
        elif scope_info.get("scope_unit_id"):
            conditions.append("scope_unit_id = :scope_unit_id")
            params["scope_unit_id"] = scope_info["scope_unit_id"]

    if target_week_monday:
        conditions.append("target_date = :target_week_monday")
        params["target_week_monday"] = target_week_monday

    if history_year:
        conditions.append("substr(target_date, 1, 4) = :history_year")
        params["history_year"] = str(history_year)

    if history_month:
        conditions.append("substr(target_date, 6, 2) = :history_month")
        params["history_month"] = f"{int(history_month):02d}"

    sql = f"""
        SELECT *
        FROM work_position_plans
        WHERE {' AND '.join(conditions)}
        ORDER BY target_date DESC, updated_at DESC
    """
    rows = db.execute(text(sql), params).fetchall()
    items = [_row_to_dict(r) for r in rows]
    if not items:
        return []

    plan_ids = [str(item.get("id")) for item in items if item.get("id")]
    item_count_map: dict[str, int] = {}
    position_types_map: dict[str, list[str]] = {}

    if plan_ids:
        placeholders = ",".join([f":p{i}" for i in range(len(plan_ids))])
        params2 = {f"p{i}": plan_id for i, plan_id in enumerate(plan_ids)}

        count_rows = db.execute(
            text(
                f"""
                SELECT plan_id, COUNT(*) AS item_count
                FROM work_position_plan_items
                WHERE plan_id IN ({placeholders})
                GROUP BY plan_id
                """
            ),
            params2,
        ).fetchall()
        item_count_map = {
            str(_row_to_dict(row).get("plan_id")): int(_row_to_dict(row).get("item_count") or 0)
            for row in count_rows
        }

        type_rows = db.execute(
            text(
                f"""
                SELECT
                    i.plan_id,
                    COALESCE(c.position_type, '') AS position_type
                FROM work_position_plan_items i
                LEFT JOIN work_position_catalog c ON c.id = i.catalog_id
                WHERE i.plan_id IN ({placeholders})
                GROUP BY i.plan_id, c.position_type
                ORDER BY i.plan_id ASC, c.position_type ASC
                """
            ),
            params2,
        ).fetchall()

        for row in type_rows:
            item = _row_to_dict(row)
            plan_id = str(item.get("plan_id") or "")
            position_type = str(item.get("position_type") or "").strip().upper()
            if not plan_id or not position_type:
                continue
            if plan_id not in position_types_map:
                position_types_map[plan_id] = []
            if position_type not in position_types_map[plan_id]:
                position_types_map[plan_id].append(position_type)

    for item in items:
        plan_id = str(item.get("id") or "")
        item["item_count"] = item_count_map.get(plan_id, 0)
        item["scope_label"] = item.get("scope_subunit_name") or item.get("scope_unit_name") or "-"
        item["week_value"] = _week_value_from_date_str(str(item.get("target_date") or ""))
        item["week_label"] = _week_label_from_date_str(str(item.get("target_date") or ""))
        item["status_label"] = _approval_status_label(str(item.get("status") or ""))
        item["position_types_label"] = ", ".join(position_types_map.get(plan_id, [])) or "-"
    return items


def _load_current_week_plan(
    db: Session,
    scope_info: dict[str, Any],
    block_tab: str,
    *,
    target_week_monday: str = "",
    selected_position_type: str = "",
) -> dict[str, Any] | None:
    if not target_week_monday:
        return None

    position_type = str(selected_position_type or "").strip().upper()
    shift_code = f"WEEKLY_MATRIX_{position_type}" if position_type else "WEEKLY_MATRIX"

    conditions = [
        "COALESCE(block_code, '') = :block_code",
        "target_date = :target_date",
        "COALESCE(shift_code, '') = :shift_code",
    ]
    params: dict[str, Any] = {
        "block_code": block_tab,
        "target_date": target_week_monday,
        "shift_code": shift_code,
    }

    if scope_info.get("scope_subunit_id"):
        conditions.append("(scope_subunit_id = :scope_subunit_id OR scope_unit_id = :scope_unit_id)")
        params["scope_subunit_id"] = scope_info["scope_subunit_id"]
        params["scope_unit_id"] = scope_info.get("scope_unit_id")
    elif scope_info.get("scope_unit_id"):
        conditions.append("scope_unit_id = :scope_unit_id")
        params["scope_unit_id"] = scope_info["scope_unit_id"]

    row = db.execute(
        text(
            f"""
            SELECT *
            FROM work_position_plans
            WHERE {' AND '.join(conditions)}
            ORDER BY updated_at DESC
            LIMIT 1
            """
        ),
        params,
    ).fetchone()
    return _row_to_dict(row) if row else None


def _find_latest_position_week(
    db: Session,
    scope_info: dict[str, Any],
    block_tab: str,
    *,
    full_block_scope: bool = False,
    hospital_summary_scope: bool = False,
) -> str:
    conditions = ["1=1"]
    params: dict[str, Any] = {}

    if hospital_summary_scope:
        conditions.append("COALESCE(block_code, '') = 'CHUYEN_MON'")
        conditions.append("status IN ('UNIT_APPROVED', 'KHTH_APPROVED')")
        conditions.append(
            """
            EXISTS (
                SELECT 1
                FROM work_position_plan_items i
                LEFT JOIN work_position_catalog c ON c.id = i.catalog_id
                WHERE i.plan_id = work_position_plans.id
                  AND COALESCE(c.position_type, '') IN ('PHONG_KHAM', 'XET_NGHIEM')
            )
            """
        )
    else:
        if block_tab in {"HANH_CHINH", "CHUYEN_MON"}:
            conditions.append("block_code = :block_code")
            params["block_code"] = block_tab

        if not full_block_scope:
            if scope_info.get("scope_subunit_id"):
                conditions.append("(scope_subunit_id = :scope_subunit_id OR scope_unit_id = :scope_unit_id)")
                params["scope_subunit_id"] = scope_info["scope_subunit_id"]
                params["scope_unit_id"] = scope_info.get("scope_unit_id")
            elif scope_info.get("scope_unit_id"):
                conditions.append("scope_unit_id = :scope_unit_id")
                params["scope_unit_id"] = scope_info["scope_unit_id"]

    row = db.execute(
        text(
            f"""
            SELECT target_date
            FROM work_position_plans
            WHERE {' AND '.join(conditions)}
            ORDER BY target_date DESC, updated_at DESC
            LIMIT 1
            """
        ),
        params,
    ).fetchone()

    if not row:
        return ""

    item = _row_to_dict(row)
    return str(item.get("target_date") or "")


def _load_khth_plan_history_rows(
    db: Session,
    *,
    history_year: str = "",
    history_month: str = "",
) -> list[dict[str, Any]]:
    conditions = [
        "COALESCE(block_code, '') = 'CHUYEN_MON'",
        "status IN ('UNIT_APPROVED', 'KHTH_APPROVED')",
    ]
    params: dict[str, Any] = {}

    if history_year:
        conditions.append("substr(target_date, 1, 4) = :history_year")
        params["history_year"] = str(history_year)

    if history_month:
        conditions.append("substr(target_date, 6, 2) = :history_month")
        params["history_month"] = f"{int(history_month):02d}"

    rows = db.execute(
        text(
            f"""
            SELECT *
            FROM work_position_plans
            WHERE {' AND '.join(conditions)}
            ORDER BY target_date DESC, scope_unit_name ASC, scope_subunit_name ASC, updated_at DESC
            """
        ),
        params,
    ).fetchall()

    raw_plan_rows = [_row_to_dict(r) for r in rows]
    plan_rows: list[dict[str, Any]] = []

    for plan in raw_plan_rows:
        plan_id = str(plan.get("id") or "")
        if not plan_id:
            continue

        item_rows = _load_plan_item_rows(db, plan_id)
        if not _plan_needs_khth_approval("CHUYEN_MON", item_rows):
            continue

        plan["status_label"] = _approval_status_label(str(plan.get("status") or ""))
        plan["scope_label"] = plan.get("scope_subunit_name") or plan.get("scope_unit_name") or "-"
        plan["week_value"] = _week_value_from_date_str(str(plan.get("target_date") or ""))
        plan["week_label"] = _week_label_from_date_str(str(plan.get("target_date") or ""))
        plan["item_count"] = len(item_rows)
        plan_rows.append(plan)

    return plan_rows
def _load_week_assignment_map(db: Session, plan_id: str | None) -> dict[str, dict[str, str]]:
    if not plan_id:
        return {}
    rows = db.execute(
        text(
            """
            SELECT *
            FROM work_position_plan_items
            WHERE plan_id = :plan_id
            ORDER BY position_name ASC, work_date ASC, session_code ASC, staff_group ASC, updated_at DESC
            """
        ),
        {"plan_id": plan_id},
    ).fetchall()
    out: dict[str, dict[str, str]] = {}
    for row in rows:
        item = _row_to_dict(row)
        key = "|".join([
            str(item.get("catalog_id") or ""),
            str(item.get("work_date") or ""),
            str(item.get("session_code") or ""),
            str(item.get("staff_group") or ""),
        ])
        assigned_ids_raw = str(item.get("assigned_user_id") or "")
        out[key] = {
            "assigned_user_id": assigned_ids_raw,
            "assigned_user_ids_csv": assigned_ids_raw,
            "assigned_user_name": str(item.get("assigned_user_name") or ""),
            "notes": str(item.get("notes") or ""),
        }
    return out


def _load_plan_item_rows(db: Session, plan_id: str) -> list[dict[str, Any]]:
    rows = db.execute(
        text(
            """
            SELECT
                i.*,
                c.position_type,
                c.apply_unit_name,
                c.owner_unit_name,
                c.block_code
            FROM work_position_plan_items i
            LEFT JOIN work_position_catalog c ON c.id = i.catalog_id
            WHERE i.plan_id = :plan_id
            ORDER BY c.apply_unit_name ASC, i.position_name ASC, i.work_date ASC, i.session_code ASC, i.staff_group ASC
            """
        ),
        {"plan_id": plan_id},
    ).fetchall()
    return [_row_to_dict(r) for r in rows]

def _role_label_from_codes(role_codes: set[str]) -> str:
    if "ROLE_GIAM_DOC" in role_codes:
        return "Giám đốc"
    if "ROLE_PHO_GIAM_DOC_TRUC" in role_codes:
        return "Phó Giám đốc trực"
    if "ROLE_PHO_GIAM_DOC" in role_codes:
        return "Phó Giám đốc"
    if "ROLE_TRUONG_KHOA" in role_codes:
        return "Trưởng khoa"
    if "ROLE_PHO_TRUONG_KHOA" in role_codes:
        return "Phó khoa"
    if "ROLE_BAC_SI" in role_codes:
        return "Bác sĩ"
    if "ROLE_DIEU_DUONG" in role_codes:
        return "Điều dưỡng"
    if "ROLE_THU_KY_Y_KHOA" in role_codes:
        return "Thư ký y khoa"
    if "ROLE_DUOC_SI" in role_codes:
        return "Dược sĩ"
    if "ROLE_KY_THUAT_VIEN" in role_codes:
        return "Kỹ thuật viên"
    if "ROLE_HO_LY" in role_codes:
        return "Hộ lý"
    if "ROLE_TRUONG_PHONG" in role_codes:
        return "Trưởng phòng"
    if "ROLE_PHO_PHONG" in role_codes:
        return "Phó phòng"
    return "Nhân sự"


def _load_scope_assignable_users(
    db: Session,
    scope_info: dict[str, Any],
    block_tab: str,
) -> dict[str, list[dict[str, Any]]]:
    q = (
        db.query(Users)
        .join(UserUnitMemberships, UserUnitMemberships.user_id == Users.id)
        .join(Units, Units.id == UserUnitMemberships.unit_id)
        .filter(Units.trang_thai == UnitStatus.ACTIVE)
    )

    if block_tab in {"HANH_CHINH", "CHUYEN_MON"}:
        q = q.filter(Units.block_code == block_tab)

    if scope_info.get("scope_subunit_id"):
        q = q.filter(
            (UserUnitMemberships.unit_id == scope_info["scope_subunit_id"]) |
            (UserUnitMemberships.unit_id == scope_info.get("scope_unit_id"))
        )
    elif scope_info.get("scope_unit_id"):
        q = q.filter(UserUnitMemberships.unit_id == scope_info["scope_unit_id"])

    users = q.distinct().all()

    all_rows: list[dict[str, Any]] = []
    bs_rows: list[dict[str, Any]] = []
    dd_rows: list[dict[str, Any]] = []

    for u in users:
        role_codes = _get_role_codes(db, str(u.id))
        role_label = _role_label_from_codes(role_codes)

        is_bs_group = bool({
            "ROLE_BAC_SI",
            "ROLE_GIAM_DOC",
            "ROLE_PHO_GIAM_DOC_TRUC",
            "ROLE_PHO_GIAM_DOC",
            "ROLE_TRUONG_KHOA",
            "ROLE_PHO_TRUONG_KHOA",
        } & role_codes)

        is_dd_group = bool({
            "ROLE_DIEU_DUONG",
            "ROLE_THU_KY_Y_KHOA",
        } & role_codes)

        row = {
            "id": str(u.id),
            "name": str(u.full_name or u.username or ""),
            "role_label": role_label,
            "role_codes": sorted(role_codes),
            "is_doctor": is_bs_group,
            "is_nurse": is_dd_group,
        }
        all_rows.append(row)
        if row["is_doctor"]:
            bs_rows.append(row)
        if row["is_nurse"]:
            dd_rows.append(row)

    all_rows.sort(key=lambda x: (x["name"], x["role_label"]))
    bs_rows.sort(key=lambda x: (x["name"], x["role_label"]))
    dd_rows.sort(key=lambda x: (x["name"], x["role_label"]))

    return {
        "all": all_rows,
        "bs": bs_rows,
        "dd": dd_rows,
    }

def _catalog_item_allowed_for_scope(
    catalog_item: dict[str, Any],
    scope_info: dict[str, Any],
    block_code: str,
    selected_position_type: str = "",
) -> bool:
    if int(catalog_item.get("is_active") or 0) != 1:
        return False

    catalog_block = str(catalog_item.get("block_code") or "").strip().upper()
    expected_block = str(block_code or "").strip().upper()
    if expected_block and catalog_block != expected_block:
        return False

    wanted_type = str(selected_position_type or "").strip().upper()
    catalog_type = str(catalog_item.get("position_type") or "").strip().upper()
    if wanted_type and catalog_type != wanted_type:
        return False

    scope_unit_id = str(scope_info.get("scope_unit_id") or "").strip()
    scope_subunit_id = str(scope_info.get("scope_subunit_id") or "").strip()
    allowed_unit_id = scope_subunit_id or scope_unit_id
    if not allowed_unit_id:
        return False

    catalog_apply_unit_id = str(catalog_item.get("apply_unit_id") or "").strip()
    catalog_owner_unit_id = str(catalog_item.get("owner_unit_id") or "").strip()
    return allowed_unit_id in {catalog_apply_unit_id, catalog_owner_unit_id}


def _assignable_user_map_for_staff_group(
    assignable_users: dict[str, list[dict[str, Any]]],
    staff_group: str,
) -> dict[str, dict[str, Any]]:
    group_code = str(staff_group or "").strip().upper()
    if group_code == "BS":
        source_rows = assignable_users.get("bs", [])
    elif group_code == "DD":
        source_rows = assignable_users.get("dd", [])
    else:
        source_rows = assignable_users.get("all", [])

    out: dict[str, dict[str, Any]] = {}
    for row in source_rows:
        uid = str(row.get("id") or "").strip()
        if uid and uid not in out:
            out[uid] = row
    return out

def _load_approved_leave_user_ids_by_date(
    db: Session,
    scope_info: dict[str, Any],
    block_tab: str,
    week_days: list[dict[str, str]],
) -> dict[str, list[str]]:
    if not _table_exists(db, "leave_requests"):
        return {}

    if not week_days:
        return {}

    start_date = week_days[0]["date"]
    end_date = week_days[-1]["date"]

    # Hợp đồng dữ liệu Nghỉ - Trực -> Vị trí làm việc:
    # - APPROVED: nghỉ đã được duyệt, bắt buộc chặn phân vị trí.
    # - PENDING: đơn đã đăng ký nhưng chưa đến kỳ/hoặc chưa hoàn tất phê duyệt,
    #   vẫn phải chặn tạm để tránh tuần vắt qua 2 tháng bị phân sai nhân sự.
    # Khi đơn bị REJECTED/CANCELLED thì tự hết chặn ở lần tải dữ liệu tiếp theo.
    conditions = [
        "status IN ('PENDING', 'APPROVED')",
        "start_date <= :end_date",
        "end_date >= :start_date",
    ]
    params: dict[str, Any] = {
        "start_date": start_date,
        "end_date": end_date,
    }

    if block_tab in {"HANH_CHINH", "CHUYEN_MON"}:
        conditions.append("COALESCE(block_code, '') = :block_code")
        params["block_code"] = block_tab

    scope_filters = []
    if scope_info.get("scope_subunit_id"):
        scope_filters.append("unit_id = :scope_subunit_id")
        params["scope_subunit_id"] = scope_info["scope_subunit_id"]
    if scope_info.get("scope_unit_id"):
        scope_filters.append("unit_id = :scope_unit_id")
        params["scope_unit_id"] = scope_info["scope_unit_id"]
    if scope_filters:
        conditions.append("(" + " OR ".join(scope_filters) + ")")

    rows = db.execute(
        text(
            f"""
            SELECT user_id, start_date, end_date
            FROM leave_requests
            WHERE {' AND '.join(conditions)}
            ORDER BY start_date ASC, end_date ASC
            """
        ),
        params,
    ).fetchall()

    out: dict[str, list[str]] = {item["date"]: [] for item in week_days}
    for row in rows:
        item = _row_to_dict(row)
        uid = str(item.get("user_id") or "")
        start_val = str(item.get("start_date") or "")
        end_val = str(item.get("end_date") or "")
        if not uid or not start_val or not end_val:
            continue
        for d in week_days:
            if start_val <= d["date"] <= end_val and uid not in out[d["date"]]:
                out[d["date"]].append(uid)
    return out


def _load_approved_roster_plus_one_user_ids_by_date(
    db: Session,
    scope_info: dict[str, Any],
    block_tab: str,
    week_days: list[dict[str, str]],
) -> dict[str, list[str]]:
    if not (_table_exists(db, "leave_roster_plans") and _table_exists(db, "leave_roster_plan_items")):
        return {item["date"]: [] for item in week_days}

    if not week_days:
        return {}
    # Quy tắc nghiệp vụ:
    # - Không chặn phân vị trí vào chính ngày trực.
    # - Chỉ chặn phân vị trí vào ngày liền sau ngày trực đã được duyệt.
    # Ví dụ: trực ngày 24 thì chỉ chặn ngày 25, không chặn ngày 24.
    start_dt = datetime.strptime(week_days[0]["date"], "%Y-%m-%d") - timedelta(days=1)
    end_dt = datetime.strptime(week_days[-1]["date"], "%Y-%m-%d") - timedelta(days=1)

    params: dict[str, Any] = {
        "start_date": start_dt.strftime("%Y-%m-%d"),
        "end_date": end_dt.strftime("%Y-%m-%d"),
    }
    conditions = [
        "p.status = 'APPROVED'",
        "i.work_date >= :start_date",
        "i.work_date <= :end_date",
    ]

    if block_tab in {"HANH_CHINH", "CHUYEN_MON"}:
        conditions.append("COALESCE(p.block_code, '') = :block_code")
        params["block_code"] = block_tab

    if scope_info.get("scope_subunit_id"):
        conditions.append("(COALESCE(p.scope_subunit_id, '') = :scope_subunit_id OR COALESCE(p.scope_unit_id, '') = :scope_unit_id)")
        params["scope_subunit_id"] = str(scope_info["scope_subunit_id"])
        params["scope_unit_id"] = str(scope_info.get("scope_unit_id") or "")
    elif scope_info.get("scope_unit_id"):
        conditions.append("COALESCE(p.scope_unit_id, '') = :scope_unit_id")
        params["scope_unit_id"] = str(scope_info["scope_unit_id"])

    rows = db.execute(
        text(
            f"""
            SELECT i.user_id, i.work_date
            FROM leave_roster_plan_items i
            JOIN leave_roster_plans p ON p.id = i.plan_id
            WHERE {' AND '.join(conditions)}
            ORDER BY i.work_date ASC
            """
        ),
        params,
    ).fetchall()

    out: dict[str, list[str]] = {item["date"]: [] for item in week_days}
    for row in rows:
        item = _row_to_dict(row)
        uid = str(item.get("user_id") or "").strip()
        work_date = str(item.get("work_date") or "").strip()
        if not uid or not work_date:
            continue
        next_day = (datetime.strptime(work_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
        if next_day in out and uid not in out[next_day]:
            out[next_day].append(uid)
    return out


def _merge_blocked_user_ids_by_date(
    leave_map: dict[str, list[str]],
    roster_plus_one_map: dict[str, list[str]],
) -> dict[str, list[str]]:
    all_dates = sorted(set(list(leave_map.keys()) + list(roster_plus_one_map.keys())))
    out: dict[str, list[str]] = {}
    for d in all_dates:
        merged = []
        for uid in leave_map.get(d, []):
            if uid not in merged:
                merged.append(uid)
        for uid in roster_plus_one_map.get(d, []):
            if uid not in merged:
                merged.append(uid)
        out[d] = merged
    return out


def _load_contract_from_leave_schedule(
    db: Session,
    scope_info: dict[str, Any],
    target_date: str | None = None,
    *,
    full_block_scope: bool = False,
    block_code: str = "",
) -> dict[str, Any]:
    if not _table_exists(db, "leave_requests"):
        return {
            "has_contract": False,
            "approved_leave_count": 0,
            "approved_leave_rows": [],
            "message": "Chưa có bảng leave_requests để đọc hợp đồng dữ liệu.",
        }

    conditions = ["status = 'APPROVED'"]
    params: dict[str, Any] = {}

    if target_date:
        conditions.append("start_date <= :target_date")
        conditions.append("end_date >= :target_date")
        params["target_date"] = target_date

    if block_code:
        conditions.append("COALESCE(block_code, '') = :block_code")
        params["block_code"] = block_code

    if not full_block_scope:
        scope_filters = []
        if scope_info.get("scope_subunit_id"):
            scope_filters.append("unit_id = :scope_subunit_id")
            params["scope_subunit_id"] = scope_info["scope_subunit_id"]
        if scope_info.get("scope_unit_id"):
            scope_filters.append("unit_id = :scope_unit_id")
            params["scope_unit_id"] = scope_info["scope_unit_id"]
        if scope_filters:
            conditions.append("(" + " OR ".join(scope_filters) + ")")

    sql = f"""
        SELECT *
        FROM leave_requests
        WHERE {' AND '.join(conditions)}
        ORDER BY start_date DESC, updated_at DESC
    """
    rows = db.execute(text(sql), params).fetchall()
    items = [_row_to_dict(r) for r in rows]

    return {
        "has_contract": True,
        "approved_leave_count": len(items),
        "approved_leave_rows": items[:50],
        "message": "Đã đọc hợp đồng dữ liệu từ leave_schedule (leave_requests đã duyệt).",
    }


def _load_unread_notice_count(db: Session, user_id: str) -> int:
    row = db.execute(
        text(
            """
            SELECT COUNT(*) AS total_count
            FROM work_position_notices
            WHERE user_id = :user_id AND is_read = 0
            """
        ),
        {"user_id": user_id},
    ).fetchone()
    if not row:
        return 0
    return int(_row_to_dict(row).get("total_count") or 0)



def _load_unread_notice_count_by_types(
    db: Session,
    user_id: str,
    notice_types: list[str],
) -> int:
    clean_types = [str(x).strip() for x in notice_types if str(x).strip()]
    if not clean_types:
        return 0

    placeholders = ",".join([f":notice_type_{idx}" for idx in range(len(clean_types))])
    params: dict[str, Any] = {"user_id": str(user_id)}
    for idx, notice_type in enumerate(clean_types):
        params[f"notice_type_{idx}"] = notice_type

    row = db.execute(
        text(
            f"""
            SELECT COUNT(*) AS total_count
            FROM work_position_notices
            WHERE user_id = :user_id
              AND is_read = 0
              AND notice_type IN ({placeholders})
            """
        ),
        params,
    ).fetchone()

    if not row:
        return 0
    return int(_row_to_dict(row).get("total_count") or 0)

def _notice_type_label(value: str) -> str:
    raw = str(value or "").strip().upper()
    return {
        "PLAN_APPROVED": "Phân vị trí",
        "ADJUSTMENT_APPROVED": "Điều chỉnh",
    }.get(raw, raw or "-")


def _load_notice_rows(
    db: Session,
    user_id: str,
    *,
    notice_type: str = "",
    week_monday: str = "",
) -> list[dict[str, Any]]:
    conditions = ["user_id = :user_id"]
    params: dict[str, Any] = {"user_id": user_id}

    if notice_type:
        conditions.append("COALESCE(notice_type, '') = :notice_type")
        params["notice_type"] = notice_type

    if week_monday:
        try:
            week_start_dt = datetime.strptime(week_monday, "%Y-%m-%d")
            week_end = (week_start_dt + timedelta(days=6)).strftime("%Y-%m-%d")
            conditions.append("COALESCE(target_date, '') >= :week_start")
            conditions.append("COALESCE(target_date, '') <= :week_end")
            params["week_start"] = week_monday
            params["week_end"] = week_end
        except Exception as ex:
            logger.warning(
                "[work_positions] week_monday không hợp lệ khi lọc thông báo vị trí: %s | lỗi: %s",
                week_monday,
                ex,
            )

    rows = db.execute(
        text(
            f"""
            SELECT *
            FROM work_position_notices
            WHERE {' AND '.join(conditions)}
            ORDER BY is_read ASC, target_date ASC, created_at DESC
            """
        ),
        params,
    ).fetchall()

    items = [_row_to_dict(r) for r in rows]
    for item in items:
        item["is_unread"] = int(item.get("is_read") or 0) == 0
        item["target_week_label"] = _week_label_from_date_str(str(item.get("target_date") or ""))
        item["notice_type_label"] = _notice_type_label(str(item.get("notice_type") or ""))
    return items
    
def _split_user_ids_csv(raw_value: str | None) -> list[str]:
    out: list[str] = []
    for part in str(raw_value or "").split(","):
        uid = str(part or "").strip()
        if uid and uid not in out:
            out.append(uid)
    return out    

def _create_position_notices_for_plan(
    db: Session,
    *,
    plan_id: str,
    approved_by_name: str,
) -> list[str]:
    plan_row = db.execute(
        text("SELECT * FROM work_position_plans WHERE id = :id"),
        {"id": plan_id},
    ).fetchone()
    if not plan_row:
        return []

    plan = _row_to_dict(plan_row)
    item_rows = db.execute(
        text(
            """
            SELECT assigned_user_id, assigned_user_name, position_name
            FROM work_position_plan_items
            WHERE plan_id = :plan_id
            ORDER BY position_name ASC, updated_at ASC
            """
        ),
        {"plan_id": plan_id},
    ).fetchall()

    scope_label = str(plan.get("scope_subunit_name") or plan.get("scope_unit_name") or "-")
    target_date = str(plan.get("target_date") or "")
    target_week_label = _week_label_from_date_str(target_date)
    created_at = _dt_to_str(_now_vn())

    notice_user_ids: list[str] = []
    for row in item_rows:
        item = _row_to_dict(row)
        assigned_ids = _split_user_ids_csv(item.get("assigned_user_id"))
        position_name = str(item.get("position_name") or "-")

        for uid in assigned_ids:
            db.execute(
                text(
                    """
                    INSERT INTO work_position_notices (
                        id, user_id, target_date, notice_type, title, message,
                        related_plan_id, is_read, created_at, approved_at
                    ) VALUES (
                        :id, :user_id, :target_date, :notice_type, :title, :message,
                        :related_plan_id, 0, :created_at, :approved_at
                    )
                    """
                ),
                {
                    "id": str(uuid.uuid4()),
                    "user_id": uid,
                    "target_date": target_date,
                    "notice_type": "PLAN_APPROVED",
                    "title": "Đã có lịch vị trí làm việc",
                    "message": f"{target_week_label} | {scope_label} | Vị trí: {position_name} | Người duyệt: {approved_by_name}",
                    "related_plan_id": plan_id,
                    "created_at": created_at,
                    "approved_at": created_at,
                },
            )
            if uid not in notice_user_ids:
                notice_user_ids.append(uid)

    return notice_user_ids


def _create_adjustment_notices_for_approved_adjustment(
    db: Session,
    *,
    adjustment_id: str,
    approved_by_name: str,
) -> list[str]:
    adjustment_row = db.execute(
        text("SELECT * FROM work_position_adjustments WHERE id = :id"),
        {"id": adjustment_id},
    ).fetchone()
    if not adjustment_row:
        return []

    adjustment = _row_to_dict(adjustment_row)
    item_rows = _load_adjustment_item_rows(db, adjustment_id)

    created_at = _dt_to_str(_now_vn())
    target_week_label = _week_label_from_date_str(str(adjustment.get("target_date") or ""))
    scope_label = str(adjustment.get("scope_subunit_name") or adjustment.get("scope_unit_name") or "-")
    reason = str(adjustment.get("reason") or "-")

    notice_user_ids: list[str] = []
    inserted_keys: set[tuple[str, str]] = set()

    for item in item_rows:
        item_id = str(item.get("id") or "")
        work_date = str(item.get("work_date") or adjustment.get("target_date") or "")
        session_code = str(item.get("session_code") or "").upper()
        session_label = "Sáng" if session_code == "SANG" else ("Chiều" if session_code == "CHIEU" else session_code or "-")
        position_name = str(item.get("position_name") or "-")
        adjust_type = str(item.get("adjust_type") or adjustment.get("adjust_type") or "")
        adjust_type_label = _adjustment_type_label(adjust_type)

        affected_ids: list[str] = []
        for raw_ids in [
            item.get("old_assigned_user_id"),
            item.get("new_assigned_user_id"),
        ]:
            for uid in _split_user_ids_csv(str(raw_ids or "")):
                if uid and uid not in affected_ids:
                    affected_ids.append(uid)

        if not affected_ids:
            continue

        message = (
            f"{target_week_label} | Ngày: {work_date} | Buổi: {session_label} | "
            f"{scope_label} | Vị trí: {position_name} | "
            f"Nội dung điều chỉnh: {adjust_type_label} | "
            f"Lý do: {reason} | Người duyệt: {approved_by_name}"
        )

        for uid in affected_ids:
            dedup_key = (uid, item_id)
            if dedup_key in inserted_keys:
                continue
            inserted_keys.add(dedup_key)

            db.execute(
                text(
                    """
                    INSERT INTO work_position_notices (
                        id, user_id, target_date, notice_type, title, message,
                        related_plan_id, is_read, created_at, approved_at
                    ) VALUES (
                        :id, :user_id, :target_date, :notice_type, :title, :message,
                        :related_plan_id, 0, :created_at, :approved_at
                    )
                    """
                ),
                {
                    "id": str(uuid.uuid4()),
                    "user_id": uid,
                    "target_date": work_date,
                    "notice_type": "ADJUSTMENT_APPROVED",
                    "title": "Đã có điều chỉnh lịch phân vị trí",
                    "message": message,
                    "related_plan_id": str(item.get("original_plan_id") or adjustment.get("original_plan_id") or ""),
                    "created_at": created_at,
                    "approved_at": created_at,
                },
            )

            if uid not in notice_user_ids:
                notice_user_ids.append(uid)

    return notice_user_ids


def _load_khth_cm_summary(db: Session, target_date: str = "") -> dict[str, Any]:
    if not (_table_exists(db, "work_position_catalog") and _table_exists(db, "work_position_plans")):
        return {
            "room_count": 0,
            "plan_count": 0,
            "approved_plan_count": 0,
            "item_count": 0,
            "room_rows": [],
            "plan_rows": [],
            "leave_rows": [],
            "message": "Chưa có dữ liệu phân vị trí để tổng hợp cho Khối chuyên môn.",
        }

    room_rows = db.execute(
        text(
            """
            SELECT apply_unit_name, position_type, position_name, sort_order
            FROM work_position_catalog
            WHERE is_active = 1
              AND COALESCE(block_code, '') = 'CHUYEN_MON'
            ORDER BY apply_unit_name ASC, sort_order ASC, position_name ASC
            """
        )
    ).fetchall()
    room_items = [_row_to_dict(r) for r in room_rows]

    params: dict[str, Any] = {}
    plan_conditions = ["COALESCE(block_code, '') = 'CHUYEN_MON'"]
    if target_date:
        plan_conditions.append("target_date = :target_date")
        params["target_date"] = target_date

    plan_rows = db.execute(
        text(
            f"""
            SELECT p.*, COUNT(i.id) AS item_count
            FROM work_position_plans p
            LEFT JOIN work_position_plan_items i ON i.plan_id = p.id
            WHERE {' AND '.join(plan_conditions)}
            GROUP BY p.id
            ORDER BY p.target_date DESC, p.scope_unit_name ASC, p.scope_subunit_name ASC, p.updated_at DESC
            """
        ),
        params,
    ).fetchall()
    plan_items = [_row_to_dict(r) for r in plan_rows]

    leave_rows: list[dict[str, Any]] = []
    if _table_exists(db, "leave_requests"):
        leave_conditions = ["status = 'APPROVED'", "COALESCE(block_code, '') = 'CHUYEN_MON'"]
        leave_params: dict[str, Any] = {}
        if target_date:
            leave_conditions.append("start_date <= :target_date")
            leave_conditions.append("end_date >= :target_date")
            leave_params["target_date"] = target_date
        leave_rows = [
            _row_to_dict(r)
            for r in db.execute(
                text(
                    f"""
                    SELECT user_name, unit_name, start_date, end_date, leave_type, symbol
                    FROM leave_requests
                    WHERE {' AND '.join(leave_conditions)}
                    ORDER BY unit_name ASC, user_name ASC, start_date ASC
                    """
                ),
                leave_params,
            ).fetchall()
        ]

    return {
        "room_count": len(room_items),
        "plan_count": len(plan_items),
        "approved_plan_count": sum(1 for item in plan_items if str(item.get("status") or "").upper() == "APPROVED"),
        "item_count": sum(int(item.get("item_count") or 0) for item in plan_items),
        "room_rows": room_items,
        "plan_rows": plan_items,
        "leave_rows": leave_rows[:50],
        "message": "Tổng hợp phạm vi Khối chuyên môn dành cho Trưởng/Phó phòng Kế hoạch tổng hợp.",
    }


def _build_matrix_from_plan_items(
    item_rows: list[dict[str, Any]],
    *,
    matrix_position_type: str = "",
) -> dict[str, Any]:
    filtered = []
    wanted_type = str(matrix_position_type or "").strip().upper()
    for item in item_rows:
        item_type = str(item.get("position_type") or "").strip().upper()
        if wanted_type and item_type != wanted_type:
            continue
        filtered.append(item)

    work_dates = sorted({str(item.get("work_date") or "") for item in filtered if str(item.get("work_date") or "").strip()})
    rows_map: dict[str, dict[str, Any]] = {}

    for item in filtered:
        row_key = str(item.get("catalog_id") or "")
        if not row_key:
            continue
        if row_key not in rows_map:
            rows_map[row_key] = {
                "catalog_id": row_key,
                "apply_unit_name": item.get("apply_unit_name") or item.get("owner_unit_name") or "-",
                "position_name": item.get("position_name") or "-",
                "cells": {},
            }

        cell_key = "|".join([
            str(item.get("work_date") or ""),
            str(item.get("session_code") or ""),
            str(item.get("staff_group") or ""),
        ])
        rows_map[row_key]["cells"][cell_key] = {
            "assigned_user_id": str(item.get("assigned_user_id") or ""),
            "assigned_user_name": str(item.get("assigned_user_name") or ""),
            "notes": str(item.get("notes") or ""),
            "is_adjusted": bool(item.get("is_adjusted")),
            "adjustment_note": str(item.get("adjustment_note") or ""),
        }

    rows = list(rows_map.values())
    rows.sort(key=lambda x: (str(x.get("apply_unit_name") or ""), str(x.get("position_name") or "")))

    return {
        "work_dates": work_dates,
        "rows": rows,
        "position_type": wanted_type,
    }

def _load_approved_adjustment_item_rows(
    db: Session,
    *,
    target_week_monday: str = "",
    original_plan_ids: list[str] | None = None,
) -> list[dict[str, Any]]:
    conditions = ["a.status = 'KHTH_APPROVED'"]
    params: dict[str, Any] = {}

    if target_week_monday:
        conditions.append("a.target_date = :target_date")
        params["target_date"] = target_week_monday

    plan_ids = [str(x) for x in (original_plan_ids or []) if str(x or "").strip()]
    if plan_ids:
        placeholders = ",".join([f":p{i}" for i in range(len(plan_ids))])
        conditions.append(f"a.original_plan_id IN ({placeholders})")
        for i, plan_id in enumerate(plan_ids):
            params[f"p{i}"] = plan_id

    rows = db.execute(
        text(
            f"""
            SELECT
                ai.*,
                a.status AS adjustment_status,
                a.reason AS adjustment_reason,
                a.khth_approved_by_name,
                a.khth_approved_at
            FROM work_position_adjustment_items ai
            JOIN work_position_adjustments a ON a.id = ai.adjustment_id
            WHERE {' AND '.join(conditions)}
            ORDER BY a.khth_approved_at ASC, ai.created_at ASC
            """
        ),
        params,
    ).fetchall()

    return [_row_to_dict(r) for r in rows]


def _apply_approved_adjustments_to_items(
    item_rows: list[dict[str, Any]],
    adjustment_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    if not item_rows or not adjustment_rows:
        return item_rows

    out: list[dict[str, Any]] = [dict(item) for item in item_rows]
    item_map: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}

    for item in out:
        key = (
            str(item.get("plan_id") or ""),
            str(item.get("catalog_id") or ""),
            str(item.get("work_date") or ""),
            str(item.get("session_code") or ""),
            str(item.get("staff_group") or ""),
        )
        item_map[key] = item

    for adj in adjustment_rows:
        adjust_type = str(adj.get("adjust_type") or "").strip().upper()
        base_key = (
            str(adj.get("original_plan_id") or ""),
            str(adj.get("catalog_id") or ""),
            str(adj.get("work_date") or ""),
            str(adj.get("session_code") or ""),
        )
        staff_group = str(adj.get("staff_group") or "").strip().upper()

        target_groups: list[str]
        if adjust_type == "CLOSE_ROOM":
            target_groups = ["BS", "DD"]
        elif adjust_type == "CHANGE_BS":
            target_groups = ["BS"]
        elif adjust_type == "CHANGE_DD":
            target_groups = ["DD"]
        elif adjust_type == "CHANGE_BOTH":
            target_groups = ["BS", "DD"]
        else:
            target_groups = [staff_group] if staff_group else []

        for group_code in target_groups:
            key = base_key + (group_code,)
            item = item_map.get(key)
            if not item:
                continue

            if adjust_type == "CLOSE_ROOM":
                item["assigned_user_id"] = ""
                item["assigned_user_name"] = ""
            elif group_code == staff_group:
                item["assigned_user_id"] = str(adj.get("new_assigned_user_id") or "")
                item["assigned_user_name"] = str(adj.get("new_assigned_user_name") or "")

            item["is_adjusted"] = True
            item["adjustment_note"] = (
                f"Đã điều chỉnh: {_adjustment_type_label(adjust_type)}"
                f" | Lý do: {adj.get('adjustment_reason') or '-'}"
                f" | Duyệt: {adj.get('khth_approved_by_name') or '-'}"
                f" lúc {adj.get('khth_approved_at') or '-'}"
            )

    return out


def _load_adjustment_rows(
    db: Session,
    *,
    scope_info: dict[str, Any] | None = None,
    target_week_monday: str = "",
    hospital_scope: bool = False,
    status_filter: str = "",
    history_year: str = "",
    history_month: str = "",
) -> list[dict[str, Any]]:
    conditions = ["1=1"]
    params: dict[str, Any] = {}

    if status_filter:
        conditions.append("status = :status")
        params["status"] = status_filter

    if target_week_monday:
        conditions.append("target_date = :target_date")
        params["target_date"] = target_week_monday

    if history_year:
        conditions.append("substr(target_date, 1, 4) = :history_year")
        params["history_year"] = history_year

    if history_month:
        conditions.append("substr(target_date, 6, 2) = :history_month")
        params["history_month"] = f"{int(history_month):02d}"

    if not hospital_scope and scope_info:
        scope_filters = []
        if scope_info.get("scope_subunit_id"):
            scope_filters.append("scope_subunit_id = :scope_subunit_id")
            params["scope_subunit_id"] = scope_info["scope_subunit_id"]
        if scope_info.get("scope_unit_id"):
            scope_filters.append("scope_unit_id = :scope_unit_id")
            params["scope_unit_id"] = scope_info["scope_unit_id"]
        if scope_filters:
            conditions.append("(" + " OR ".join(scope_filters) + ")")

    rows = db.execute(
        text(
            f"""
            SELECT *
            FROM work_position_adjustments
            WHERE {' AND '.join(conditions)}
            ORDER BY target_date DESC, requested_at DESC, updated_at DESC
            """
        ),
        params,
    ).fetchall()

    items = [_row_to_dict(r) for r in rows]
    for item in items:
        item["status_label"] = _adjustment_status_label(str(item.get("status") or ""))
        item["adjust_type_label"] = _adjustment_type_label(str(item.get("adjust_type") or ""))
        item["scope_label"] = item.get("scope_subunit_name") or item.get("scope_unit_name") or "-"
        item["week_value"] = _week_value_from_date_str(str(item.get("target_date") or ""))
        item["week_label"] = _week_label_from_date_str(str(item.get("target_date") or ""))
    return items


def _load_adjustment_item_rows(db: Session, adjustment_id: str) -> list[dict[str, Any]]:
    rows = db.execute(
        text(
            """
            SELECT *
            FROM work_position_adjustment_items
            WHERE adjustment_id = :adjustment_id
            ORDER BY work_date ASC, session_code ASC, staff_group ASC
            """
        ),
        {"adjustment_id": adjustment_id},
    ).fetchall()

    items = [_row_to_dict(r) for r in rows]
    for item in items:
        item["adjust_type_label"] = _adjustment_type_label(str(item.get("adjust_type") or ""))
    return items


def _load_clinic_adjustment_targets(
    db: Session,
    scope_info: dict[str, Any],
    target_week_monday: str,
) -> list[dict[str, Any]]:
    rows = db.execute(
        text(
            """
            SELECT
                p.id AS plan_id,
                p.target_date,
                p.scope_unit_id,
                p.scope_unit_name,
                p.scope_subunit_id,
                p.scope_subunit_name,
                i.id AS item_id,
                i.catalog_id,
                i.position_name,
                i.work_date,
                i.session_code,
                i.staff_group,
                i.assigned_user_id,
                i.assigned_user_name,
                c.apply_unit_name,
                c.position_type
            FROM work_position_plan_items i
            JOIN work_position_plans p ON p.id = i.plan_id
            LEFT JOIN work_position_catalog c ON c.id = i.catalog_id
            WHERE p.status = 'KHTH_APPROVED'
              AND p.target_date = :target_date
              AND COALESCE(c.position_type, '') = 'PHONG_KHAM'
              AND (
                    p.scope_subunit_id = :scope_subunit_id
                    OR p.scope_unit_id = :scope_unit_id
                  )
            ORDER BY c.apply_unit_name ASC, i.position_name ASC, i.work_date ASC, i.session_code ASC, i.staff_group ASC
            """
        ),
        {
            "target_date": target_week_monday,
            "scope_subunit_id": str(scope_info.get("scope_subunit_id") or ""),
            "scope_unit_id": str(scope_info.get("scope_unit_id") or ""),
        },
    ).fetchall()

    raw_items = [_row_to_dict(r) for r in rows]
    grouped: dict[tuple[str, str, str, str], dict[str, Any]] = {}

    for item in raw_items:
        key = (
            str(item.get("plan_id") or ""),
            str(item.get("catalog_id") or ""),
            str(item.get("work_date") or ""),
            str(item.get("session_code") or ""),
        )
        if key not in grouped:
            grouped[key] = {
                "plan_id": item.get("plan_id"),
                "target_date": item.get("target_date"),
                "scope_unit_name": item.get("scope_unit_name"),
                "scope_subunit_name": item.get("scope_subunit_name"),
                "catalog_id": item.get("catalog_id"),
                "position_name": item.get("position_name"),
                "apply_unit_name": item.get("apply_unit_name"),
                "work_date": item.get("work_date"),
                "session_code": item.get("session_code"),
                "bs_item_id": "",
                "bs_user_id": "",
                "bs_user_name": "",
                "dd_item_id": "",
                "dd_user_id": "",
                "dd_user_name": "",
            }

        staff_group = str(item.get("staff_group") or "").strip().upper()
        if staff_group == "BS":
            grouped[key]["bs_item_id"] = item.get("item_id") or ""
            grouped[key]["bs_user_id"] = item.get("assigned_user_id") or ""
            grouped[key]["bs_user_name"] = item.get("assigned_user_name") or ""
        elif staff_group == "DD":
            grouped[key]["dd_item_id"] = item.get("item_id") or ""
            grouped[key]["dd_user_id"] = item.get("assigned_user_id") or ""
            grouped[key]["dd_user_name"] = item.get("assigned_user_name") or ""

    return list(grouped.values())


def _load_unit_approve_summary(
    db: Session,
    scope_info: dict[str, Any],
    block_tab: str,
    target_week_monday: str,
) -> dict[str, Any]:
    plan_rows = _load_position_plan_rows(
        db,
        scope_info,
        block_tab,
        full_block_scope=False,
        target_week_monday=target_week_monday,
    )

    all_plan_items: list[dict[str, Any]] = []
    for plan in plan_rows:
        plan_id = str(plan.get("id") or "")
        if not plan_id:
            continue
        all_plan_items.extend(_load_plan_item_rows(db, plan_id))

    matrix_phong_kham = _build_matrix_from_plan_items(all_plan_items, matrix_position_type="PHONG_KHAM")
    matrix_phong_dieu_tri = _build_matrix_from_plan_items(all_plan_items, matrix_position_type="PHONG_DIEU_TRI")
    matrix_xet_nghiem = _build_matrix_from_plan_items(all_plan_items, matrix_position_type="XET_NGHIEM")

    return {
        "plan_rows": plan_rows,
        "plan_items": all_plan_items,
        "matrix_phong_kham": matrix_phong_kham,
        "matrix_phong_dieu_tri": matrix_phong_dieu_tri,
        "matrix_xet_nghiem": matrix_xet_nghiem,
    }


def _load_khth_approve_summary(
    db: Session,
    target_week_monday: str,
) -> dict[str, Any]:
    rows = db.execute(
        text(
            """
            SELECT *
            FROM work_position_plans
            WHERE COALESCE(block_code, '') = 'CHUYEN_MON'
              AND target_date = :target_date
              AND status IN ('UNIT_APPROVED', 'KHTH_APPROVED')
            ORDER BY scope_unit_name ASC, scope_subunit_name ASC, updated_at DESC
            """
        ),
        {"target_date": target_week_monday},
    ).fetchall()

    raw_plan_rows = [_row_to_dict(r) for r in rows]
    plan_rows: list[dict[str, Any]] = []
    all_item_rows: list[dict[str, Any]] = []

    for plan in raw_plan_rows:
        plan_id = str(plan.get("id") or "")
        if not plan_id:
            continue
        item_rows = _load_plan_item_rows(db, plan_id)
        if not _plan_needs_khth_approval("CHUYEN_MON", item_rows):
            continue
        plan["status_label"] = _approval_status_label(str(plan.get("status") or ""))
        plan["scope_label"] = plan.get("scope_subunit_name") or plan.get("scope_unit_name") or "-"
        plan["week_value"] = _week_value_from_date_str(str(plan.get("target_date") or ""))
        plan["week_label"] = _week_label_from_date_str(str(plan.get("target_date") or ""))
        plan["item_count"] = len(item_rows)
        plan_rows.append(plan)
        all_item_rows.extend(item_rows)

    plan_ids = [str(plan.get("id") or "") for plan in plan_rows if str(plan.get("id") or "").strip()]
    approved_adjustment_items = _load_approved_adjustment_item_rows(
        db,
        target_week_monday=target_week_monday,
        original_plan_ids=plan_ids,
    )
    effective_item_rows = _apply_approved_adjustments_to_items(all_item_rows, approved_adjustment_items)

    matrix_phong_kham = _build_matrix_from_plan_items(effective_item_rows, matrix_position_type="PHONG_KHAM")
    matrix_xet_nghiem = _build_matrix_from_plan_items(effective_item_rows, matrix_position_type="XET_NGHIEM")

    return {
        "plans": plan_rows,
        "matrix_phong_kham": matrix_phong_kham,
        "matrix_xet_nghiem": matrix_xet_nghiem,
    }


def _split_assigned_people(user_ids_raw: str, user_names_raw: str) -> list[dict[str, str]]:
    user_ids = [x.strip() for x in str(user_ids_raw or "").split(",") if x.strip()]
    user_names = [x.strip() for x in str(user_names_raw or "").split(",") if x.strip()]

    max_len = max(len(user_ids), len(user_names))
    people: list[dict[str, str]] = []
    for idx in range(max_len):
        people.append({
            "user_id": user_ids[idx] if idx < len(user_ids) else "",
            "name": user_names[idx] if idx < len(user_names) else "",
        })
    return people


def _clinic_cell_payload(cell: dict[str, Any] | None) -> dict[str, Any]:
    cell = cell or {}
    return {
        "assigned_user_id": str(cell.get("assigned_user_id") or ""),
        "assigned_user_name": str(cell.get("assigned_user_name") or ""),
        "people": _split_assigned_people(
            str(cell.get("assigned_user_id") or ""),
            str(cell.get("assigned_user_name") or ""),
        ),
        "is_adjusted": bool(cell.get("is_adjusted")),
        "adjustment_note": str(cell.get("adjustment_note") or ""),
        "notes": str(cell.get("notes") or ""),
    }


def _build_hospital_clinic_matrix_payload(db: Session, target_week_monday: str) -> dict[str, Any]:
    week_value = _week_value_from_date_str(target_week_monday)
    week_days = _build_week_days(target_week_monday)
    summary = _load_khth_approve_summary(db, target_week_monday)
    matrix = summary.get("matrix_phong_kham") or {}
    matrix_rows = matrix.get("rows") or []

    clinics: list[dict[str, Any]] = []
    for row in matrix_rows:
        cells = row.get("cells") or {}
        clinic_days: list[dict[str, Any]] = []

        for day in week_days:
            day_date = str(day.get("date") or "")
            day_payload = {
                "date": day_date,
                "label": day.get("label") or "",
                "display": day.get("display") or "",
                "shifts": [],
            }

            for session_code, session_label in [("SANG", "Sáng"), ("CHIEU", "Chiều")]:
                bs_cell = _clinic_cell_payload(cells.get(f"{day_date}|{session_code}|BS"))
                dd_cell = _clinic_cell_payload(cells.get(f"{day_date}|{session_code}|DD"))

                has_people = bool(bs_cell["assigned_user_name"] or dd_cell["assigned_user_name"])
                is_adjusted = bool(bs_cell["is_adjusted"] or dd_cell["is_adjusted"])

                if has_people and is_adjusted:
                    status = "OPEN_ADJUSTED"
                    status_label = "Hoạt động sau điều chỉnh"
                elif has_people:
                    status = "OPEN"
                    status_label = "Hoạt động"
                elif is_adjusted:
                    status = "CLOSED_ADJUSTED"
                    status_label = "Đóng phòng theo điều chỉnh"
                else:
                    status = "CLOSED"
                    status_label = "Đóng phòng"

                adjustment_notes = [
                    note for note in [bs_cell.get("adjustment_note"), dd_cell.get("adjustment_note")]
                    if str(note or "").strip()
                ]

                day_payload["shifts"].append({
                    "code": session_code,
                    "label": session_label,
                    "status": status,
                    "status_label": status_label,
                    "is_adjusted": is_adjusted,
                    "adjustment_note": " | ".join(adjustment_notes),
                    "doctor": bs_cell,
                    "nurse": dd_cell,
                })

            clinic_days.append(day_payload)

        clinics.append({
            "catalog_id": str(row.get("catalog_id") or ""),
            "unit_name": str(row.get("apply_unit_name") or ""),
            "room_name": str(row.get("position_name") or ""),
            "days": clinic_days,
        })

    generated_at = _dt_to_str(_now_vn()) or ""
    return {
        "schema_version": "1.0",
        "metadata": {
            "hospital_code": "HVGL",
            "hospital_name": "Bệnh viện Hùng Vương Gia Lai",
            "week": week_value,
            "from_date": target_week_monday,
            "to_date": (datetime.strptime(target_week_monday, "%Y-%m-%d") + timedelta(days=6)).strftime("%Y-%m-%d"),
            "position_type": "PHONG_KHAM",
            "data_status": "EFFECTIVE",
            "generated_at": generated_at,
            "source_system": "HVGL_QuanLyDieuHanh",
            "clinic_count": len(clinics),
        },
        "clinics": clinics,
    }


def _append_people_xml(parent: Any, tag_name: str, staff_payload: dict[str, Any]) -> None:
    holder = ET.SubElement(
        parent,
        tag_name,
        {
            "assigned_user_id": str(staff_payload.get("assigned_user_id") or ""),
            "assigned_user_name": str(staff_payload.get("assigned_user_name") or ""),
        },
    )

    for person in staff_payload.get("people") or []:
        ET.SubElement(
            holder,
            "person",
            {
                "user_id": str(person.get("user_id") or ""),
                "name": str(person.get("name") or ""),
            },
        )


def _build_hospital_clinic_matrix_xml(payload: dict[str, Any]) -> bytes:
    metadata = payload.get("metadata") or {}

    root = ET.Element(
        "hospital_clinic_matrix",
        {
            "schema_version": str(payload.get("schema_version") or "1.0"),
        },
    )

    meta_el = ET.SubElement(root, "metadata")
    ET.SubElement(meta_el, "hospital_code").text = str(metadata.get("hospital_code") or "")
    ET.SubElement(meta_el, "hospital_name").text = str(metadata.get("hospital_name") or "")
    ET.SubElement(
        meta_el,
        "week",
        {
            "value": str(metadata.get("week") or ""),
            "from": str(metadata.get("from_date") or ""),
            "to": str(metadata.get("to_date") or ""),
        },
    )
    ET.SubElement(meta_el, "position_type").text = str(metadata.get("position_type") or "")
    ET.SubElement(meta_el, "data_status").text = str(metadata.get("data_status") or "")
    ET.SubElement(meta_el, "generated_at").text = str(metadata.get("generated_at") or "")
    ET.SubElement(meta_el, "source_system").text = str(metadata.get("source_system") or "")
    ET.SubElement(meta_el, "clinic_count").text = str(metadata.get("clinic_count") or 0)

    clinics_el = ET.SubElement(root, "clinics")
    for clinic in payload.get("clinics") or []:
        clinic_el = ET.SubElement(
            clinics_el,
            "clinic",
            {
                "catalog_id": str(clinic.get("catalog_id") or ""),
            },
        )
        ET.SubElement(clinic_el, "unit_name").text = str(clinic.get("unit_name") or "")
        ET.SubElement(clinic_el, "room_name").text = str(clinic.get("room_name") or "")

        days_el = ET.SubElement(clinic_el, "days")
        for day in clinic.get("days") or []:
            day_el = ET.SubElement(
                days_el,
                "day",
                {
                    "date": str(day.get("date") or ""),
                    "label": str(day.get("label") or ""),
                    "display": str(day.get("display") or ""),
                },
            )

            for shift in day.get("shifts") or []:
                shift_el = ET.SubElement(
                    day_el,
                    "shift",
                    {
                        "code": str(shift.get("code") or ""),
                        "label": str(shift.get("label") or ""),
                    },
                )
                ET.SubElement(shift_el, "status").text = str(shift.get("status") or "")
                ET.SubElement(shift_el, "status_label").text = str(shift.get("status_label") or "")
                ET.SubElement(shift_el, "is_adjusted").text = "1" if shift.get("is_adjusted") else "0"
                ET.SubElement(shift_el, "adjustment_note").text = str(shift.get("adjustment_note") or "")
                _append_people_xml(shift_el, "doctor", shift.get("doctor") or {})
                _append_people_xml(shift_el, "nurse", shift.get("nurse") or {})

    ET.indent(root, space="  ")
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _xlsx_safe_text(value: Any) -> str:
    return str(value or "").strip()


def _xlsx_safe_sheet_title(value: str) -> str:
    raw = str(value or "Sheet").strip() or "Sheet"
    for ch in ["\\", "/", "*", "?", ":", "[", "]"]:
        raw = raw.replace(ch, "_")
    return raw[:31] or "Sheet"


def _xlsx_safe_filename(value: str) -> str:
    raw = str(value or "export.xlsx").strip() or "export.xlsx"
    safe = []
    for ch in raw:
        if ch.isalnum() or ch in {"-", "_", "."}:
            safe.append(ch)
        else:
            safe.append("_")
    out = "".join(safe).strip("_")
    if not out.lower().endswith(".xlsx"):
        out += ".xlsx"
    return out


def _xlsx_cell_display(cell: dict[str, Any] | None, *, include_adjust_mark: bool = True) -> str:
    cell = cell or {}
    name = _xlsx_safe_text(cell.get("assigned_user_name"))
    notes = _xlsx_safe_text(cell.get("notes"))
    adjustment_note = _xlsx_safe_text(cell.get("adjustment_note"))
    is_adjusted = bool(cell.get("is_adjusted"))

    parts: list[str] = []
    if name:
        parts.append(name)
    if notes:
        parts.append(notes)
    if include_adjust_mark and is_adjusted:
        parts.append("*Đã điều chỉnh")
    if adjustment_note:
        parts.append(adjustment_note)

    return "\n".join(parts)


def _xlsx_apply_common_styles(ws) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color="94A3B8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    title_fill = PatternFill("solid", fgColor="0B4F9C")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    sub_header_fill = PatternFill("solid", fgColor="EFF6FF")
    bs_fill = PatternFill("solid", fgColor="CFE2F3")
    dd_fill = PatternFill("solid", fgColor="FFF2CC")
    closed_fill = PatternFill("solid", fgColor="F3F4F6")

    for row_idx in [1, 2]:
        for cell in ws[row_idx]:
            cell.fill = title_fill
            cell.font = title_font

    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = Font(bold=True)

    for cell in ws[5]:
        cell.fill = sub_header_fill
        cell.font = Font(bold=True)

    for row in ws.iter_rows(min_row=6):
        for cell in row:
            if cell.column <= 2:
                cell.font = Font(bold=True)
                cell.fill = sub_header_fill
            elif "Đóng phòng" in str(cell.value or ""):
                cell.font = Font(bold=True, color="991B1B")
                cell.fill = closed_fill
            elif cell.column >= 3:
                sub_head = str(ws.cell(row=5, column=cell.column).value or "").upper()
                if "BS" in sub_head:
                    cell.fill = bs_fill
                elif "ĐD" in sub_head or "TKYK" in sub_head:
                    cell.fill = dd_fill

    ws.freeze_panes = "C6"


def _xlsx_autofit_columns(ws) -> None:
    from openpyxl.utils import get_column_letter

    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_len = 0

        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = str(cell.value or "")
            if not value:
                continue

            line_lengths = [len(line) for line in value.splitlines()] or [0]
            max_len = max(max_len, min(max(line_lengths), 45))

        ws.column_dimensions[column_letter].width = max(12, min(max_len + 3, 50))


def _write_four_slot_matrix_sheet(
    wb,
    *,
    sheet_title: str,
    title: str,
    subtitle: str,
    matrix: dict[str, Any],
    first_col_title: str,
    second_col_title: str,
    show_closed_room: bool = False,
) -> None:
    ws = wb.create_sheet(title=_xlsx_safe_sheet_title(sheet_title))

    work_dates = list(matrix.get("work_dates") or [])
    rows = list(matrix.get("rows") or [])

    total_cols = 2 + len(work_dates) * 4
    if total_cols < 2:
        total_cols = 2

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.cell(row=1, column=1).value = title
    ws.cell(row=2, column=1).value = subtitle

    ws.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)
    ws.merge_cells(start_row=4, start_column=2, end_row=5, end_column=2)
    ws.cell(row=4, column=1).value = first_col_title
    ws.cell(row=4, column=2).value = second_col_title

    col = 3
    for d in work_dates:
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 3)
        ws.cell(row=4, column=col).value = d
        ws.cell(row=5, column=col).value = "Sáng\nBS"
        ws.cell(row=5, column=col + 1).value = "Sáng\nĐD/TKYK"
        ws.cell(row=5, column=col + 2).value = "Chiều\nBS"
        ws.cell(row=5, column=col + 3).value = "Chiều\nĐD/TKYK"
        col += 4

    row_idx = 6
    for item in rows:
        ws.cell(row=row_idx, column=1).value = _xlsx_safe_text(item.get("apply_unit_name")) or "-"
        ws.cell(row=row_idx, column=2).value = _xlsx_safe_text(item.get("position_name")) or "-"

        col = 3
        cells = item.get("cells") or {}
        for d in work_dates:
            sang_bs = cells.get(f"{d}|SANG|BS") or {}
            sang_dd = cells.get(f"{d}|SANG|DD") or {}
            chieu_bs = cells.get(f"{d}|CHIEU|BS") or {}
            chieu_dd = cells.get(f"{d}|CHIEU|DD") or {}

            sang_bs_text = _xlsx_cell_display(sang_bs)
            sang_dd_text = _xlsx_cell_display(sang_dd)
            chieu_bs_text = _xlsx_cell_display(chieu_bs)
            chieu_dd_text = _xlsx_cell_display(chieu_dd)

            if show_closed_room and not sang_bs_text and not sang_dd_text:
                ws.merge_cells(start_row=row_idx, start_column=col, end_row=row_idx, end_column=col + 1)
                ws.cell(row=row_idx, column=col).value = "Đóng phòng"
            else:
                ws.cell(row=row_idx, column=col).value = sang_bs_text
                ws.cell(row=row_idx, column=col + 1).value = sang_dd_text

            if show_closed_room and not chieu_bs_text and not chieu_dd_text:
                ws.merge_cells(start_row=row_idx, start_column=col + 2, end_row=row_idx, end_column=col + 3)
                ws.cell(row=row_idx, column=col + 2).value = "Đóng phòng"
            else:
                ws.cell(row=row_idx, column=col + 2).value = chieu_bs_text
                ws.cell(row=row_idx, column=col + 3).value = chieu_dd_text

            col += 4

        row_idx += 1

    if not rows:
        ws.cell(row=6, column=1).value = "Không có dữ liệu."

    _xlsx_apply_common_styles(ws)
    _xlsx_autofit_columns(ws)


def _write_lab_matrix_sheet(
    wb,
    *,
    sheet_title: str,
    title: str,
    subtitle: str,
    matrix: dict[str, Any],
) -> None:
    ws = wb.create_sheet(title=_xlsx_safe_sheet_title(sheet_title))

    work_dates = list(matrix.get("work_dates") or [])
    rows = list(matrix.get("rows") or [])
    total_cols = 2 + len(work_dates)
    if total_cols < 2:
        total_cols = 2

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.cell(row=1, column=1).value = title
    ws.cell(row=2, column=1).value = subtitle

    ws.cell(row=4, column=1).value = "Đơn vị"
    ws.cell(row=4, column=2).value = "Vị trí xét nghiệm"

    col = 3
    for d in work_dates:
        ws.cell(row=4, column=col).value = d
        col += 1

    row_idx = 5
    for item in rows:
        ws.cell(row=row_idx, column=1).value = _xlsx_safe_text(item.get("apply_unit_name")) or "-"
        ws.cell(row=row_idx, column=2).value = _xlsx_safe_text(item.get("position_name")) or "-"

        col = 3
        cells = item.get("cells") or {}
        for d in work_dates:
            ws.cell(row=row_idx, column=col).value = _xlsx_cell_display(cells.get(f"{d}|ALLDAY|LAB"))
            col += 1

        row_idx += 1

    if not rows:
        ws.cell(row=5, column=1).value = "Không có dữ liệu."

    _xlsx_apply_common_styles(ws)
    _xlsx_autofit_columns(ws)
    ws.freeze_panes = "C5"


def _build_unit_matrix_xlsx_bytes(
    *,
    summary: dict[str, Any],
    target_week_label: str,
    block_label: str,
    scope_display: str,
) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    title = "BẢNG TỔNG HỢP MA TRẬN PHÂN VỊ TRÍ CẤP ĐƠN VỊ"
    subtitle = f"{block_label} - {scope_display} - {target_week_label}"

    matrix_phong_kham = summary.get("matrix_phong_kham") or {}
    matrix_phong_dieu_tri = summary.get("matrix_phong_dieu_tri") or {}
    matrix_xet_nghiem = summary.get("matrix_xet_nghiem") or {}

    if matrix_phong_kham.get("rows"):
        _write_four_slot_matrix_sheet(
            wb,
            sheet_title="PHONG_KHAM",
            title=title,
            subtitle=subtitle,
            matrix=matrix_phong_kham,
            first_col_title="Số phòng / Đơn vị",
            second_col_title="Phòng khám / Tên vị trí",
            show_closed_room=False,
        )

    if matrix_phong_dieu_tri.get("rows"):
        _write_four_slot_matrix_sheet(
            wb,
            sheet_title="PHONG_DIEU_TRI",
            title=title,
            subtitle=subtitle,
            matrix=matrix_phong_dieu_tri,
            first_col_title="Đơn vị",
            second_col_title="Phòng điều trị / Tên vị trí",
            show_closed_room=False,
        )

    if matrix_xet_nghiem.get("rows"):
        _write_lab_matrix_sheet(
            wb,
            sheet_title="XET_NGHIEM",
            title=title,
            subtitle=subtitle,
            matrix=matrix_xet_nghiem,
        )

    if not wb.sheetnames:
        ws = wb.create_sheet(title="Khong_co_du_lieu")
        ws["A1"] = title
        ws["A2"] = subtitle
        ws["A4"] = "Không có dữ liệu phân vị trí trong tuần đang chọn."

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _build_hospital_clinic_matrix_xlsx_bytes(
    *,
    summary: dict[str, Any],
    target_week_label: str,
) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    _write_four_slot_matrix_sheet(
        wb,
        sheet_title="PHONG_KHAM_BV",
        title="MA TRẬN TỔNG HỢP PHÒNG KHÁM CẤP BỆNH VIỆN",
        subtitle=target_week_label,
        matrix=summary.get("matrix_phong_kham") or {},
        first_col_title="Đơn vị",
        second_col_title="Tên vị trí / Phòng khám",
        show_closed_room=True,
    )

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _xlsx_response(content: bytes, filename: str) -> StreamingResponse:
    safe_filename = _xlsx_safe_filename(filename)
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_filename}"'},
    )

@router.get("/api/nav-badge")
def work_position_nav_badge(request: Request, db: Session = Depends(get_db)):
    _ensure_tables(db)
    user = login_required(request, db)
    count = _load_unread_notice_count(db, user.id)
    return JSONResponse({"ok": True, "unread_work_position_count": count})

@router.get("/export/unit-matrix.xlsx")
def export_unit_matrix_xlsx(
    request: Request,
    block_tab: str = "",
    target_week: str = "",
    scope_unit_id: str = "",
    db: Session = Depends(get_db),
):
    _ensure_tables(db)

    user = login_required(request, db)
    scope_info = _scope_payload(db, user, scope_unit_id)
    role_codes = set(scope_info.get("role_codes") or [])

    if not _can_assign_positions(role_codes):
        raise HTTPException(status_code=403, detail="Bạn không có quyền xuất Excel ma trận cấp đơn vị.")

    current_block_tab = block_tab if block_tab in {"HANH_CHINH", "CHUYEN_MON"} else (scope_info.get("block_code") or "CHUYEN_MON")
    raw_target_week = str(target_week or "").strip() or _current_week_value()
    target_week_monday = _parse_week_value(raw_target_week).strftime("%Y-%m-%d")

    summary = _load_unit_approve_summary(
        db,
        scope_info,
        current_block_tab,
        target_week_monday,
    )

    target_week_label = _week_label_from_date_str(target_week_monday)
    content = _build_unit_matrix_xlsx_bytes(
        summary=summary,
        target_week_label=target_week_label,
        block_label=_block_label(current_block_tab),
        scope_display=str(scope_info.get("scope_display") or "-"),
    )

    filename = f"ma_tran_phan_vi_tri_don_vi_{raw_target_week}.xlsx"
    return _xlsx_response(content, filename)


@router.get("/export/hospital-clinic-matrix.xlsx")
def export_hospital_clinic_matrix_xlsx(
    request: Request,
    target_week: str = "",
    db: Session = Depends(get_db),
):
    _ensure_tables(db)

    user = login_required(request, db)
    role_codes = _get_role_codes(db, user.id)
    if not _can_view_hospital_position_summary(db, user, role_codes):
        raise HTTPException(status_code=403, detail="Bạn không có quyền xuất Excel ma trận phòng khám cấp Bệnh viện.")

    raw_target_week = str(target_week or "").strip() or _current_week_value()
    target_week_monday = _parse_week_value(raw_target_week).strftime("%Y-%m-%d")
    summary = _load_khth_approve_summary(db, target_week_monday)

    content = _build_hospital_clinic_matrix_xlsx_bytes(
        summary=summary,
        target_week_label=_week_label_from_date_str(target_week_monday),
    )

    filename = f"ma_tran_phong_kham_benh_vien_{raw_target_week}.xlsx"
    return _xlsx_response(content, filename)


@router.get("/api/hospital-clinic-matrix")
def hospital_clinic_matrix_api(
    request: Request,
    target_week: str = "",
    db: Session = Depends(get_db),
):
    _ensure_tables(db)

    user = login_required(request, db)
    role_codes = _get_role_codes(db, user.id)
    if not _can_view_hospital_position_summary(db, user, role_codes):
        raise HTTPException(status_code=403, detail="Bạn không có quyền xem ma trận phòng khám cấp Bệnh viện.")

    raw_target_week = str(target_week or "").strip() or _current_week_value()
    target_week_monday = _parse_week_value(raw_target_week).strftime("%Y-%m-%d")
    payload = _build_hospital_clinic_matrix_payload(db, target_week_monday)

    return JSONResponse({"ok": True, "data": payload})


@router.get("/export/hospital-clinic-matrix.xml")
def export_hospital_clinic_matrix_xml(
    request: Request,
    target_week: str = "",
    db: Session = Depends(get_db),
):
    _ensure_tables(db)

    user = login_required(request, db)
    role_codes = _get_role_codes(db, user.id)
    if not _can_view_hospital_position_summary(db, user, role_codes):
        raise HTTPException(status_code=403, detail="Bạn không có quyền xuất XML ma trận phòng khám cấp Bệnh viện.")

    raw_target_week = str(target_week or "").strip() or _current_week_value()
    target_week_monday = _parse_week_value(raw_target_week).strftime("%Y-%m-%d")
    payload = _build_hospital_clinic_matrix_payload(db, target_week_monday)
    xml_bytes = _build_hospital_clinic_matrix_xml(payload)

    safe_week = str(payload.get("metadata", {}).get("week") or raw_target_week).replace("/", "-").replace("\\", "-")
    filename = f"hospital_clinic_matrix_{safe_week}.xml"

    return Response(
        content=xml_bytes,
        media_type="application/xml; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("")
def work_positions_index(
    request: Request,
    main_tab: str = "positions",
    block_tab: str = "",
    sub_tab: str = "",
    selected_position_type: str = "",
    target_week: str = "",
    scope_unit_id: str = "",
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    scope_info = _scope_payload(db, user, scope_unit_id)
    current_main_tab = "positions" if main_tab == "positions" else "leave_roster"
    current_block_tab = block_tab if block_tab in {"HANH_CHINH", "CHUYEN_MON"} else (scope_info.get("block_code") or "CHUYEN_MON")

    role_codes = set(scope_info.get("role_codes") or [])

    # Tab phê duyệt tổng hợp cấp Bệnh viện là khu vực tổng hợp của Khối chuyên môn,
    # nhưng người được quyền xem/phê duyệt có thể thuộc Phòng Kế hoạch - Tổng hợp
    # trong Khối hành chính. Vì vậy không được khóa quyền xem theo current_block_tab.
    can_approve_hospital_position_summary = bool(scope_info.get("is_khth_manager"))
    can_view_hospital_position_summary = _can_view_hospital_position_summary(db, user, role_codes)
    allow_khth_summary = can_view_hospital_position_summary

    if sub_tab == "khth_summary" and can_view_hospital_position_summary:
        current_block_tab = "CHUYEN_MON"

    target_week_was_selected = bool(str(target_week or "").strip())
    raw_target_week = str(target_week or "").strip()

    allowed_position_tabs = {"catalog", "assign", "notice", "approve"}
    if can_view_hospital_position_summary:
        allowed_position_tabs.add("khth_summary")

    if current_main_tab == "positions":
        current_sub_tab = sub_tab if sub_tab in allowed_position_tabs else "catalog"
    else:
        current_sub_tab = sub_tab if sub_tab in {"handoff"} else "handoff"

    full_block_scope = can_view_hospital_position_summary and current_sub_tab == "khth_summary"

    if target_week_was_selected:
        effective_target_week = raw_target_week
        target_week_monday = _parse_week_value(effective_target_week).strftime("%Y-%m-%d")
    else:
        latest_target_date = _find_latest_position_week(
            db,
            scope_info,
            current_block_tab,
            full_block_scope=full_block_scope,
            hospital_summary_scope=(current_sub_tab == "khth_summary" and can_view_hospital_position_summary),
        )
        if latest_target_date:
            effective_target_week = _week_value_from_date_str(latest_target_date)
            target_week_monday = latest_target_date
        else:
            effective_target_week = _current_week_value()
            target_week_monday = _parse_week_value(effective_target_week).strftime("%Y-%m-%d")

    target_date = target_week_monday

    all_catalog_rows = _load_catalog_rows(db, scope_info, current_block_tab, full_block_scope=full_block_scope)

    effective_selected_position_type = str(selected_position_type or "").strip().upper()
    assign_position_types = sorted(
        {
            str(item.get("position_type") or "").strip().upper()
            for item in all_catalog_rows
            if str(item.get("position_type") or "").strip()
        }
    )

    if current_sub_tab == "assign" and effective_selected_position_type:
        catalog_rows = [
            item for item in all_catalog_rows
            if str(item.get("position_type") or "").strip().upper() == effective_selected_position_type
        ]
    else:
        catalog_rows = all_catalog_rows

    plan_history_year = str(request.query_params.get("plan_history_year") or _now_vn().year)
    plan_history_month = str(request.query_params.get("plan_history_month") or _now_vn().month)

    plan_rows = _load_position_plan_rows(
        db,
        scope_info,
        current_block_tab,
        full_block_scope=full_block_scope,
        target_week_monday=target_week_monday,
    )
    plan_history_rows = _load_position_plan_rows(
        db,
        scope_info,
        current_block_tab,
        full_block_scope=full_block_scope,
        target_week_monday="",
        history_year=plan_history_year,
        history_month=plan_history_month,
    )
    khth_plan_history_rows = _load_khth_plan_history_rows(
        db,
        history_year=plan_history_year,
        history_month=plan_history_month,
    ) if can_view_hospital_position_summary else []
    
    current_week_plan = _load_current_week_plan(
        db,
        scope_info,
        current_block_tab,
        target_week_monday=target_week_monday,
        selected_position_type=effective_selected_position_type,
    ) if current_main_tab == "positions" and current_sub_tab == "assign" else None
    
    week_assignment_map = _load_week_assignment_map(db, current_week_plan.get("id") if current_week_plan else None)
    week_days = _build_week_days(target_week_monday)
    assignable_users = _load_scope_assignable_users(db, scope_info, current_block_tab)
    approved_leave_user_ids_by_date = _load_approved_leave_user_ids_by_date(
        db,
        scope_info,
        current_block_tab,
        week_days,
    )
    approved_roster_plus_one_user_ids_by_date = _load_approved_roster_plus_one_user_ids_by_date(
        db,
        scope_info,
        current_block_tab,
        week_days,
    )
    blocked_user_ids_by_date = _merge_blocked_user_ids_by_date(
        approved_leave_user_ids_by_date,
        approved_roster_plus_one_user_ids_by_date,
    )
    leave_contract = _load_contract_from_leave_schedule(
        db,
        scope_info,
        target_date or None,
        full_block_scope=current_main_tab == "leave_roster",
        block_code=current_block_tab if current_main_tab == "leave_roster" else "",
    )
    unread_notice_count = _load_unread_notice_count(db, user.id)
    unread_khth_pending_count = _load_unread_notice_count_by_types(
        db,
        user.id,
        [
            "KHTH_PLAN_PENDING",
            "KHTH_ADJUSTMENT_PENDING",
        ],
    )

    notice_plan_week = str(request.query_params.get("notice_plan_week") or effective_target_week or _current_week_value())
    notice_adjustment_week = str(request.query_params.get("notice_adjustment_week") or effective_target_week or _current_week_value())

    try:
        notice_plan_week_monday = _parse_week_value(notice_plan_week).strftime("%Y-%m-%d")
    except Exception:
        notice_plan_week = _current_week_value()
        notice_plan_week_monday = _parse_week_value(notice_plan_week).strftime("%Y-%m-%d")

    try:
        notice_adjustment_week_monday = _parse_week_value(notice_adjustment_week).strftime("%Y-%m-%d")
    except Exception:
        notice_adjustment_week = _current_week_value()
        notice_adjustment_week_monday = _parse_week_value(notice_adjustment_week).strftime("%Y-%m-%d")

    plan_notice_rows = _load_notice_rows(
        db,
        user.id,
        notice_type="PLAN_APPROVED",
        week_monday=notice_plan_week_monday,
    )
    adjustment_notice_rows = _load_notice_rows(
        db,
        user.id,
        notice_type="ADJUSTMENT_APPROVED",
        week_monday=notice_adjustment_week_monday,
    )
    notice_rows = plan_notice_rows + adjustment_notice_rows

    approve_summary = _load_unit_approve_summary(db, scope_info, current_block_tab, target_week_monday)
    khth_approve_summary = _load_khth_approve_summary(db, target_week_monday) if can_view_hospital_position_summary else None
    khth_summary = _load_khth_cm_summary(db, target_date or "") if can_view_hospital_position_summary else None

    adjustment_history_year = str(request.query_params.get("adjustment_year") or _now_vn().year)
    adjustment_history_month = str(request.query_params.get("adjustment_month") or _now_vn().month)

    adjustment_target_rows = _load_clinic_adjustment_targets(
        db,
        scope_info,
        target_week_monday,
    ) if current_main_tab == "positions" and current_sub_tab == "approve" else []

    unit_adjustment_rows = _load_adjustment_rows(
        db,
        scope_info=scope_info,
        target_week_monday="",
        hospital_scope=False,
        history_year=adjustment_history_year,
        history_month=adjustment_history_month,
    ) if current_main_tab == "positions" and current_sub_tab == "approve" else []

    khth_pending_adjustment_rows = _load_adjustment_rows(
        db,
        hospital_scope=True,
        status_filter="UNIT_APPROVED",
    ) if can_view_hospital_position_summary else []

    khth_adjustment_history_rows = _load_adjustment_rows(
        db,
        hospital_scope=True,
        history_year=adjustment_history_year,
        history_month=adjustment_history_month,
    ) if can_view_hospital_position_summary else []

    return templates.TemplateResponse(
        "work_positions.html",
        {
            "request": request,
            "app_name": settings.APP_NAME,
            "company_name": settings.COMPANY_NAME,
            "main_tab": current_main_tab,
            "block_tab": current_block_tab,
            "sub_tab": current_sub_tab,
            "scope_info": scope_info,
            "scope_options": scope_info.get("scope_options", []),
            "selected_scope_unit_id": scope_info.get("selected_scope_unit_id", ""),
            "scope_display": scope_info["scope_display"],
            "block_label": scope_info["block_label"],
            "can_manage_catalog": _can_manage_catalog(db, user),
            "can_assign_positions": scope_info["can_assign_positions"],
            "can_view_khth_summary": can_view_hospital_position_summary,
            "can_view_hospital_position_summary": can_view_hospital_position_summary,
            "can_approve_hospital_position_summary": can_approve_hospital_position_summary,
            "catalog_rows": catalog_rows,
            "plan_rows": plan_rows,
            "plan_history_rows": plan_history_rows,
            "khth_plan_history_rows": khth_plan_history_rows,
            "plan_history_year": plan_history_year,
            "plan_history_month": plan_history_month,
            "target_week_was_selected": target_week_was_selected,
            "leave_contract": leave_contract,
            "khth_summary": khth_summary,
            "approve_summary": approve_summary,
            "khth_approve_summary": khth_approve_summary,
            "adjustment_target_rows": adjustment_target_rows,
            "unit_adjustment_rows": unit_adjustment_rows,
            "khth_pending_adjustment_rows": khth_pending_adjustment_rows,
            "khth_adjustment_history_rows": khth_adjustment_history_rows,
            "adjustment_history_year": adjustment_history_year,
            "adjustment_history_month": adjustment_history_month,
            "notice_rows": notice_rows,
            "plan_notice_rows": plan_notice_rows,
            "adjustment_notice_rows": adjustment_notice_rows,
            "notice_plan_week": notice_plan_week,
            "notice_adjustment_week": notice_adjustment_week,
            "notice_plan_week_label": _week_label_from_date_str(notice_plan_week_monday),
            "notice_adjustment_week_label": _week_label_from_date_str(notice_adjustment_week_monday),
            "unread_notice_count": unread_notice_count,
            "unread_khth_pending_count": unread_khth_pending_count,
            "target_week": effective_target_week,
            "selected_position_type": effective_selected_position_type,
            "assign_position_types": assign_position_types,
            "target_week_label": _week_label_from_date_str(target_week_monday) if target_week_monday else "",
            "week_days": week_days,
            "current_week_plan": current_week_plan,
            "week_assignment_map": week_assignment_map,
            "assignable_users": assignable_users,
            "approved_leave_user_ids_by_date": approved_leave_user_ids_by_date,
            "approved_roster_plus_one_user_ids_by_date": approved_roster_plus_one_user_ids_by_date,
            "blocked_user_ids_by_date": blocked_user_ids_by_date,
            "message": request.query_params.get("msg", ""),
            "error": request.query_params.get("error", ""),
        },
    )


@router.post("/catalog/add")
def add_work_position_catalog(
    request: Request,
    position_type: str = Form(...),
    position_name: str = Form(...),
    sort_order: int = Form(0),
    scope_unit_id: str = Form(""),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    if not _can_manage_catalog(db, user):
        raise HTTPException(status_code=403, detail="Bạn không có quyền tạo danh mục vị trí làm việc.")

    scope_info = _scope_payload(db, user, scope_unit_id)
    scope_unit = scope_info.get("scope_unit")
    scope_subunit = scope_info.get("scope_subunit")
    apply_unit = scope_subunit or scope_unit

    if apply_unit is None:
        return RedirectResponse(
            url="/work-positions?error=Chưa xác định được phạm vi đơn vị để tạo vị trí làm việc.",
            status_code=302,
        )

    now_str = _dt_to_str(_now_vn())
    db.execute(
        text(
            """
            INSERT INTO work_position_catalog (
                id, owner_unit_id, owner_unit_name, apply_unit_id, apply_unit_name,
                block_code, position_type, position_name, sort_order, is_active,
                created_by_id, created_by_name, created_at, updated_at
            ) VALUES (
                :id, :owner_unit_id, :owner_unit_name, :apply_unit_id, :apply_unit_name,
                :block_code, :position_type, :position_name, :sort_order, 1,
                :created_by_id, :created_by_name, :created_at, :updated_at
            )
            """
        ),
        {
            "id": str(uuid.uuid4()),
            "owner_unit_id": getattr(scope_unit, "id", None) or getattr(apply_unit, "id", None),
            "owner_unit_name": getattr(scope_unit, "ten_don_vi", None) or getattr(apply_unit, "ten_don_vi", None),
            "apply_unit_id": getattr(apply_unit, "id", None),
            "apply_unit_name": getattr(apply_unit, "ten_don_vi", None),
            "block_code": scope_info.get("block_code") or _unit_block_value(apply_unit),
            "position_type": (position_type or "").strip(),
            "position_name": (position_name or "").strip(),
            "sort_order": int(sort_order or 0),
            "created_by_id": user.id,
            "created_by_name": user.full_name or user.username,
            "created_at": now_str,
            "updated_at": now_str,
        },
    )
    db.commit()

    _emit_work_position_realtime_sync(
        _work_position_scope_realtime_user_ids(
            db,
            scope_unit_id=getattr(scope_unit, "id", None),
            scope_subunit_id=getattr(scope_subunit, "id", None),
        ),
        _work_position_all_changed_payload(
            "work_position_catalog_changed",
            actor_user_id=str(user.id),
            block_code=str(scope_info.get("block_code") or ""),
            sub_tab="catalog",
        ),
    )

    return RedirectResponse(
        url=(
            f"/work-positions?main_tab=positions"
            f"&block_tab={scope_info.get('block_code') or 'CHUYEN_MON'}"
            f"&sub_tab=catalog"
            f"&scope_unit_id={scope_info.get('selected_scope_unit_id') or ''}"
            f"&msg=Đã tạo danh mục vị trí làm việc."
        ),
        status_code=302,
    )

@router.post("/plans/save-weekly")
def save_weekly_position_plan(
    request: Request,
    target_week: str = Form(...),
    selected_position_type: str = Form(""),
    scope_unit_id: str = Form(""),
    entry_catalog_ids: list[str] = Form([]),
    entry_work_dates: list[str] = Form([]),
    entry_session_codes: list[str] = Form([]),
    entry_staff_groups: list[str] = Form([]),
    assigned_user_ids: list[str] = Form([]),
    assigned_user_names: list[str] = Form([]),
    entry_notes: list[str] = Form([]),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    scope_info = _scope_payload(db, user, scope_unit_id)
    if not scope_info.get("can_assign_positions"):
        raise HTTPException(status_code=403, detail="Bạn không có quyền phân vị trí.")

    week_start = _parse_week_value(target_week)
    target_date = week_start.strftime("%Y-%m-%d")
    now_str = _dt_to_str(_now_vn())
    position_type = str(selected_position_type or "").strip().upper()
    shift_code = f"WEEKLY_MATRIX_{position_type}" if position_type else "WEEKLY_MATRIX"

    scope_unit = scope_info.get("scope_unit")
    scope_subunit = scope_info.get("scope_subunit")

    row = db.execute(
        text(
            """
            SELECT *
            FROM work_position_plans
            WHERE COALESCE(block_code, '') = :block_code
              AND COALESCE(scope_unit_id, '') = :scope_unit_id
              AND COALESCE(scope_subunit_id, '') = :scope_subunit_id
              AND target_date = :target_date
              AND COALESCE(shift_code, '') = :shift_code
            ORDER BY updated_at DESC
            LIMIT 1
            """
        ),
        {
            "block_code": scope_info.get("block_code") or "",
            "scope_unit_id": getattr(scope_unit, "id", None) or "",
            "scope_subunit_id": getattr(scope_subunit, "id", None) or "",
            "target_date": target_date,
            "shift_code": shift_code,
        },
    ).fetchone()

    if row:
        plan_id = str(_row_to_dict(row).get("id"))
        db.execute(text("DELETE FROM work_position_plan_items WHERE plan_id = :plan_id"), {"plan_id": plan_id})
        db.execute(
            text(
                """
                UPDATE work_position_plans
                SET status = 'DRAFT',
                    submitted_by_id = :submitted_by_id,
                    submitted_by_name = :submitted_by_name,
                    updated_at = :updated_at,
                    shift_code = :shift_code
                WHERE id = :id
                """
            ),
            {
                "id": plan_id,
                "submitted_by_id": user.id,
                "submitted_by_name": user.full_name or user.username,
                "updated_at": now_str,
                "shift_code": shift_code,
            },
        )
    else:
        plan_id = str(uuid.uuid4())
        db.execute(
            text(
                """
                INSERT INTO work_position_plans (
                    id, block_code, scope_unit_id, scope_unit_name, scope_subunit_id, scope_subunit_name,
                    target_date, shift_code, status,
                    submitted_by_id, submitted_by_name, submitted_at,
                    created_at, updated_at
                ) VALUES (
                    :id, :block_code, :scope_unit_id, :scope_unit_name, :scope_subunit_id, :scope_subunit_name,
                    :target_date, :shift_code, 'DRAFT',
                    :submitted_by_id, :submitted_by_name, :submitted_at,
                    :created_at, :updated_at
                )
                """
            ),
            {
                "id": plan_id,
                "block_code": scope_info.get("block_code") or "",
                "scope_unit_id": getattr(scope_unit, "id", None),
                "scope_unit_name": getattr(scope_unit, "ten_don_vi", None),
                "scope_subunit_id": getattr(scope_subunit, "id", None),
                "scope_subunit_name": getattr(scope_subunit, "ten_don_vi", None),
                "target_date": target_date,
                "shift_code": shift_code,
                "submitted_by_id": user.id,
                "submitted_by_name": user.full_name or user.username,
                "submitted_at": now_str,
                "created_at": now_str,
                "updated_at": now_str,
            },
        )

    row_count = min(
        len(entry_catalog_ids),
        len(entry_work_dates),
        len(entry_session_codes),
        len(entry_staff_groups),
        len(assigned_user_ids),
        len(assigned_user_names),
        len(entry_notes),
    )
    saved_count = 0
    week_days_for_save = _build_week_days(target_date)
    valid_dates = {item["date"] for item in week_days_for_save}
    valid_sessions = {"SANG", "CHIEU", "ALLDAY"}
    valid_groups = {"BS", "DD", "GENERAL", "GENERAL_MULTI", "LAB"}
    assignable_users_for_save = _load_scope_assignable_users(
        db,
        scope_info,
        scope_info.get("block_code") or "",
    )

    leave_blocked_user_ids_by_date = _load_approved_leave_user_ids_by_date(
        db,
        scope_info,
        scope_info.get("block_code") or "",
        week_days_for_save,
    )
    roster_plus_one_user_ids_by_date = _load_approved_roster_plus_one_user_ids_by_date(
        db,
        scope_info,
        scope_info.get("block_code") or "",
        week_days_for_save,
    )
    blocked_user_ids_by_date_for_save = _merge_blocked_user_ids_by_date(
        leave_blocked_user_ids_by_date,
        roster_plus_one_user_ids_by_date,
    )

    seen_same_type_position_keys: set[tuple[str, str, str, str]] = set()
    seen_cross_type_keys: dict[tuple[str, str, str], str] = {}

    for idx in range(row_count):
        catalog_id = str(entry_catalog_ids[idx] or "").strip()
        work_date = str(entry_work_dates[idx] or "").strip()
        session_code = str(entry_session_codes[idx] or "").strip().upper()
        staff_group = str(entry_staff_groups[idx] or "").strip().upper()
        assigned_ids_val = str(assigned_user_ids[idx] or "").strip()
        assigned_name = str(assigned_user_names[idx] or "").strip()
        note_val = str(entry_notes[idx] or "").strip()

        if not catalog_id or not work_date or session_code not in valid_sessions or staff_group not in valid_groups:
            continue
        if work_date not in valid_dates:
            continue
        if not assigned_name and not note_val:
            continue

        catalog_row = db.execute(
            text("SELECT * FROM work_position_catalog WHERE id = :id"),
            {"id": catalog_id},
        ).fetchone()
        if not catalog_row:
            raise HTTPException(status_code=400, detail="Danh mục vị trí không tồn tại hoặc đã bị chỉnh sửa.")
        catalog_item = _row_to_dict(catalog_row)
        if not _catalog_item_allowed_for_scope(
            catalog_item,
            scope_info,
            scope_info.get("block_code") or "",
            position_type,
        ):
            raise HTTPException(
                status_code=403,
                detail="Danh mục vị trí không thuộc phạm vi được phép phân công.",
            )

        current_position_type = str(catalog_item.get("position_type") or "").strip().upper() or "UNKNOWN"

        parsed_user_ids = [x.strip() for x in assigned_ids_val.split(",") if x.strip()]
        if parsed_user_ids:
            allowed_user_map = _assignable_user_map_for_staff_group(assignable_users_for_save, staff_group)
            invalid_assigned_ids = [uid for uid in parsed_user_ids if uid not in allowed_user_map]
            if invalid_assigned_ids:
                raise HTTPException(
                    status_code=403,
                    detail="Có nhân sự không thuộc phạm vi hoặc không đúng nhóm vai trò được phép phân vị trí.",
                )
            assigned_name = ", ".join(
                str(allowed_user_map[uid].get("name") or "").strip()
                for uid in parsed_user_ids
                if uid in allowed_user_map
            )

            blocked_ids_for_date = set(blocked_user_ids_by_date_for_save.get(work_date, []))
            blocked_assigned_ids = [uid for uid in parsed_user_ids if uid in blocked_ids_for_date]
            if blocked_assigned_ids:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"Ngày {work_date} có user đang bị chặn phân vị trí do đã đăng ký/được duyệt nghỉ "
                        f"hoặc do ngày này là ngày liền sau lịch trực đã được duyệt. "
                        f"Không chặn chính ngày trực; chỉ chặn ngày trực + 1."
                    ),
                )

            effective_sessions = ["SANG", "CHIEU"] if session_code == "ALLDAY" else [session_code]
            for uid in parsed_user_ids:
                for sess in effective_sessions:
                    cross_type_key = (work_date, sess, uid)
                    previous_type = seen_cross_type_keys.get(cross_type_key)

                    if previous_type and previous_type != current_position_type:
                        raise HTTPException(
                            status_code=400,
                            detail=(
                                f"User đã được phân ở loại vị trí {previous_type} trong cùng buổi {sess} ngày {work_date}. "
                                f"Không được phân đồng thời sang loại vị trí {current_position_type}."
                            ),
                        )

                    seen_cross_type_keys[cross_type_key] = current_position_type

                    if current_position_type != "PHONG_DIEU_TRI":
                        same_type_key = (work_date, sess, uid, current_position_type)
                        if same_type_key in seen_same_type_position_keys:
                            raise HTTPException(
                                status_code=400,
                                detail=(
                                    f"User đã được phân từ 2 tên vị trí trở lên trong cùng loại vị trí {current_position_type}, "
                                    f"cùng buổi {sess} ngày {work_date}."
                                ),
                            )

                        seen_same_type_position_keys.add(same_type_key)

        db.execute(
            text(
                """
                INSERT INTO work_position_plan_items (
                    id, plan_id, catalog_id, position_name, assigned_user_id, assigned_user_name,
                    source_contract_type, source_contract_ref, notes, created_at, updated_at,
                    work_date, session_code, staff_group
                ) VALUES (
                    :id, :plan_id, :catalog_id, :position_name, :assigned_user_id, :assigned_user_name,
                    :source_contract_type, :source_contract_ref, :notes, :created_at, :updated_at,
                    :work_date, :session_code, :staff_group
                )
                """
            ),
            {
                "id": str(uuid.uuid4()),
                "plan_id": plan_id,
                "catalog_id": catalog_id,
                "position_name": catalog_item.get("position_name") or "",
                "assigned_user_id": assigned_ids_val,
                "assigned_user_name": assigned_name,
                "source_contract_type": "LEAVE_SCHEDULE",
                "source_contract_ref": target_week,
                "notes": note_val,
                "created_at": now_str,
                "updated_at": now_str,
                "work_date": work_date,
                "session_code": session_code,
                "staff_group": staff_group,
            },
        )
        saved_count += 1

    db.commit()

    _emit_work_position_realtime_sync(
        _work_position_scope_realtime_user_ids(
            db,
            scope_unit_id=getattr(scope_unit, "id", None),
            scope_subunit_id=getattr(scope_subunit, "id", None),
        ),
        _work_position_all_changed_payload(
            "work_position_plan_saved",
            actor_user_id=str(user.id),
            block_code=str(scope_info.get("block_code") or ""),
            target_week=str(target_week or ""),
            target_date=str(target_date or ""),
            plan_id=str(plan_id or ""),
            selected_position_type=str(selected_position_type or ""),
        ),
    )

    _emit_work_position_realtime_sync(
        _unit_member_user_ids(
            db,
            [
                x for x in [
                    getattr(scope_unit, "id", None),
                    getattr(scope_subunit, "id", None),
                ] if x
            ],
        ),
        _position_notify_payload(
            "weekly_plan_saved",
            block_code=scope_info.get("block_code") or "",
            target_week=target_week,
            target_date=target_date,
            plan_id=plan_id,
        ),
    )

    return RedirectResponse(
        url=(
            f"/work-positions?main_tab=positions"
            f"&block_tab={scope_info.get('block_code') or 'CHUYEN_MON'}"
            f"&sub_tab=approve"
            f"&target_week={target_week}"
            f"&scope_unit_id={scope_info.get('selected_scope_unit_id') or ''}"
            f"&msg=Đã lưu phân vị trí theo tuần ({saved_count} ô thao tác). Vui lòng kiểm tra Bảng tổng hợp phân vị trí và phê duyệt cấp đơn vị."
        ),
        status_code=302,
    )

@router.post("/plans/approve-unit")
def approve_unit_position_plan(
    request: Request,
    plan_id: str = Form(...),
    scope_unit_id: str = Form(""),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)
    scope_info = _scope_payload(db, user, scope_unit_id)

    if not scope_info.get("can_assign_positions"):
        raise HTTPException(status_code=403, detail="Bạn không có quyền duyệt cấp đơn vị.")

    plan_row = db.execute(
        text("SELECT * FROM work_position_plans WHERE id = :id"),
        {"id": plan_id},
    ).fetchone()
    if not plan_row:
        raise HTTPException(status_code=404, detail="Không tìm thấy bản phân vị trí.")

    plan = _row_to_dict(plan_row)
    plan_block_code = str(plan.get("block_code") or "").strip().upper()
    scope_block_code = str(scope_info.get("block_code") or "").strip().upper()
    plan_scope_unit_id = str(plan.get("scope_unit_id") or "").strip()
    plan_scope_subunit_id = str(plan.get("scope_subunit_id") or "").strip()
    scope_unit_id_value = str(scope_info.get("scope_unit_id") or "").strip()
    scope_subunit_id_value = str(scope_info.get("scope_subunit_id") or "").strip()

    if plan_block_code != scope_block_code:
        raise HTTPException(status_code=403, detail="Bản phân vị trí không thuộc khối/phạm vi được phép duyệt.")

    if plan_scope_unit_id != scope_unit_id_value or plan_scope_subunit_id != scope_subunit_id_value:
        raise HTTPException(status_code=403, detail="Bản phân vị trí không thuộc đơn vị được phép duyệt.")

    current_status = str(plan.get("status") or "").strip().upper()
    if current_status != "DRAFT":
        raise HTTPException(status_code=400, detail="Chỉ được duyệt cấp đơn vị đối với bản phân vị trí đang ở trạng thái chờ duyệt.")

    plan_items = _load_plan_item_rows(db, plan_id)
    now_str = _dt_to_str(_now_vn())
    approver_name = user.full_name or user.username or ""

    needs_khth = _plan_needs_khth_approval(str(plan.get("block_code") or ""), plan_items)
    next_status = "UNIT_APPROVED" if needs_khth else "APPROVED"

    db.execute(
        text(
            """
            UPDATE work_position_plans
            SET status = :status,
                unit_approved_by_id = :unit_approved_by_id,
                unit_approved_by_name = :unit_approved_by_name,
                unit_approved_at = :unit_approved_at,
                approved_by_id = CASE WHEN :status = 'APPROVED' THEN :approved_by_id ELSE approved_by_id END,
                approved_by_name = CASE WHEN :status = 'APPROVED' THEN :approved_by_name ELSE approved_by_name END,
                approved_at = CASE WHEN :status = 'APPROVED' THEN :approved_at ELSE approved_at END,
                updated_at = :updated_at
            WHERE id = :id
            """
        ),
        {
            "id": plan_id,
            "status": next_status,
            "unit_approved_by_id": user.id,
            "unit_approved_by_name": approver_name,
            "unit_approved_at": now_str,
            "approved_by_id": user.id,
            "approved_by_name": approver_name,
            "approved_at": now_str,
            "updated_at": now_str,
        },
    )

    notice_user_ids: list[str] = []
    khth_notice_user_ids: list[str] = []

    if next_status == "APPROVED":
        notice_user_ids = _create_position_notices_for_plan(
            db,
            plan_id=plan_id,
            approved_by_name=approver_name,
        )
    elif next_status == "UNIT_APPROVED":
        scope_label = str(plan.get("scope_subunit_name") or plan.get("scope_unit_name") or "-")
        week_label = _week_label_from_date_str(str(plan.get("target_date") or ""))
        position_types = ", ".join(
            sorted({
                str(item.get("position_type") or "").strip().upper()
                for item in plan_items
                if str(item.get("position_type") or "").strip()
            })
        ) or "-"

        khth_notice_user_ids = _create_khth_pending_notice(
            db,
            notice_type="KHTH_PLAN_PENDING",
            title="Có bản phân vị trí chờ phê duyệt cấp Bệnh viện",
            message=(
                f"{week_label} | Đơn vị: {scope_label} | "
                f"Loại vị trí: {position_types} | "
                f"Người duyệt cấp đơn vị: {approver_name}"
            ),
            target_date=str(plan.get("target_date") or ""),
            related_plan_id=plan_id,
        )

    db.commit()

    _emit_work_position_realtime_sync(
        _work_position_scope_realtime_user_ids(
            db,
            scope_unit_id=str(plan.get("scope_unit_id") or ""),
            scope_subunit_id=str(plan.get("scope_subunit_id") or ""),
            extra_user_ids=khth_notice_user_ids,
        ),
        _work_position_all_changed_payload(
            "work_position_plan_unit_approved",
            actor_user_id=str(user.id),
            plan_id=str(plan_id),
            status=str(next_status or ""),
            target_date=str(plan.get("target_date") or ""),
            target_week=_week_value_from_date_str(str(plan.get("target_date") or "")),
        ),
    )

    if notice_user_ids:
        _emit_work_position_realtime_sync(
            notice_user_ids,
            _position_notify_payload(
                "work_position_notice_created",
                plan_id=plan_id,
                target_date=str(plan.get("target_date") or ""),
            ),
        )

    if khth_notice_user_ids:
        _emit_work_position_realtime_sync(
            khth_notice_user_ids,
            _position_notify_payload(
                "work_position_notice_created",
                notice_type="KHTH_PLAN_PENDING",
                plan_id=plan_id,
                target_date=str(plan.get("target_date") or ""),
                target_week=_week_value_from_date_str(str(plan.get("target_date") or "")),
            ),
        )

    return RedirectResponse(
        url=(
            f"/work-positions?main_tab=positions"
            f"&block_tab={scope_info.get('block_code') or 'CHUYEN_MON'}"
            f"&sub_tab=approve"
            f"&target_week={_week_value_from_date_str(str(plan.get('target_date') or ''))}"
            f"&scope_unit_id={scope_info.get('selected_scope_unit_id') or ''}"
            f"&msg=Đã duyệt cấp đơn vị."
        ),
        status_code=302,
    )
    
@router.post("/adjustments/create")
def create_work_position_adjustment(
    request: Request,
    original_plan_id: str = Form(...),
    catalog_id: str = Form(...),
    work_date: str = Form(...),
    session_code: str = Form(...),
    position_name: str = Form(...),
    adjust_type: str = Form(...),
    bs_item_id: str = Form(""),
    dd_item_id: str = Form(""),
    new_bs_user_id: str = Form(""),
    new_dd_user_id: str = Form(""),
    reason: str = Form(...),
    scope_unit_id: str = Form(""),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)
    scope_info = _scope_payload(db, user, scope_unit_id)

    adjust_type_val = str(adjust_type or "").strip().upper()
    if adjust_type_val not in {"CLOSE_ROOM", "CHANGE_BS", "CHANGE_DD", "CHANGE_BOTH"}:
        raise HTTPException(status_code=400, detail="Loại điều chỉnh không hợp lệ.")

    plan_row = db.execute(
        text("SELECT * FROM work_position_plans WHERE id = :id"),
        {"id": original_plan_id},
    ).fetchone()
    if not plan_row:
        raise HTTPException(status_code=404, detail="Không tìm thấy bản phân vị trí gốc.")

    plan = _row_to_dict(plan_row)
    if str(plan.get("status") or "").upper() != "KHTH_APPROVED":
        raise HTTPException(status_code=400, detail="Chỉ được điều chỉnh bản đã phê duyệt cấp Bệnh viện.")

    scope_unit_id = str(scope_info.get("scope_unit_id") or "")
    scope_subunit_id = str(scope_info.get("scope_subunit_id") or "")
    plan_scope_unit_id = str(plan.get("scope_unit_id") or "")
    plan_scope_subunit_id = str(plan.get("scope_subunit_id") or "")

    if scope_unit_id not in {plan_scope_unit_id, plan_scope_subunit_id} and scope_subunit_id not in {plan_scope_unit_id, plan_scope_subunit_id}:
        raise HTTPException(status_code=403, detail="Bạn không có quyền lập điều chỉnh cho bản phân vị trí này.")

    now_str = _dt_to_str(_now_vn())
    adjustment_id = str(uuid.uuid4())
    requester_name = user.full_name or user.username or ""

    def _user_name(user_id: str) -> str:
        uid = str(user_id or "").strip()
        if not uid:
            return ""
        row_user = db.get(Users, uid)
        if not row_user:
            return ""
        return str(row_user.full_name or row_user.username or "")

    def _old_item(item_id: str) -> dict[str, Any]:
        iid = str(item_id or "").strip()
        if not iid:
            return {}
        row = db.execute(
            text("SELECT * FROM work_position_plan_items WHERE id = :id"),
            {"id": iid},
        ).fetchone()
        return _row_to_dict(row) if row else {}

    bs_old = _old_item(bs_item_id)
    dd_old = _old_item(dd_item_id)

    db.execute(
        text(
            """
            INSERT INTO work_position_adjustments (
                id, original_plan_id, block_code, scope_unit_id, scope_unit_name,
                scope_subunit_id, scope_subunit_name, target_date, status,
                adjust_type, reason, requested_by_id, requested_by_name, requested_at,
                created_at, updated_at
            ) VALUES (
                :id, :original_plan_id, :block_code, :scope_unit_id, :scope_unit_name,
                :scope_subunit_id, :scope_subunit_name, :target_date, 'UNIT_APPROVED',
                :adjust_type, :reason, :requested_by_id, :requested_by_name, :requested_at,
                :created_at, :updated_at
            )
            """
        ),
        {
            "id": adjustment_id,
            "original_plan_id": original_plan_id,
            "block_code": plan.get("block_code") or "CHUYEN_MON",
            "scope_unit_id": plan.get("scope_unit_id"),
            "scope_unit_name": plan.get("scope_unit_name"),
            "scope_subunit_id": plan.get("scope_subunit_id"),
            "scope_subunit_name": plan.get("scope_subunit_name"),
            "target_date": plan.get("target_date"),
            "adjust_type": adjust_type_val,
            "reason": str(reason or "").strip(),
            "requested_by_id": user.id,
            "requested_by_name": requester_name,
            "requested_at": now_str,
            "created_at": now_str,
            "updated_at": now_str,
        },
    )

    adjustment_items: list[dict[str, Any]] = []

    if adjust_type_val in {"CLOSE_ROOM", "CHANGE_BS", "CHANGE_BOTH"}:
        adjustment_items.append({
            "original_item_id": bs_item_id,
            "staff_group": "BS",
            "old_assigned_user_id": bs_old.get("assigned_user_id") or "",
            "old_assigned_user_name": bs_old.get("assigned_user_name") or "",
            "new_assigned_user_id": "" if adjust_type_val == "CLOSE_ROOM" else new_bs_user_id,
            "new_assigned_user_name": "" if adjust_type_val == "CLOSE_ROOM" else _user_name(new_bs_user_id),
        })

    if adjust_type_val in {"CLOSE_ROOM", "CHANGE_DD", "CHANGE_BOTH"}:
        adjustment_items.append({
            "original_item_id": dd_item_id,
            "staff_group": "DD",
            "old_assigned_user_id": dd_old.get("assigned_user_id") or "",
            "old_assigned_user_name": dd_old.get("assigned_user_name") or "",
            "new_assigned_user_id": "" if adjust_type_val == "CLOSE_ROOM" else new_dd_user_id,
            "new_assigned_user_name": "" if adjust_type_val == "CLOSE_ROOM" else _user_name(new_dd_user_id),
        })

    for row in adjustment_items:
        db.execute(
            text(
                """
                INSERT INTO work_position_adjustment_items (
                    id, adjustment_id, original_plan_id, original_item_id,
                    catalog_id, position_name, work_date, session_code, staff_group,
                    adjust_type, old_assigned_user_id, old_assigned_user_name,
                    new_assigned_user_id, new_assigned_user_name, notes,
                    created_at, updated_at
                ) VALUES (
                    :id, :adjustment_id, :original_plan_id, :original_item_id,
                    :catalog_id, :position_name, :work_date, :session_code, :staff_group,
                    :adjust_type, :old_assigned_user_id, :old_assigned_user_name,
                    :new_assigned_user_id, :new_assigned_user_name, :notes,
                    :created_at, :updated_at
                )
                """
            ),
            {
                "id": str(uuid.uuid4()),
                "adjustment_id": adjustment_id,
                "original_plan_id": original_plan_id,
                "original_item_id": row.get("original_item_id") or "",
                "catalog_id": catalog_id,
                "position_name": position_name,
                "work_date": work_date,
                "session_code": session_code,
                "staff_group": row.get("staff_group") or "",
                "adjust_type": adjust_type_val,
                "old_assigned_user_id": row.get("old_assigned_user_id") or "",
                "old_assigned_user_name": row.get("old_assigned_user_name") or "",
                "new_assigned_user_id": row.get("new_assigned_user_id") or "",
                "new_assigned_user_name": row.get("new_assigned_user_name") or "",
                "notes": str(reason or "").strip(),
                "created_at": now_str,
                "updated_at": now_str,
            },
        )

    scope_label = str(plan.get("scope_subunit_name") or plan.get("scope_unit_name") or "-")
    week_label = _week_label_from_date_str(str(plan.get("target_date") or ""))
    session_label = "Sáng" if str(session_code or "").upper() == "SANG" else ("Chiều" if str(session_code or "").upper() == "CHIEU" else str(session_code or "-"))

    khth_notice_user_ids = _create_khth_pending_notice(
        db,
        notice_type="KHTH_ADJUSTMENT_PENDING",
        title="Có phiếu điều chỉnh phân vị trí chờ phê duyệt cấp Bệnh viện",
        message=(
            f"{week_label} | Ngày: {work_date} | Buổi: {session_label} | "
            f"Đơn vị: {scope_label} | Vị trí: {position_name} | "
            f"Loại điều chỉnh: {_adjustment_type_label(adjust_type_val)} | "
            f"Lý do: {str(reason or '').strip()}"
        ),
        target_date=str(plan.get("target_date") or work_date or ""),
        related_plan_id=original_plan_id,
    )

    db.commit()

    _emit_work_position_realtime_sync(
        _work_position_scope_realtime_user_ids(
            db,
            scope_unit_id=str(plan.get("scope_unit_id") or ""),
            scope_subunit_id=str(plan.get("scope_subunit_id") or ""),
            extra_user_ids=khth_notice_user_ids,
        ),
        _work_position_all_changed_payload(
            "work_position_adjustment_created",
            actor_user_id=str(user.id),
            adjustment_id=str(adjustment_id),
            plan_id=str(original_plan_id),
            target_date=str(plan.get("target_date") or work_date or ""),
            target_week=_week_value_from_date_str(str(plan.get("target_date") or work_date or "")),
        ),
    )

    if khth_notice_user_ids:
        _emit_work_position_realtime_sync(
            khth_notice_user_ids,
            _position_notify_payload(
                "work_position_notice_created",
                notice_type="KHTH_ADJUSTMENT_PENDING",
                adjustment_id=adjustment_id,
                plan_id=original_plan_id,
                target_date=str(plan.get("target_date") or work_date or ""),
                target_week=_week_value_from_date_str(str(plan.get("target_date") or work_date or "")),
            ),
        )

    return RedirectResponse(
        url=(
            f"/work-positions?main_tab=positions"
            f"&block_tab={scope_info.get('block_code') or 'CHUYEN_MON'}"
            f"&sub_tab=approve"
            f"&target_week={_week_value_from_date_str(str(plan.get('target_date') or work_date or ''))}"
            f"&scope_unit_id={scope_info.get('selected_scope_unit_id') or ''}"
            f"&msg=Đã lập phiếu điều chỉnh và chuyển cấp Bệnh viện phê duyệt."
        ),
        status_code=302,
    )

@router.post("/adjustments/approve-khth")
def approve_work_position_adjustment_khth(
    request: Request,
    adjustment_id: str = Form(...),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)
    scope_info = _scope_payload(db, user)

    if not scope_info.get("is_khth_manager"):
        raise HTTPException(status_code=403, detail="Bạn không có quyền phê duyệt điều chỉnh cấp Bệnh viện.")

    row = db.execute(
        text("SELECT * FROM work_position_adjustments WHERE id = :id"),
        {"id": adjustment_id},
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Không tìm thấy phiếu điều chỉnh.")

    adjustment = _row_to_dict(row)
    if str(adjustment.get("status") or "").upper() != "UNIT_APPROVED":
        raise HTTPException(status_code=400, detail="Phiếu điều chỉnh không ở trạng thái chờ phê duyệt cấp Bệnh viện.")

    now_str = _dt_to_str(_now_vn())
    approver_name = user.full_name or user.username or ""

    db.execute(
        text(
            """
            UPDATE work_position_adjustments
            SET status = 'KHTH_APPROVED',
                khth_approved_by_id = :khth_approved_by_id,
                khth_approved_by_name = :khth_approved_by_name,
                khth_approved_at = :khth_approved_at,
                updated_at = :updated_at
            WHERE id = :id
            """
        ),
        {
            "id": adjustment_id,
            "khth_approved_by_id": user.id,
            "khth_approved_by_name": approver_name,
            "khth_approved_at": now_str,
            "updated_at": now_str,
        },
    )

    notice_user_ids = _create_adjustment_notices_for_approved_adjustment(
        db,
        adjustment_id=adjustment_id,
        approved_by_name=approver_name,
    )

    db.execute(
        text(
            """
            UPDATE work_position_notices
            SET is_read = 1,
                approved_at = :approved_at
            WHERE notice_type = 'KHTH_ADJUSTMENT_PENDING'
              AND related_plan_id = :related_plan_id
              AND is_read = 0
            """
        ),
        {
            "related_plan_id": str(adjustment.get("original_plan_id") or ""),
            "approved_at": now_str,
        },
    )

    db.commit()

    _emit_work_position_realtime_sync(
        _work_position_scope_realtime_user_ids(
            db,
            scope_unit_id=str(adjustment.get("scope_unit_id") or ""),
            scope_subunit_id=str(adjustment.get("scope_subunit_id") or ""),
            include_hospital_viewers=True,
            extra_user_ids=notice_user_ids,
        ),
        _work_position_all_changed_payload(
            "work_position_adjustment_khth_approved",
            actor_user_id=str(user.id),
            adjustment_id=str(adjustment_id),
            target_date=str(adjustment.get("target_date") or ""),
            target_week=_week_value_from_date_str(str(adjustment.get("target_date") or "")),
        ),
    )

    if notice_user_ids:
        _emit_work_position_realtime_sync(
            notice_user_ids,
            _position_notify_payload(
                "work_position_notice_created",
                notice_type="ADJUSTMENT_APPROVED",
                adjustment_id=adjustment_id,
                target_date=str(adjustment.get("target_date") or ""),
                target_week=_week_value_from_date_str(str(adjustment.get("target_date") or "")),
            ),
        )

    return RedirectResponse(
        url=(
            f"/work-positions?main_tab=positions"
            f"&block_tab=CHUYEN_MON"
            f"&sub_tab=khth_summary"
            f"&target_week={_week_value_from_date_str(str(adjustment.get('target_date') or ''))}"
            f"&msg=Đã phê duyệt điều chỉnh cấp Bệnh viện."
        ),
        status_code=302,
    )

    
@router.post("/plans/approve-khth")
def approve_khth_position_plan(
    request: Request,
    plan_id: str = Form(...),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)
    scope_info = _scope_payload(db, user)

    if not scope_info.get("is_khth_manager"):
        raise HTTPException(status_code=403, detail="Bạn không có quyền phê duyệt cấp Bệnh viện.")

    plan_row = db.execute(
        text("SELECT * FROM work_position_plans WHERE id = :id"),
        {"id": plan_id},
    ).fetchone()
    if not plan_row:
        raise HTTPException(status_code=404, detail="Không tìm thấy bản phân vị trí.")

    plan = _row_to_dict(plan_row)
    plan_items = _load_plan_item_rows(db, plan_id)
    if not _plan_needs_khth_approval(str(plan.get("block_code") or ""), plan_items):
        raise HTTPException(status_code=400, detail="Bản phân vị trí này không thuộc luồng phê duyệt cấp Bệnh viện.")

    current_status = str(plan.get("status") or "").strip().upper()
    if current_status != "UNIT_APPROVED":
        raise HTTPException(
            status_code=400,
            detail="Chỉ được phê duyệt cấp Bệnh viện sau khi bản phân vị trí đã được phê duyệt cấp đơn vị.",
        )

    now_str = _dt_to_str(_now_vn())
    approver_name = user.full_name or user.username or ""

    db.execute(
        text(
            """
            UPDATE work_position_plans
            SET status = 'KHTH_APPROVED',
                khth_approved_by_id = :khth_approved_by_id,
                khth_approved_by_name = :khth_approved_by_name,
                khth_approved_at = :khth_approved_at,
                approved_by_id = :approved_by_id,
                approved_by_name = :approved_by_name,
                approved_at = :approved_at,
                updated_at = :updated_at
            WHERE id = :id
            """
        ),
        {
            "id": plan_id,
            "khth_approved_by_id": user.id,
            "khth_approved_by_name": approver_name,
            "khth_approved_at": now_str,
            "approved_by_id": user.id,
            "approved_by_name": approver_name,
            "approved_at": now_str,
            "updated_at": now_str,
        },
    )

    notice_user_ids = _create_position_notices_for_plan(
        db,
        plan_id=plan_id,
        approved_by_name=approver_name,
    )

    db.execute(
        text(
            """
            UPDATE work_position_notices
            SET is_read = 1,
                approved_at = :approved_at
            WHERE notice_type = 'KHTH_PLAN_PENDING'
              AND related_plan_id = :plan_id
              AND is_read = 0
            """
        ),
        {
            "plan_id": str(plan_id),
            "approved_at": now_str,
        },
    )

    db.commit()

    _emit_work_position_realtime_sync(
        _work_position_scope_realtime_user_ids(
            db,
            scope_unit_id=str(plan.get("scope_unit_id") or ""),
            scope_subunit_id=str(plan.get("scope_subunit_id") or ""),
            include_hospital_viewers=True,
            extra_user_ids=notice_user_ids,
        ),
        _work_position_all_changed_payload(
            "work_position_plan_khth_approved",
            actor_user_id=str(user.id),
            plan_id=str(plan_id),
            target_date=str(plan.get("target_date") or ""),
            target_week=_week_value_from_date_str(str(plan.get("target_date") or "")),
        ),
    )

    if notice_user_ids:
        _emit_work_position_realtime_sync(
            notice_user_ids,
            _position_notify_payload(
                "work_position_notice_created",
                notice_type="PLAN_APPROVED",
                plan_id=plan_id,
                target_date=str(plan.get("target_date") or ""),
                target_week=_week_value_from_date_str(str(plan.get("target_date") or "")),
            ),
        )

    return RedirectResponse(
        url=(
            f"/work-positions?main_tab=positions"
            f"&block_tab=CHUYEN_MON"
            f"&sub_tab=khth_summary"
            f"&target_week={_week_value_from_date_str(str(plan.get('target_date') or ''))}"
            f"&msg=Đã phê duyệt cấp Bệnh viện và phát thông báo vị trí."
        ),
        status_code=302,
    )
    
    
@router.post("/notices/read")
def mark_work_position_notice_read(
    request: Request,
    notice_id: str = Form(...),
    scope_unit_id: str = Form(""),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    db.execute(
        text(
            """
            UPDATE work_position_notices
            SET is_read = 1
            WHERE id = :id AND user_id = :user_id
            """
        ),
        {"id": notice_id, "user_id": str(user.id)},
    )
    db.commit()

    return RedirectResponse(
        url=(
            f"/work-positions?main_tab=positions"
            f"&sub_tab=notice"
            f"&scope_unit_id={scope_unit_id}"
            f"&msg=Đã đánh dấu đã đọc."
        ),
        status_code=302,
    )

@router.post("/notices/read-all")
def mark_all_work_position_notices_read(
    request: Request,
    scope_unit_id: str = Form(""),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    db.execute(
        text(
            """
            UPDATE work_position_notices
            SET is_read = 1
            WHERE user_id = :user_id AND is_read = 0
            """
        ),
        {"user_id": str(user.id)},
    )
    db.commit()

    return RedirectResponse(
        url=(
            f"/work-positions?main_tab=positions"
            f"&sub_tab=notice"
            f"&scope_unit_id={scope_unit_id}"
            f"&msg=Đã đánh dấu đọc toàn bộ."
        ),
        status_code=302,
    )


@router.post("/catalog/toggle")
def toggle_work_position_catalog(
    request: Request,
    catalog_id: str = Form(...),
    scope_unit_id: str = Form(""),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)
    scope_info = _scope_payload(db, user, scope_unit_id)

    if not _can_manage_catalog(db, user):
        raise HTTPException(status_code=403, detail="Bạn không có quyền cập nhật danh mục vị trí làm việc.")

    row = db.execute(
        text("SELECT * FROM work_position_catalog WHERE id = :id"),
        {"id": catalog_id},
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Không tìm thấy danh mục vị trí làm việc.")

    item = _row_to_dict(row)
    scope_unit_id_val = str(scope_info.get("scope_unit_id") or "").strip()
    scope_subunit_id_val = str(scope_info.get("scope_subunit_id") or "").strip()
    allowed_scope_ids = {x for x in {scope_unit_id_val, scope_subunit_id_val} if x}
    item_scope_ids = {
        str(item.get("owner_unit_id") or "").strip(),
        str(item.get("apply_unit_id") or "").strip(),
    }

    if not allowed_scope_ids.intersection(item_scope_ids):
        raise HTTPException(status_code=403, detail="Danh mục vị trí không thuộc phạm vi đang thao tác.")

    new_active = 0 if int(item.get("is_active") or 0) == 1 else 1

    db.execute(
        text(
            """
            UPDATE work_position_catalog
            SET is_active = :is_active,
                updated_at = :updated_at
            WHERE id = :id
            """
        ),
        {
            "id": catalog_id,
            "is_active": new_active,
            "updated_at": _dt_to_str(_now_vn()),
        },
    )
    db.commit()

    _emit_work_position_realtime_sync(
        _work_position_scope_realtime_user_ids(
            db,
            scope_unit_id=str(item.get("owner_unit_id") or item.get("apply_unit_id") or ""),
            scope_subunit_id=str(item.get("apply_unit_id") or ""),
        ),
        _work_position_all_changed_payload(
            "work_position_catalog_changed",
            actor_user_id=str(user.id),
            catalog_id=str(catalog_id),
            is_active=int(new_active),
        ),
    )

    msg = "Đã ngưng sử dụng vị trí làm việc." if new_active == 0 else "Đã mở lại vị trí làm việc."
    return RedirectResponse(
        url=(
            f"/work-positions?main_tab=positions"
            f"&block_tab={scope_info.get('block_code') or 'CHUYEN_MON'}"
            f"&sub_tab=catalog"
            f"&scope_unit_id={scope_info.get('selected_scope_unit_id') or ''}"
            f"&msg={msg}"
        ),
        status_code=302,
    )
